import pandas as pd
from flask import (
    Flask, 
    request, 
    send_file, 
    jsonify, 
    render_template_string, 
    render_template,
    redirect,
    url_for,
    session
)
import io
import time
import requests
from geopy.geocoders import Nominatim
import math
import threading
import concurrent.futures
from diskcache import Cache
import logging
import os
import joblib
import polyline
from polyline import decode as decode_polyline
import json
from functools import wraps
import re
import unicodedata
from rapidfuzz import process, fuzz
import csv
from ptv_api_manager import PTVRouteManager
import openpyxl
import hashlib
import secrets
import atexit
from datetime import timedelta
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

# Import nowych modułów do zarządzania sesjami
from session_manager import SessionManager, SessionCleanupScheduler
from user_session_data import UserSessionData

# Blokada dla bezpiecznej aktualizacji zmiennych globalnych (używana przez starszy kod)
# TODO: Stopniowo usunąć po pełnej migracji do SessionManager
progress_lock = threading.Lock()

# Globalny słownik lookup
LOOKUP_DICT = {}

# Nowa zmienna globalna - słownik mapujący (kod kraju, kod pocztowy) -> region
REGION_MAPPING = {}

# Słownik kodów ISO krajów
ISO_CODES = {
    "Poland": "pl",
    "Germany": "de",
    "France": "fr",
    "Italy": "it",
    "Spain": "es",
    "Netherlands": "nl",
    "Belgium": "be",
    "Czech Republic": "cz",
    "Austria": "at",
    "Slovakia": "sk",
    "Slovenia": "si",
    "Hungary": "hu",
    "Portugal": "pt",
    "Greece": "gr",
    "Switzerland": "ch",
    "Sweden": "se",
    "Finland": "fi",
    "Norway": "no",
    "Denmark": "dk",
    "Luxembourg": "lu"
}

# Globalne zmienne
PTV_API_KEY = "RVVfZmQ1YTcyY2E4ZjNiNDhmOTlhYjE5NjRmNGZhYTdlNTc6NGUyM2VhMmEtZTc2YS00YmVkLWIyMTMtZDc2YjE0NWZjZjE1"
DEFAULT_ROUTING_MODE = "FAST"  # Stały tryb wyznaczania trasy
DEFAULT_FUEL_COST = 0.40  # Domyślny koszt paliwa EUR/km
DEFAULT_DRIVER_COST = 210  # Domyślny koszt kierowcy EUR/dzień

# Konfiguracja loggera - zmiana poziomu na ERROR aby ograniczyć logi
logging.basicConfig(
    level=logging.ERROR,  # Zmieniono z INFO na ERROR
    format='%(message)s',  # Uproszczony format bez timestampów
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Wyłączenie debugowych logów urllib3
logging.getLogger('urllib3').setLevel(logging.ERROR)


# ============================================================================
# DATACLASSES - Modele danych dla tras z punktami pośrednimi
# ============================================================================

@dataclass
class WaypointData:
    """
    Reprezentacja pojedynczego punktu na trasie.
    
    Attributes:
        country: Kod kraju (ISO 2-letter, np. 'PL')
        postal_code: Kod pocztowy
        city: Nazwa miasta (opcjonalne)
        coordinates: Współrzędne (lat, lon) po geokodowaniu
    """
    country: str
    postal_code: str
    city: Optional[str] = None
    coordinates: Optional[Tuple[float, float]] = None
    
    def __post_init__(self):
        """Normalizacja danych wejściowych"""
        if self.country:
            self.country = self.country.upper().strip()
        if self.postal_code:
            self.postal_code = self.postal_code.strip()
        if self.city:
            self.city = self.city.strip()
    
    def is_geocoded(self) -> bool:
        """Sprawdza czy punkt ma przypisane współrzędne"""
        return self.coordinates is not None and len(self.coordinates) == 2
    
    def __str__(self) -> str:
        """String representation dla logowania"""
        city_str = f" ({self.city})" if self.city else ""
        return f"{self.country} {self.postal_code}{city_str}"


@dataclass
class RouteRequest:
    """
    Żądanie obliczenia trasy z punktami pośrednimi.
    
    Attributes:
        start: Punkt startowy
        end: Punkt końcowy
        waypoints: Lista punktów pośrednich (0-5)
        fuel_cost: Koszt paliwa [EUR/km]
        driver_cost: Koszt kierowcy [EUR/dzień]
        matrix_type: Typ matrycy marży ('klient' lub 'targi')
        avoid_switzerland: Czy unikać Szwajcarii
        routing_mode: Tryb routingu ('FAST', 'ECO', 'SHORT')
    """
    start: WaypointData
    end: WaypointData
    waypoints: List[WaypointData] = field(default_factory=list)
    fuel_cost: float = DEFAULT_FUEL_COST
    driver_cost: float = DEFAULT_DRIVER_COST
    matrix_type: str = "klient"
    avoid_switzerland: bool = True
    routing_mode: str = DEFAULT_ROUTING_MODE
    
    def __post_init__(self):
        """Walidacja danych"""
        if len(self.waypoints) > 5:
            raise ValueError("Maksymalnie 5 punktów pośrednich dozwolone")
        
        if self.fuel_cost < 0 or self.fuel_cost > 5:
            raise ValueError("Koszt paliwa musi być w zakresie 0-5 EUR/km")
        
        if self.driver_cost < 0 or self.driver_cost > 1000:
            raise ValueError("Koszt kierowcy musi być w zakresie 0-1000 EUR/dzień")
        
        if self.matrix_type not in ['klient', 'targi']:
            raise ValueError("matrix_type musi być 'klient' lub 'targi'")
    
    def get_all_points_ordered(self) -> List[WaypointData]:
        """Zwraca wszystkie punkty w kolejności: start → waypoints → end"""
        return [self.start] + self.waypoints + [self.end]
    
    def total_waypoints_count(self) -> int:
        """Liczba wszystkich punktów włączając start i end"""
        return 2 + len(self.waypoints)
    
    def has_waypoints(self) -> bool:
        """Czy trasa zawiera punkty pośrednie"""
        return len(self.waypoints) > 0
    
    def __str__(self) -> str:
        """String representation dla logowania"""
        points_str = " → ".join(str(p) for p in self.get_all_points_ordered())
        return f"Route[{self.total_waypoints_count()} points]: {points_str}"


# ============================================================================
# FLASK & LOGGING CONFIGURATION
# ============================================================================

# Filtr dla logów postępu
class FilterProgress(logging.Filter):
    def filter(self, record):
        return not record.getMessage().startswith('Progress:')

logger.addFilter(FilterProgress())

# Ustawienie poziomu ERROR dla logów związanych z punktami trasy
route_logger = logging.getLogger(__name__ + '.route')
route_logger.setLevel(logging.ERROR)
route_logger.handlers = []  # Usuwamy domyślne handlery
route_logger.addHandler(logging.StreamHandler())
route_logger.propagate = False  # Zapobiegamy propagacji logów do rodzica

# Inicjalizacja geolokatora
geolocator = Nominatim(user_agent="wycena_transportu", timeout=15)

# Inicjalizacja pamięci podręcznych
geo_cache = Cache("geo_cache")
route_cache = Cache("route_cache")
locations_cache = Cache("locations_cache")

# Inicjalizacja menedżera PTV
ptv_manager = PTVRouteManager(PTV_API_KEY)

# === ZMIENNE GLOBALNE - TYLKO DLA LEGACY MODE DEKORATORA ===
GEOCODING_TOTAL = 0
GEOCODING_CURRENT = 0

# Inicjalizacja SessionManager - zarządzanie sesjami użytkowników
session_manager = SessionManager(max_age_hours=24)

# Mapowanie krajów – ujednolicone nazwy
COUNTRY_MAPPING = {
    'PL': 'Poland', 'Polska': 'Poland',
    'DE': 'Germany', 'Niemcy': 'Germany',
    'FR': 'France', 'Francja': 'France',
    'IT': 'Italy', 'Włochy': 'Italy',
    'ES': 'Spain', 'Hiszpania': 'Spain',
    'NL': 'Netherlands', 'Holandia': 'Netherlands',
    'BE': 'Belgium', 'Belgia': 'Belgium',
    'CZ': 'Czech Republic', 'Czechy': 'Czech Republic',
    'AT': 'Austria', 'Austria': 'Austria',
    'SK': 'Slovakia', 'Słowacja': 'Slovakia',
    'SI': 'Slovenia', 'Słowenia': 'Slovenia',
    'HU': 'Hungary', 'Węgry': 'Hungary',
    'PT': 'Portugal', 'Portugalia': 'Portugal',
    'GR': 'Greece', 'Grecja': 'Greece',
    'CH': 'Switzerland', 'Szwajcaria': 'Switzerland',
    'SE': 'Sweden', 'Szwecja': 'Sweden',
    'FI': 'Finland', 'Finlandia': 'Finland',
    'NO': 'Norway', 'Norwegia': 'Norway',
    'DK': 'Denmark', 'Dania': 'Denmark',
    'BG': 'Bulgaria', 'Bułgaria': 'Bulgaria',
    'EE': 'Estonia', 'Estonia': 'Estonia',
    'HR': 'Croatia', 'Chorwacja': 'Croatia',
    'IE': 'Ireland', 'Irlandia': 'Ireland',  # Użyto "Irlandia" jako poprawnej wersji
    'LT': 'Lithuania', 'Litwa': 'Lithuania',
    'LV': 'Latvia', 'Łotwa': 'Latvia',
    'RO': 'Romania', 'Rumunia': 'Romania',
    'UK': 'United Kingdom', 'Wielka Brytania': 'United Kingdom',
    'LU': 'Luxembourg', 'Luksemburg': 'Luxembourg'

}

# Definicja wyjątku
class GeocodeException(Exception):
    """Wyjątek sygnalizujący potrzebę ręcznego geokodowania."""

    def __init__(self, ungeocoded_locations):
        self.ungeocoded_locations = ungeocoded_locations
        super().__init__("Znaleziono nierozpoznane lokalizacje")


class LocationVerificationRequired(Exception):
    """Wyjątek sygnalizujący potrzebę weryfikacji lokalizacji przez użytkownika."""

    def __init__(self, locations_to_verify):
        self.locations_to_verify = locations_to_verify
        super().__init__("Znaleziono lokalizacje wymagające weryfikacji")


def normalize_country(country):
    return COUNTRY_MAPPING.get(str(country).strip(), str(country).strip())


# Funkcja wczytująca mapowania regionów
def load_region_mapping():
    global REGION_MAPPING

    region_data = """ 
    ('10', 'AT', 'AT WSCHÓD'),
('11', 'AT', 'AT WSCHÓD'),
('12', 'AT', 'AT WSCHÓD'),
('13', 'AT', 'AT WSCHÓD'),
('14', 'AT', 'AT WSCHÓD'),
('15', 'AT', 'AT WSCHÓD'),
('16', 'AT', 'AT WSCHÓD'),
('17', 'AT', 'AT WSCHÓD'),
('18', 'AT', 'AT WSCHÓD'),
('19', 'AT', 'AT WSCHÓD'),
('20', 'AT', 'AT WSCHÓD'),
('21', 'AT', 'AT WSCHÓD'),
('22', 'AT', 'AT WSCHÓD'),
('23', 'AT', 'AT WSCHÓD'),
('24', 'AT', 'AT WSCHÓD'),
('25', 'AT', 'AT WSCHÓD'),
('26', 'AT', 'AT WSCHÓD'),
('27', 'AT', 'AT WSCHÓD'),
('28', 'AT', 'AT WSCHÓD'),
('29', 'AT', 'AT WSCHÓD'),
('30', 'AT', 'AT WSCHÓD'),
('31', 'AT', 'AT WSCHÓD'),
('32', 'AT', 'AT WSCHÓD'),
('33', 'AT', 'AT WSCHÓD'),
('34', 'AT', 'AT WSCHÓD'),
('35', 'AT', 'AT WSCHÓD'),
('36', 'AT', 'AT WSCHÓD'),
('37', 'AT', 'AT WSCHÓD'),
('38', 'AT', 'AT WSCHÓD'),
('39', 'AT', 'AT WSCHÓD'),
('40', 'AT', 'AT WSCHÓD'),
('41', 'AT', 'AT WSCHÓD'),
('42', 'AT', 'AT WSCHÓD'),
('43', 'AT', 'AT WSCHÓD'),
('44', 'AT', 'AT WSCHÓD'),
('45', 'AT', 'AT WSCHÓD'),
('46', 'AT', 'AT WSCHÓD'),
('47', 'AT', 'AT WSCHÓD'),
('48', 'AT', 'AT WSCHÓD'),
('49', 'AT', 'AT WSCHÓD'),
('50', 'AT', 'AT WSCHÓD'),
('51', 'AT', 'AT WSCHÓD'),
('52', 'AT', 'AT WSCHÓD'),
('53', 'AT', 'AT WSCHÓD'),
('54', 'AT', 'AT WSCHÓD'),
('55', 'AT', 'AT WSCHÓD'),
('56', 'AT', 'AT WSCHÓD'),
('57', 'AT', 'AT WSCHÓD'),
('58', 'AT', 'AT WSCHÓD'),
('59', 'AT', 'AT WSCHÓD'),
('60', 'AT', 'AT ZACHÓD'),
('61', 'AT', 'AT ZACHÓD'),
('62', 'AT', 'AT ZACHÓD'),
('63', 'AT', 'AT ZACHÓD'),
('64', 'AT', 'AT ZACHÓD'),
('65', 'AT', 'AT ZACHÓD'),
('66', 'AT', 'AT ZACHÓD'),
('67', 'AT', 'AT ZACHÓD'),
('68', 'AT', 'AT ZACHÓD'),
('69', 'AT', 'AT ZACHÓD'),
('70', 'AT', 'AT WSCHÓD'),
('71', 'AT', 'AT WSCHÓD'),
('72', 'AT', 'AT WSCHÓD'),
('73', 'AT', 'AT WSCHÓD'),
('74', 'AT', 'AT WSCHÓD'),
('75', 'AT', 'AT WSCHÓD'),
('76', 'AT', 'AT WSCHÓD'),
('77', 'AT', 'AT WSCHÓD'),
('78', 'AT', 'AT WSCHÓD'),
('79', 'AT', 'AT WSCHÓD'),
('80', 'AT', 'AT WSCHÓD'),
('81', 'AT', 'AT WSCHÓD'),
('82', 'AT', 'AT WSCHÓD'),
('83', 'AT', 'AT WSCHÓD'),
('84', 'AT', 'AT WSCHÓD'),
('85', 'AT', 'AT WSCHÓD'),
('86', 'AT', 'AT WSCHÓD'),
('87', 'AT', 'AT WSCHÓD'),
('88', 'AT', 'AT WSCHÓD'),
('89', 'AT', 'AT WSCHÓD'),
('90', 'AT', 'AT WSCHÓD'),
('91', 'AT', 'AT WSCHÓD'),
('92', 'AT', 'AT WSCHÓD'),
('93', 'AT', 'AT WSCHÓD'),
('94', 'AT', 'AT WSCHÓD'),
('95', 'AT', 'AT WSCHÓD'),
('96', 'AT', 'AT WSCHÓD'),
('97', 'AT', 'AT WSCHÓD'),
('98', 'AT', 'AT WSCHÓD'),
('99', 'AT', 'AT WSCHÓD'),
('', 'BE', 'BE'),
('', 'BG', 'BG'),
('', 'CH', 'CH'),
('10', 'CZ', 'CZ PÓŁNOC'),
('11', 'CZ', 'CZ PÓŁNOC'),
('12', 'CZ', 'CZ PÓŁNOC'),
('13', 'CZ', 'CZ PÓŁNOC'),
('14', 'CZ', 'CZ PÓŁNOC'),
('15', 'CZ', 'CZ PÓŁNOC'),
('16', 'CZ', 'CZ PÓŁNOC'),
('17', 'CZ', 'CZ PÓŁNOC'),
('18', 'CZ', 'CZ PÓŁNOC'),
('19', 'CZ', 'CZ PÓŁNOC'),
('20', 'CZ', 'CZ PÓŁNOC'),
('21', 'CZ', 'CZ PÓŁNOC'),
('22', 'CZ', 'CZ PÓŁNOC'),
('23', 'CZ', 'CZ PÓŁNOC'),
('24', 'CZ', 'CZ PÓŁNOC'),
('25', 'CZ', 'CZ PÓŁNOC'),
('26', 'CZ', 'CZ PÓŁNOC'),
('27', 'CZ', 'CZ PÓŁNOC'),
('28', 'CZ', 'CZ PÓŁNOC'),
('29', 'CZ', 'CZ PÓŁNOC'),
('30', 'CZ', 'CZ PÓŁNOC'),
('31', 'CZ', 'CZ PÓŁNOC'),
('32', 'CZ', 'CZ PÓŁNOC'),
('33', 'CZ', 'CZ PÓŁNOC'),
('34', 'CZ', 'CZ PÓŁNOC'),
('35', 'CZ', 'CZ PÓŁNOC'),
('36', 'CZ', 'CZ PÓŁNOC'),
('37', 'CZ', 'CZ POŁUDNIE'),
('38', 'CZ', 'CZ POŁUDNIE'),
('39', 'CZ', 'CZ PÓŁNOC'),
('40', 'CZ', 'CZ PÓŁNOC'),
('41', 'CZ', 'CZ PÓŁNOC'),
('42', 'CZ', 'CZ PÓŁNOC'),
('43', 'CZ', 'CZ PÓŁNOC'),
('44', 'CZ', 'CZ PÓŁNOC'),
('45', 'CZ', 'CZ PÓŁNOC'),
('46', 'CZ', 'CZ PÓŁNOC'),
('47', 'CZ', 'CZ PÓŁNOC'),
('48', 'CZ', 'CZ PÓŁNOC'),
('49', 'CZ', 'CZ PÓŁNOC'),
('50', 'CZ', 'CZ PÓŁNOC'),
('51', 'CZ', 'CZ PÓŁNOC'),
('52', 'CZ', 'CZ PÓŁNOC'),
('53', 'CZ', 'CZ PÓŁNOC'),
('54', 'CZ', 'CZ PÓŁNOC'),
('55', 'CZ', 'CZ PÓŁNOC'),
('56', 'CZ', 'CZ PÓŁNOC'),
('57', 'CZ', 'CZ PÓŁNOC'),
('58', 'CZ', 'CZ PÓŁNOC'),
('59', 'CZ', 'CZ PÓŁNOC'),
('60', 'CZ', 'CZ PÓŁNOC'),
('61', 'CZ', 'CZ PÓŁNOC'),
('62', 'CZ', 'CZ PÓŁNOC'),
('63', 'CZ', 'CZ PÓŁNOC'),
('64', 'CZ', 'CZ PÓŁNOC'),
('65', 'CZ', 'CZ PÓŁNOC'),
('66', 'CZ', 'CZ POŁUDNIE'),
('67', 'CZ', 'CZ POŁUDNIE'),
('68', 'CZ', 'CZ POŁUDNIE'),
('69', 'CZ', 'CZ POŁUDNIE'),
('70', 'CZ', 'CZ PÓŁNOC'),
('71', 'CZ', 'CZ PÓŁNOC'),
('72', 'CZ', 'CZ PÓŁNOC'),
('73', 'CZ', 'CZ PÓŁNOC'),
('74', 'CZ', 'CZ PÓŁNOC'),
('75', 'CZ', 'CZ PÓŁNOC'),
('76', 'CZ', 'CZ POŁUDNIE'),
('77', 'CZ', 'CZ POŁUDNIE'),
('78', 'CZ', 'CZ POŁUDNIE'),
('79', 'CZ', 'CZ POŁUDNIE'),
('01', 'DE', 'DE0'),
('02', 'DE', 'DE0'),
('03', 'DE', 'DE0'),
('04', 'DE', 'DE0'),
('05', 'DE', 'DE0'),
('06', 'DE', 'DE0'),
('07', 'DE', 'DE0'),
('08', 'DE', 'DE0'),
('09', 'DE', 'DE0'),
('10', 'DE', 'DE1 POŁUDNIE'),
('11', 'DE', 'DE1 POŁUDNIE'),
('12', 'DE', 'DE1 POŁUDNIE'),
('13', 'DE', 'DE1 POŁUDNIE'),
('14', 'DE', 'DE1 POŁUDNIE'),
('15', 'DE', 'DE1 POŁUDNIE'),
('16', 'DE', 'DE1 POŁUDNIE'),
('17', 'DE', 'DE1 PÓŁNOC'),
('18', 'DE', 'DE1 PÓŁNOC'),
('19', 'DE', 'DE1 PÓŁNOC'),
('20', 'DE', 'DE2 POŁUDNIE'),
('21', 'DE', 'DE2 POŁUDNIE'),
('22', 'DE', 'DE2 POŁUDNIE'),
('23', 'DE', 'DE2 PÓŁNOC'),
('24', 'DE', 'DE2 PÓŁNOC'),
('25', 'DE', 'DE2 PÓŁNOC'),
('26', 'DE', 'DE2 POŁUDNIE'),
('27', 'DE', 'DE2 POŁUDNIE'),
('28', 'DE', 'DE2 POŁUDNIE'),
('29', 'DE', 'DE2 POŁUDNIE'),
('30', 'DE', 'DE3'),
('31', 'DE', 'DE3'),
('32', 'DE', 'DE3'),
('33', 'DE', 'DE3'),
('34', 'DE', 'DE3'),
('35', 'DE', 'DE3'),
('36', 'DE', 'DE3'),
('37', 'DE', 'DE3'),
('38', 'DE', 'DE3'),
('39', 'DE', 'DE3'),
('40', 'DE', 'DE4'),
('41', 'DE', 'DE4'),
('42', 'DE', 'DE4'),
('43', 'DE', 'DE4'),
('44', 'DE', 'DE4'),
('45', 'DE', 'DE4'),
('46', 'DE', 'DE4'),
('47', 'DE', 'DE4'),
('48', 'DE', 'DE4'),
('49', 'DE', 'DE4'),
('50', 'DE', 'DE5'),
('51', 'DE', 'DE5'),
('52', 'DE', 'DE5'),
('53', 'DE', 'DE5'),
('54', 'DE', 'DE5'),
('55', 'DE', 'DE5'),
('56', 'DE', 'DE5'),
('57', 'DE', 'DE5'),
('58', 'DE', 'DE5'),
('59', 'DE', 'DE5'),
('60', 'DE', 'DE6'),
('61', 'DE', 'DE6'),
('62', 'DE', 'DE6'),
('63', 'DE', 'DE6'),
('64', 'DE', 'DE6'),
('65', 'DE', 'DE6'),
('66', 'DE', 'DE6'),
('67', 'DE', 'DE6'),
('68', 'DE', 'DE6'),
('69', 'DE', 'DE6'),
('70', 'DE', 'DE7 PÓŁNOC'),
('71', 'DE', 'DE7 PÓŁNOC'),
('72', 'DE', 'DE7 PÓŁNOC'),
('73', 'DE', 'DE7 PÓŁNOC'),
('74', 'DE', 'DE7 PÓŁNOC'),
('75', 'DE', 'DE7 PÓŁNOC'),
('76', 'DE', 'DE7 PÓŁNOC'),
('77', 'DE', 'DE7 POŁUDNIE'),
('78', 'DE', 'DE7 POŁUDNIE'),
('79', 'DE', 'DE7 POŁUDNIE'),
('80', 'DE', 'DE8 PÓŁNOC'),
('81', 'DE', 'DE8 PÓŁNOC'),
('82', 'DE', 'DE8 PÓŁNOC'),
('83', 'DE', 'DE8 POŁUDNIE'),
('84', 'DE', 'DE8 PÓŁNOC'),
('85', 'DE', 'DE8 PÓŁNOC'),
('86', 'DE', 'DE8 PÓŁNOC'),
('87', 'DE', 'DE8 POŁUDNIE'),
('88', 'DE', 'DE8 POŁUDNIE'),
('89', 'DE', 'DE8 PÓŁNOC'),
('90', 'DE', 'DE9 PÓŁNOC'),
('91', 'DE', 'DE9 PÓŁNOC'),
('92', 'DE', 'DE9 PÓŁNOC'),
('93', 'DE', 'DE9 POŁUDNIE'),
('94', 'DE', 'DE9 POŁUDNIE'),
('95', 'DE', 'DE9 PÓŁNOC'),
('96', 'DE', 'DE9 PÓŁNOC'),
('97', 'DE', 'DE9 PÓŁNOC'),
('98', 'DE', 'DE9 PÓŁNOC'),
('99', 'DE', 'DE9 PÓŁNOC'),
('', 'DK', 'DK'),
('', 'EE', 'EE'),
('01', 'ES', 'ES PÓŁNOC'),
('02', 'ES', 'ES CENTRUM'),
('03', 'ES', 'ES POŁUDNIE'),
('04', 'ES', 'ES POŁUDNIE'),
('05', 'ES', 'ES CENTRUM'),
('06', 'ES', 'ES POŁUDNIE'),
('08', 'ES', 'ES PÓŁNOC'),
('09', 'ES', 'ES PÓŁNOC'),
('10', 'ES', 'ES POŁUDNIE'),
('11', 'ES', 'ES POŁUDNIE'),
('12', 'ES', 'ES CENTRUM'),
('13', 'ES', 'ES CENTRUM'),
('14', 'ES', 'ES POŁUDNIE'),
('15', 'ES', 'ES POŁUDNIE'),
('16', 'ES', 'ES CENTRUM'),
('17', 'ES', 'ES PÓŁNOC'),
('18', 'ES', 'ES POŁUDNIE'),
('19', 'ES', 'ES CENTRUM'),
('20', 'ES', 'ES PÓŁNOC'),
('21', 'ES', 'ES POŁUDNIE'),
('22', 'ES', 'ES PÓŁNOC'),
('23', 'ES', 'ES POŁUDNIE'),
('24', 'ES', 'ES PÓŁNOC'),
('25', 'ES', 'ES PÓŁNOC'),
('26', 'ES', 'ES PÓŁNOC'),
('27', 'ES', 'ES POŁUDNIE'),
('28', 'ES', 'ES CENTRUM'),
('29', 'ES', 'ES POŁUDNIE'),
('30', 'ES', 'ES POŁUDNIE'),
('31', 'ES', 'ES PÓŁNOC'),
('32', 'ES', 'ES POŁUDNIE'),
('33', 'ES', 'ES PÓŁNOC'),
('34', 'ES', 'ES PÓŁNOC'),
('35', 'ES', 'ES POŁUDNIE'),
('36', 'ES', 'ES POŁUDNIE'),
('37', 'ES', 'ES POŁUDNIE'),
('38', 'ES', 'ES POŁUDNIE'),
('39', 'ES', 'ES PÓŁNOC'),
('40', 'ES', 'ES CENTRUM'),
('41', 'ES', 'ES POŁUDNIE'),
('42', 'ES', 'ES PÓŁNOC'),
('43', 'ES', 'ES PÓŁNOC'),
('44', 'ES', 'ES CENTRUM'),
('45', 'ES', 'ES CENTRUM'),
('46', 'ES', 'ES CENTRUM'),
('47', 'ES', 'ES CENTRUM'),
('48', 'ES', 'ES PÓŁNOC'),
('49', 'ES', 'ES POŁUDNIE'),
('50', 'ES', 'ES PÓŁNOC'),
('52', 'ES', 'ES POŁUDNIE'),
('', 'FI', 'FI'),
('01', 'FR', 'FR WSCHÓD'),
('02', 'FR', 'FR PÓŁNOC'),
('03', 'FR', 'FR WSCHÓD'),
('04', 'FR', 'FR POŁUDNIE'),
('05', 'FR', 'FR WSCHÓD'),
('06', 'FR', 'FR POŁUDNIE'),
('07', 'FR', 'FR WSCHÓD'),
('08', 'FR', 'FR PÓŁNOC'),
('09', 'FR', 'FR POŁUDNIE'),
('10', 'FR', 'FR WSCHÓD'),
('11', 'FR', 'FR POŁUDNIE'),
('12', 'FR', 'FR POŁUDNIE'),
('13', 'FR', 'FR POŁUDNIE'),
('14', 'FR', 'FR ZACHÓD'),
('15', 'FR', 'FR WSCHÓD'),
('16', 'FR', 'FR ZACHÓD'),
('17', 'FR', 'FR ZACHÓD'),
('18', 'FR', 'FR CENTRUM'),
('19', 'FR', 'FR ZACHÓD'),
('20', 'FR', 'FR POŁUDNIE'),
('21', 'FR', 'FR WSCHÓD'),
('22', 'FR', 'FR ZACHÓD'),
('23', 'FR', 'FR ZACHÓD'),
('24', 'FR', 'FR ZACHÓD'),
('25', 'FR', 'FR WSCHÓD'),
('26', 'FR', 'FR WSCHÓD'),
('27', 'FR', 'FR ZACHÓD'),
('28', 'FR', 'FR CENTRUM'),
('29', 'FR', 'FR ZACHÓD'),
('30', 'FR', 'FR POŁUDNIE'),
('31', 'FR', 'FR POŁUDNIE'),
('32', 'FR', 'FR POŁUDNIE'),
('33', 'FR', 'FR ZACHÓD'),
('34', 'FR', 'FR POŁUDNIE'),
('35', 'FR', 'FR ZACHÓD'),
('36', 'FR', 'FR CENTRUM'),
('37', 'FR', 'FR CENTRUM'),
('38', 'FR', 'FR WSCHÓD'),
('39', 'FR', 'FR WSCHÓD'),
('40', 'FR', 'FR POŁUDNIE'),
('41', 'FR', 'FR CENTRUM'),
('42', 'FR', 'FR WSCHÓD'),
('43', 'FR', 'FR WSCHÓD'),
('44', 'FR', 'FR ZACHÓD'),
('45', 'FR', 'FR CENTRUM'),
('46', 'FR', 'FR POŁUDNIE'),
('47', 'FR', 'FR POŁUDNIE'),
('48', 'FR', 'FR POŁUDNIE'),
('49', 'FR', 'FR ZACHÓD'),
('50', 'FR', 'FR ZACHÓD'),
('51', 'FR', 'FR PÓŁNOC'),
('52', 'FR', 'FR WSCHÓD'),
('53', 'FR', 'FR ZACHÓD'),
('54', 'FR', 'FR PÓŁNOC'),
('55', 'FR', 'FR PÓŁNOC'),
('56', 'FR', 'FR ZACHÓD'),
('57', 'FR', 'FR PÓŁNOC'),
('58', 'FR', 'FR CENTRUM'),
('59', 'FR', 'FR PÓŁNOC'),
('60', 'FR', 'FR PÓŁNOC'),
('61', 'FR', 'FR ZACHÓD'),
('62', 'FR', 'FR PÓŁNOC'),
('63', 'FR', 'FR WSCHÓD'),
('64', 'FR', 'FR POŁUDNIE'),
('65', 'FR', 'FR POŁUDNIE'),
('66', 'FR', 'FR POŁUDNIE'),
('67', 'FR', 'FR WSCHÓD'),
('68', 'FR', 'FR WSCHÓD'),
('69', 'FR', 'FR WSCHÓD'),
('70', 'FR', 'FR WSCHÓD'),
('71', 'FR', 'FR WSCHÓD'),
('72', 'FR', 'FR ZACHÓD'),
('73', 'FR', 'FR WSCHÓD'),
('74', 'FR', 'FR WSCHÓD'),
('75', 'FR', 'FR CENTRUM'),
('76', 'FR', 'FR PÓŁNOC'),
('77', 'FR', 'FR CENTRUM'),
('78', 'FR', 'FR CENTRUM'),
('79', 'FR', 'FR ZACHÓD'),
('80', 'FR', 'FR PÓŁNOC'),
('81', 'FR', 'FR POŁUDNIE'),
('82', 'FR', 'FR POŁUDNIE'),
('83', 'FR', 'FR POŁUDNIE'),
('84', 'FR', 'FR POŁUDNIE'),
('85', 'FR', 'FR ZACHÓD'),
('86', 'FR', 'FR ZACHÓD'),
('87', 'FR', 'FR ZACHÓD'),
('88', 'FR', 'FR WSCHÓD'),
('89', 'FR', 'FR CENTRUM'),
('90', 'FR', 'FR WSCHÓD'),
('91', 'FR', 'FR CENTRUM'),
('92', 'FR', 'FR CENTRUM'),
('93', 'FR', 'FR CENTRUM'),
('94', 'FR', 'FR CENTRUM'),
('95', 'FR', 'FR CENTRUM'),
('98', 'FR', 'FR POŁUDNIE'),
('AB', 'GB', 'GB PÓŁNOC'),
('AL', 'GB', 'GB POŁUDNIE'),
('B', 'GB', 'GB CENTRUM'),
('B1', 'GB', 'GB CENTRUM'),
('B2', 'GB', 'GB CENTRUM'),
('B4', 'GB', 'GB CENTRUM'),
('B7', 'GB', 'GB CENTRUM'),
('B9', 'GB', 'GB CENTRUM'),
('BA', 'GB', 'GB CENTRUM'),
('BB', 'GB', 'GB CENTRUM'),
('BD', 'GB', 'GB CENTRUM'),
('BH', 'GB', 'GB CENTRUM'),
('BL', 'GB', 'GB CENTRUM'),
('BN', 'GB', 'GB POŁUDNIE'),
('BR', 'GB', 'GB POŁUDNIE'),
('BS', 'GB', 'GB CENTRUM'),
('CA', 'GB', 'GB PÓŁNOC'),
('CB', 'GB', 'GB POŁUDNIE'),
('CF', 'GB', 'GB CENTRUM'),
('CH', 'GB', 'GB CENTRUM'),
('CM', 'GB', 'GB POŁUDNIE'),
('CO', 'GB', 'GB POŁUDNIE'),
('CR', 'GB', 'GB POŁUDNIE'),
('CT', 'GB', 'GB POŁUDNIE'),
('CV', 'GB', 'GB CENTRUM'),
('CW', 'GB', 'GB CENTRUM'),
('DA', 'GB', 'GB POŁUDNIE'),
('DD', 'GB', 'GB PÓŁNOC'),
('DE', 'GB', 'GB CENTRUM'),
('DG', 'GB', 'GB PÓŁNOC'),
('DH', 'GB', 'GB PÓŁNOC'),
('DL', 'GB', 'GB CENTRUM'),
('DN', 'GB', 'GB CENTRUM'),
('DT', 'GB', 'GB CENTRUM'),
('DY', 'GB', 'GB CENTRUM'),
('E', 'GB', 'GB POŁUDNIE'),
('E1', 'GB', 'GB POŁUDNIE'),
('EC', 'GB', 'GB POŁUDNIE'),
('EH', 'GB', 'GB PÓŁNOC'),
('EN', 'GB', 'GB POŁUDNIE'),
('EX', 'GB', 'GB CENTRUM'),
('FK', 'GB', 'GB PÓŁNOC'),
('FY', 'GB', 'GB CENTRUM'),
('G', 'GB', 'GB PÓŁNOC'),
('G3', 'GB', 'GB PÓŁNOC'),
('G7', 'GB', 'GB PÓŁNOC'),
('GL', 'GB', 'GB CENTRUM'),
('GU', 'GB', 'GB POŁUDNIE'),
('HA', 'GB', 'GB POŁUDNIE'),
('HD', 'GB', 'GB CENTRUM'),
('HG', 'GB', 'GB CENTRUM'),
('HP', 'GB', 'GB POŁUDNIE'),
('HR', 'GB', 'GB CENTRUM'),
('HU', 'GB', 'GB CENTRUM'),
('HX', 'GB', 'GB CENTRUM'),
('IG', 'GB', 'GB POŁUDNIE'),
('IP', 'GB', 'GB POŁUDNIE'),
('IV', 'GB', 'GB PÓŁNOC'),
('KT', 'GB', 'GB POŁUDNIE'),
('KW', 'GB', 'GB PÓŁNOC'),
('KY', 'GB', 'GB PÓŁNOC'),
('L', 'GB', 'GB CENTRUM'),
('L2', 'GB', 'GB CENTRUM'),
('L3', 'GB', 'GB CENTRUM'),
('LA', 'GB', 'GB CENTRUM'),
('LD', 'GB', 'GB CENTRUM'),
('LE', 'GB', 'GB CENTRUM'),
('LL', 'GB', 'GB CENTRUM'),
('LN', 'GB', 'GB CENTRUM'),
('LS', 'GB', 'GB CENTRUM'),
('LU', 'GB', 'GB POŁUDNIE'),
('M', 'GB', 'GB CENTRUM'),
('M2', 'GB', 'GB CENTRUM'),
('M3', 'GB', 'GB CENTRUM'),
('M9', 'GB', 'GB CENTRUM'),
('ME', 'GB', 'GB POŁUDNIE'),
('MK', 'GB', 'GB CENTRUM'),
('ML', 'GB', 'GB PÓŁNOC'),
('N', 'GB', 'GB POŁUDNIE'),
('N1', 'GB', 'GB POŁUDNIE'),
('NE', 'GB', 'GB PÓŁNOC'),
('NG', 'GB', 'GB CENTRUM'),
('NN', 'GB', 'GB CENTRUM'),
('NP', 'GB', 'GB CENTRUM'),
('NR', 'GB', 'GB POŁUDNIE'),
('NW', 'GB', 'GB POŁUDNIE'),
('OL', 'GB', 'GB CENTRUM'),
('OX', 'GB', 'GB CENTRUM'),
('PA', 'GB', 'GB PÓŁNOC'),
('PE', 'GB', 'GB CENTRUM'),
('PH', 'GB', 'GB PÓŁNOC'),
('PL', 'GB', 'GB CENTRUM'),
('PO', 'GB', 'GB POŁUDNIE'),
('PR', 'GB', 'GB CENTRUM'),
('RG', 'GB', 'GB POŁUDNIE'),
('RH', 'GB', 'GB POŁUDNIE'),
('RM', 'GB', 'GB POŁUDNIE'),
('S', 'GB', 'GB CENTRUM'),
('S2', 'GB', 'GB CENTRUM'),
('S6', 'GB', 'GB CENTRUM'),
('S7', 'GB', 'GB CENTRUM'),
('SA', 'GB', 'GB CENTRUM'),
('SE', 'GB', 'GB POŁUDNIE'),
('SG', 'GB', 'GB POŁUDNIE'),
('SK', 'GB', 'GB CENTRUM'),
('SL', 'GB', 'GB POŁUDNIE'),
('SM', 'GB', 'GB POŁUDNIE'),
('SN', 'GB', 'GB CENTRUM'),
('SO', 'GB', 'GB POŁUDNIE'),
('SP', 'GB', 'GB CENTRUM'),
('SP', 'GB', 'GB CENTRUM'),
('SS', 'GB', 'GB POŁUDNIE'),
('ST', 'GB', 'GB CENTRUM'),
('SW', 'GB', 'GB POŁUDNIE'),
('SY', 'GB', 'GB CENTRUM'),
('TA', 'GB', 'GB CENTRUM'),
('TD', 'GB', 'GB PÓŁNOC'),
('TF', 'GB', 'GB CENTRUM'),
('TN', 'GB', 'GB POŁUDNIE'),
('TQ', 'GB', 'GB CENTRUM'),
('TR', 'GB', 'GB CENTRUM'),
('TS', 'GB', 'GB CENTRUM'),
('TW', 'GB', 'GB POŁUDNIE'),
('UB', 'GB', 'GB POŁUDNIE'),
('W', 'GB', 'GB POŁUDNIE'),
('WA', 'GB', 'GB CENTRUM'),
('WD', 'GB', 'GB POŁUDNIE'),
('WF', 'GB', 'GB CENTRUM'),
('WN', 'GB', 'GB CENTRUM'),
('WR', 'GB', 'GB CENTRUM'),
('WS', 'GB', 'GB CENTRUM'),
('WV', 'GB', 'GB CENTRUM'),
('YO', 'GB', 'GB CENTRUM'),
('', 'GR', 'GR'),
('', 'HR', 'HR'),
('', 'HU', 'HU'),
('', 'IE', 'IE'),
('0', 'IT', 'IT CENTRUM'),
('1', 'IT', 'IT CENTRUM'),
('2', 'IT', 'IT PÓŁNOC'),
('3', 'IT', 'IT PÓŁNOC'),
('4', 'IT', 'IT PÓŁNOC'),
('5', 'IT', 'IT CENTRUM'),
('6', 'IT', 'IT POŁUDNIE'),
('7', 'IT', 'IT POŁUDNIE'),
('8', 'IT', 'IT POŁUDNIE'),
('9', 'IT', 'IT POŁUDNIE'),
('', 'LI', 'LI'),
('', 'LU', 'LU'),
('', 'MC', 'FR POŁUDNIE'),
('', 'NL', 'NL'),
('0', 'PL', 'PL WSCHÓD'),
('1', 'PL', 'PL WSCHÓD'),
('2', 'PL', 'PL WSCHÓD'),
('3', 'PL', 'PL WSCHÓD'),
('4', 'PL', 'PL WSCHÓD'),
('5', 'PL', 'PL ZACHÓD'),
('6', 'PL', 'PL ZACHÓD'),
('7', 'PL', 'PL ZACHÓD'),
('8', 'PL', 'PL WSCHÓD'),
('9', 'PL', 'PL WSCHÓD'),
('', 'PT', 'PT'),
('', 'RO', 'RO'),
('', 'SE', 'SE'),
('1', 'SI', 'SI ZACHÓD'),
('2', 'SI', 'SI WSCHÓD'),
('3', 'SI', 'SI WSCHÓD'),
('4', 'SI', 'SI ZACHÓD'),
('5', 'SI', 'SI ZACHÓD'),
('6', 'SI', 'SI ZACHÓD'),
('7', 'SI', 'SI WSCHÓD'),
('8', 'SI', 'SI WSCHÓD'),
('9', 'SI', 'SI WSCHÓD'),
('', 'SK', 'SK'),
('M1', 'GB', 'GB CENTRUM'),
('L1', 'GB', 'GB CENTRUM'),
('L4', 'GB', 'GB CENTRUM'),
('W3', 'GB', 'GB POŁUDNIE'),
('G8', 'GB', 'GB PÓŁNOC'),
('G6', 'GB', 'GB PÓŁNOC')
('N9', 'GB', 'GB POŁUDNIE'),
('WA', 'GB', 'GB CENTRUM'),
    """

    # wyrażenie do złapania trójek (kod pocztowy, kod kraju, nazwa regionu)
    pattern = r"\('([^']*)',\s*'([^']*)',\s*'([^']*)'\)"
    matches = re.findall(pattern, region_data)

    for postal_code, country_code, region in matches:
        # null -> pusta stringa
        if postal_code.lower() == "null":
            postal_code = ""
        # normalizacja kraju tak, jak w normalize_country
        norm_country = normalize_country(country_code)
        # zapisujemy pod ujednoliconym kluczem
        REGION_MAPPING[(norm_country, postal_code)] = region

    print(f"Wczytano {len(REGION_MAPPING)} mapowań regionów.")

# Funkcja określająca region na podstawie kodu kraju i kodu pocztowego
def get_region(country, postal_code):
    postal_code = str(postal_code).strip()
    prefix2 = postal_code[:2] if len(postal_code) >= 2 else postal_code

    # 1) spróbuj dwu-cyfrowego
    region = REGION_MAPPING.get((country, prefix2))
    if region:
        return region

    # 2) spróbuj jednocyfrowego
    if len(postal_code) >= 1:
        prefix1 = postal_code[0]
        region = REGION_MAPPING.get((country, prefix1))
        if region:
            return region

    # 3) fallback na kraj-bez-prefixu
    return REGION_MAPPING.get((country, ""), None)


# Funkcja pobierająca stawki na podstawie relacji region-region
def get_region_based_rates(lc, lp, uc, up):
    """Pobiera stawki na podstawie relacji region–region oraz sumę zleceń."""
    # Określ regiony dla kodów pocztowych
    lc_region = get_region(normalize_country(lc), lp)
    uc_region = get_region(normalize_country(uc), up)

    if not lc_region or not uc_region:
        # Nie można określić regionów
        return {
            'region_gielda_stawka_3m': None,
            'region_gielda_stawka_6m': None,
            'region_gielda_stawka_12m': None,
            'region_klient_stawka_3m': None,
            'region_klient_stawka_6m': None,
            'region_klient_stawka_12m': None,
            'region_relacja': f"{lc_region or lc} - {uc_region or uc}",
            'region_dopasowanie': "Brak dopasowania regionalnego",
            'region_gielda_dopasowanie': None,
            'region_klient_dopasowanie': None
        }

    try:
        hist_df = pd.read_excel("historical_rates.xlsx",
                                dtype={'kod pocztowy zaladunku': str, 'kod pocztowy rozladunku': str})
        gielda_df = pd.read_excel("historical_rates_gielda.xlsx",
                                  dtype={'kod pocztowy zaladunku': str, 'kod pocztowy rozladunku': str})
    except Exception as e:
        return {
            'region_gielda_stawka_3m': None,
            'region_gielda_stawka_6m': None,
            'region_gielda_stawka_12m': None,
            'region_klient_stawka_3m': None,
            'region_klient_stawka_6m': None,
            'region_klient_stawka_12m': None,
            'region_podlot': None,  # Dodane pole dla regionalnego podlotu
            'region_relacja': f"{lc_region} - {uc_region}",
            'region_dopasowanie': f"Błąd: {e}",
            'region_gielda_dopasowanie': None,
            'region_klient_dopasowanie': None
        }

    # Dodaj kolumny regionów
    hist_df['region_zaladunku'] = hist_df.apply(
        lambda r: get_region(normalize_country(r['kraj zaladunku']), r['kod pocztowy zaladunku']), axis=1)
    hist_df['region_rozladunku'] = hist_df.apply(
        lambda r: get_region(normalize_country(r['kraj rozladunku']), r['kod pocztowy rozladunku']), axis=1)

    gielda_df['region_zaladunku'] = gielda_df.apply(
        lambda r: get_region(normalize_country(r['kraj zaladunku']), r['kod pocztowy zaladunku']), axis=1)
    gielda_df['region_rozladunku'] = gielda_df.apply(
        lambda r: get_region(normalize_country(r['kraj rozladunku']), r['kod pocztowy rozladunku']), axis=1)

    # Filtruj po regionach
    hist_matches = hist_df[
        (hist_df['region_zaladunku'] == lc_region) &
        (hist_df['region_rozladunku'] == uc_region)
    ]
    gielda_matches = gielda_df[
        (gielda_df['region_zaladunku'] == lc_region) &
        (gielda_df['region_rozladunku'] == uc_region)
    ]

    # Przygotuj wynik
    result = {
        'region_gielda_stawka_3m': None,
        'region_gielda_stawka_6m': None,
        'region_gielda_stawka_12m': None,
        'region_klient_stawka_3m': None,
        'region_klient_stawka_6m': None,
        'region_klient_stawka_12m': None,
        'region_podlot': None,  # Dodane pole dla regionalnego podlotu
        'region_relacja': f"{lc_region} - {uc_region}",
        'region_dopasowanie': "Brak dopasowań",
        'region_gielda_dopasowanie': None,
        'region_klient_dopasowanie': None
    }

    # --- Klient (historyczne) ---
    if not hist_matches.empty:
        # suma zleceń
        if 'Liczba zlecen' in hist_matches.columns:
            total_orders_hist = hist_matches['Liczba zlecen'].sum()
        else:
            total_orders_hist = len(hist_matches)
        result['region_dopasowanie'] = f"Dopasowano {total_orders_hist} zleceń (klient)"
        result['region_klient_dopasowanie'] = total_orders_hist

        # średnie ważone stawek
        for period, col in [('3m', 'stawka_3m'), ('6m', 'stawka_6m'), ('12m', 'stawka_12m')]:
            if col in hist_matches.columns:
                valid = hist_matches.dropna(subset=[col])
                if not valid.empty:
                    if 'Liczba zlecen' in valid.columns:
                        orders = valid['Liczba zlecen'].sum()
                        weights = valid['Liczba zlecen'] / orders if orders else None
                    else:
                        weights = pd.Series(1 / len(valid), index=valid.index)
                    if weights is not None:
                        result[f'region_klient_stawka_{period}'] = (valid[col] * weights).sum()

        # Regionalny podlot (klient/historyczne)
        region_podlot_hist = calculate_podlot_from_data(hist_matches, "regionalny podlot historyczny")
        if region_podlot_hist is not None:
            result['region_podlot'] = region_podlot_hist

    # --- Giełda ---
    if not gielda_matches.empty:
        # suma zleceń
        if 'Liczba zlecen' in gielda_matches.columns:
            total_orders_gielda = gielda_matches['Liczba zlecen'].sum()
        else:
            total_orders_gielda = len(gielda_matches)
        if result['region_dopasowanie'] == "Brak dopasowań":
            result['region_dopasowanie'] = f"Dopasowano {total_orders_gielda} zleceń (giełda)"
        else:
            result['region_dopasowanie'] += f", {total_orders_gielda} zleceń (giełda)"
        result['region_gielda_dopasowanie'] = total_orders_gielda

        # średnie ważone stawek giełdowych
        for period, col in [('3m', 'stawka_3m'), ('6m', 'stawka_6m'), ('12m', 'stawka_12m')]:
            if col in gielda_matches.columns:
                valid = gielda_matches.dropna(subset=[col])
                if not valid.empty:
                    if 'Liczba zlecen' in valid.columns:
                        orders = valid['Liczba zlecen'].sum()
                        weights = valid['Liczba zlecen'] / orders if orders else None
                    else:
                        weights = pd.Series(1 / len(valid), index=valid.index)
                    if weights is not None:
                        result[f'region_gielda_stawka_{period}'] = (valid[col] * weights).sum()

        # Regionalny podlot (giełda) - tylko jeśli nie mamy jeszcze z danych historycznych
        if result['region_podlot'] is None:
            region_podlot_gielda = calculate_podlot_from_data(gielda_matches, "regionalny podlot giełda")
            if region_podlot_gielda is not None:
                result['region_podlot'] = region_podlot_gielda

    return result


# Funkcja wczytująca dane z global_data.csv – klucze tworzymy jako stringi, np. "Poland_36"
def load_global_data(filepath):
    with open(filepath, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=",")
        for row in reader:
            country_norm = normalize_country(row['country'].strip())
            prefix = row['prefix'].strip()
            try:
                lat = float(row['latitude'])
                lon = float(row['longitude'])
                key = f"{country_norm}_{prefix}"
                LOOKUP_DICT[key] = (lat, lon)
            except Exception as e:
                print("Błąd wczytywania rekordu:", row, e)
    #print("Wczytany LOOKUP_DICT:", LOOKUP_DICT)


# Flaga do kontroli synchronizacji
ENABLE_CACHE_SYNC = True  # Synchronizacja tymczasowo wyłączona

def sync_geo_cache_with_lookup():
    """Synchronizuje geo_cache z danymi z LOOKUP_DICT dla spójności."""
    if not ENABLE_CACHE_SYNC:
        print("Synchronizacja cache jest tymczasowo wyłączona.")
        return 0
        
    print("Synchronizowanie geo_cache z LOOKUP_DICT...")
    synced_count = 0

    for key, (lat, lon) in LOOKUP_DICT.items():
        if key not in geo_cache or geo_cache[key][0] is None:
            geo_cache[key] = (lat, lon, 'lookup_sync', 'sync')
            synced_count += 1

    print(f"Zsynchronizowano {synced_count} wpisów z LOOKUP_DICT do geo_cache")
    return synced_count


# Na starcie aplikacji
load_global_data("global_data.csv")
sync_geo_cache_with_lookup()  # Synchronizuj cache
load_region_mapping()  # Wczytaj mapowania regionów


def clean_text(text):
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-zA-Z0-9\s]', '', text)
    return text.lower().strip()


def get_geocoding_progress():
    if GEOCODING_TOTAL > 0:
        progress = int((GEOCODING_CURRENT / GEOCODING_TOTAL) * 100)
    else:
        progress = 0
    return f"Geokodowanie: {progress}%"


def generate_query_variants(country, postal_code, city=None):
    norm_country = normalize_country(country)
    norm_postal = str(postal_code).strip()
    variants = []
    
    # Sprawdź czy miasto jest prawidłowe (nie puste, nie "nan")
    is_city_valid = (city and 
                    isinstance(city, str) and 
                    city.strip() and 
                    not pd.isna(city) and
                    str(city).lower().strip() not in ['nan', 'none', 'null'])
    
    # Dla dwucyfrowych kodów – pierwszy wariant to prosty: "Poland, 36" z kluczem "Poland_36"
    if len(norm_postal) == 2:
        variants.append((f"{norm_country}, {norm_postal}", f"{norm_country}_{norm_postal}"))
        if is_city_valid:
            norm_city = city.strip().title()
            clean_city = clean_text(norm_city)
            variants.append((f"{norm_country} postal code {norm_postal}, {norm_city}",
                             f"{norm_country}_postal_{norm_postal}_{clean_city}"))
        return variants
    else:
        if is_city_valid:
            norm_city = city.strip().title()
            clean_city = clean_text(norm_city)
            variants.append(
                (f"{norm_country}, {norm_postal}, {norm_city}", f"{norm_country}_{norm_postal}_{clean_city}"))
            variants.append(
                (f"{norm_city}, {norm_postal}, {norm_country}", f"{clean_city}_{norm_postal}_{norm_country}"))
            variants.append(
                (f"{norm_city}, {norm_country}, {norm_postal}", f"{clean_city}_{norm_country}_{norm_postal}"))
            variants.append((f"{norm_city}, {norm_country}", f"{clean_city}_{norm_country}"))
            variants.append((f"{norm_postal}, {norm_country}", f"{norm_country}_{norm_postal}"))
        else:
            variants.append((f"{norm_postal}, {norm_country}", f"{norm_country}_{norm_postal}"))
        return variants


def ptv_geocode_by_address(country, postal_code=None, city=None, api_key=None, language="pl"):
    """
    Geokodowanie używając strukturyzowanego endpoint locations/by-address PTV API
    Z logiką fallback:
    1. Próbuje kraj + kod pocztowy + miasto
    2. Jeśli nie znajdzie i ma kod: kraj + kod pocztowy
    3. Jeśli nie znajdzie i ma miasto: kraj + miasto
    
    Eliminuje problem z "STREET" poprzez użycie strukturyzowanych parametrów
    """
    endpoint = "https://api.myptv.com/geocoding/v1/locations/by-address"
    logger = logging.getLogger("ptv_api_manager")
    
    # Sprawdź czy miasto jest prawidłowe (nie "nan")
    has_valid_city = (city and 
                      str(city).strip() and 
                      not pd.isna(city) and
                      str(city).lower().strip() not in ['nan', 'none', 'null'])
    
    has_postal_code = postal_code and str(postal_code).strip()
    
    # Lista kombinacji do sprawdzenia (w kolejności priorytetów)
    attempts = []
    
    # 1. Pełna kombinacja: kraj + kod + miasto (jeśli oba dostępne)
    if has_postal_code and has_valid_city:
        attempts.append({
            "country": country,
            "postalCode": str(postal_code).strip(),
            "locality": str(city).strip(),
            "description": f"country={country}, postal_code={postal_code}, city={city}"
        })
    
    # 2. Kraj + kod pocztowy (jeśli kod dostępny)
    if has_postal_code:
        attempts.append({
            "country": country,
            "postalCode": str(postal_code).strip(),
            "description": f"country={country}, postal_code={postal_code}"
        })
    
    # 3. Kraj + miasto (jeśli miasto dostępne)
    if has_valid_city:
        attempts.append({
            "country": country,
            "locality": str(city).strip(),
            "description": f"country={country}, city={city}"
        })
    
    # Jeśli nie ma ani kodu ani miasta, nie próbuj
    if not attempts:
        logger.warning(f"PTV API (by-address): Brak wystarczających danych dla kraju {country}")
        return (None, None, 'nieznane', 'brak danych')
    
    # Próbuj każdą kombinację
    for i, attempt in enumerate(attempts):
        params = {
            "apiKey": api_key or PTV_API_KEY,
            "language": language
        }
        
        # Dodaj parametry dla tej próby
        if "country" in attempt:
            params["country"] = attempt["country"]
        if "postalCode" in attempt:
            params["postalCode"] = attempt["postalCode"]
        if "locality" in attempt:
            params["locality"] = attempt["locality"]
        
        logger.info(f"PTV API (by-address) próba {i+1}/{len(attempts)}: {attempt['description']}")
        
        try:
            response = requests.get(endpoint, params=params, timeout=10)
            time.sleep(0.1)  # Rate limiting
            
            if response.status_code == 200:
                data = response.json()
                if data.get("locations"):
                    result = data["locations"][0]
                    
                    # Sprawdź czy zwrócona lokalizacja ma odpowiedni kraj
                    returned_country = result.get("address", {}).get("countryCode", "").upper()
                    if country:
                        country_iso = ISO_CODES.get(normalize_country(country), "").upper()
                        if country_iso and returned_country != country_iso:
                            logger.warning(f"PTV API (by-address): BŁĄD KRAJU - Zwrócony: {returned_country}, oczekiwany: {country_iso}")
                            continue  # Próbuj następną kombinację
                    
                    # Sprawdź kod pocztowy jeśli został podany w tej próbie (podstawowe sprawdzenie)
                    if "postalCode" in attempt:
                        returned_postal = result.get("address", {}).get("postalCode", "").strip()
                        requested_postal = attempt["postalCode"]
                        
                        # Loguj szczegóły odpowiedzi dla diagnostyki
                        logger.debug(f"PTV API (by-address): Zwrócony kod pocztowy: '{returned_postal}', szukany: '{requested_postal}'")
                        
                        # Sprawdzenie kodu pocztowego
                        if not returned_postal and requested_postal:
                            # PTV API nie zwróciło kodu pocztowego, ale szukaliśmy konkretnego kodu
                            logger.warning(f"PTV API (by-address): Brak kodu pocztowego w odpowiedzi, ale szukano: {requested_postal}")
                            continue  # Próbuj następną kombinację
                        elif returned_postal and requested_postal:
                            # Sprawdź zgodność kodów pocztowych
                            if returned_postal != requested_postal:
                                # Sprawdź czy pierwsza cyfra się zgadza (dla mniej oczywistych błędów)
                                requested_first_digit = ''.join(c for c in requested_postal if c.isdigit())
                                returned_first_digit = ''.join(c for c in returned_postal if c.isdigit())
                                
                                if (requested_first_digit and returned_first_digit and 
                                    len(requested_first_digit) > 0 and len(returned_first_digit) > 0 and
                                    requested_first_digit[0] != returned_first_digit[0]):
                                    logger.warning(f"PTV API (by-address): Bardzo różne kody pocztowe - zwrócony: {returned_postal}, szukany: {requested_postal}")
                                    continue  # Próbuj następną kombinację dla oczywistych błędów
                                else:
                                    logger.info(f"PTV API (by-address): Różne kody pocztowe, ale podobne - zwrócony: {returned_postal}, szukany: {requested_postal}")
                            else:
                                logger.debug(f"PTV API (by-address): Kod pocztowy zgodny: {returned_postal}")
                    
                    lat = result["referencePosition"]["latitude"]
                    lon = result["referencePosition"]["longitude"]
                    quality = result.get("locationType", "STRUCTURED_ADDRESS")
                    source = f"PTV API (structured-{i+1})"
                    
                    logger.info(f"PTV API (by-address): Znaleziono w próbie {i+1}: {lat}, {lon}")
                    return (lat, lon, quality, source)
                else:
                    logger.warning(f"PTV API (by-address): Brak wyników dla próby {i+1}: {attempt['description']}")
            else:
                logger.error(f"PTV API (by-address) błąd w próbie {i+1}: {response.status_code} - {response.text}")
        
        except requests.exceptions.Timeout:
            logger.warning(f"PTV API (by-address): Timeout w próbie {i+1}")
        except Exception as e:
            logger.error(f"PTV API (by-address) wyjątek w próbie {i+1}: {str(e)}")
    
    # Jeśli wszystkie próby się nie powiodły
    logger.warning(f"PTV API (by-address): Wszystkie próby nieudane dla country={country}, postal_code={postal_code}, city={city}")
    return (None, None, 'nieznane', 'brak danych')


def ptv_geocode_by_text(search_text, api_key, language="pl", country_code=None):
    endpoint = "https://api.myptv.com/geocoding/v1/locations/by-text"
    params = {
        "searchText": search_text,
        "apiKey": api_key,
        "language": language
    }

    # Dodaj parametr countryFilter, jeśli podano kod kraju
    if country_code:
        params["countryFilter"] = country_code.upper()

    logger = logging.getLogger("ptv_api_manager")
    logger.info(f"PTV API: Wysyłam zapytanie: '{search_text}'" + (f" z filtrem kraju: {country_code}" if country_code else ""))
    
    try:
        # Dodajemy timeout 10 sekund
        response = requests.get(endpoint, params=params, timeout=10)
        time.sleep(0.1)  # Rate limiting
        
        if response.status_code == 200:
            data = response.json()
            if data.get("locations"):
                result = data["locations"][0]
                
                # Sprawdź czy zwrócona lokalizacja jest w odpowiednim kraju
                if country_code:
                    returned_country = result.get("address", {}).get("countryCode", "").upper()
                    expected_country = country_code.upper()
                    
                    if returned_country != expected_country:
                        logger.warning(f"PTV API: BŁĄD KRAJU - Zwrócony kraj ({returned_country}) nie zgadza się z oczekiwanym ({expected_country}) dla zapytania '{search_text}'")
                        # Dodatkowy log ze szczegółami odpowiedzi
                        logger.warning(f"PTV API: Pełna odpowiedź adresu: {result.get('address', {})}")
                        return (None, None, 'nieznane', f'niezgodny kraj: oczekiwano {expected_country}, otrzymano {returned_country}')
                
                # Sprawdź czy to zapytanie o miasto
                if "," in search_text and not search_text.replace(" ", "").split(",")[0].replace("-", "").isdigit():
                    # To może być zapytanie typu "MIASTO, KRAJ"
                    requested_city = search_text.split(",")[0].strip().upper()
                    returned_city = result.get("address", {}).get("locality", "").strip().upper()
                    
                    # Jeśli nie ma locality, sprawdź municipality lub city
                    if not returned_city:
                        returned_city = result.get("address", {}).get("municipality", "").strip().upper()
                    if not returned_city:
                        returned_city = result.get("address", {}).get("city", "").strip().upper()
                    
                    if returned_city and requested_city:
                        # Sprawdź czy nazwy miast są podobne (przynajmniej 70% zgodności)
                        from difflib import SequenceMatcher
                        similarity = SequenceMatcher(None, requested_city, returned_city).ratio()
                        
                        logger.debug(f"PTV API: Porównuję miasta - szukane: '{requested_city}', zwrócone: '{returned_city}', podobieństwo: {similarity:.2f}")
                        
                        if similarity < 0.7:  # Mniej niż 70% podobieństwa
                            logger.warning(f"PTV API: Zwrócone miasto ({returned_city}) ma zbyt małe podobieństwo ({similarity:.2f}) do szukanego ({requested_city})")
                            return (None, None, 'nieznane', f'niezgodne miasto: {returned_city} vs {requested_city}')
                
                # Sprawdź czy to zapytanie o kod pocztowy
                elif "," in search_text and search_text.replace(" ", "").split(",")[0].replace("-", "").isdigit():
                    requested_postal = search_text.split(",")[0].strip()
                    returned_postal = result["address"].get("postalCode", "").strip()
                    logger.debug(f"PTV API: Porównuję kody - szukany: '{requested_postal}', zwrócony: '{returned_postal}'")
                    
                    # Sprawdź czy pierwsza cyfra się zgadza
                    requested_first_digit = ''.join(c for c in requested_postal if c.isdigit())[0]
                    returned_first_digit = ''.join(c for c in returned_postal if c.isdigit())[0]
                    
                    if requested_first_digit != returned_first_digit:
                        logger.warning(f"PTV API: Pierwsza cyfra kodu ({returned_first_digit}) nie zgadza się z szukaną ({requested_first_digit})")
                        return (None, None, 'nieznane', 'niezgodna pierwsza cyfra kodu')
                    
                    # Jeśli zwrócony kod jest inny niż szukany, traktuj jako błąd
                    if requested_postal != returned_postal:
                        logger.warning(f"PTV API: Zwrócony kod pocztowy ({returned_postal}) nie zgadza się z szukanym ({requested_postal})")
                        return (None, None, 'nieznane', 'niezgodny kod pocztowy')
                
                lat = result["referencePosition"]["latitude"]
                lon = result["referencePosition"]["longitude"]
                quality = result.get("locationType", "PTV_API")
                source = "PTV API"
                logger.info(f"PTV API: Znaleziono wynik: {lat}, {lon}")
                return (lat, lon, quality, source)
            else:
                logger.warning(f"PTV API: Brak wyników dla '{search_text}'")
                return (None, None, 'nieznane', 'brak danych')
        else:
            logger.error(f"PTV API błąd: {response.status_code} - {response.text}")
            return (None, None, 'nieznane', f'błąd PTV API: {response.status_code}')
    except requests.exceptions.Timeout:
        logger.warning(f"PTV API: Timeout dla zapytania '{search_text}'")
        return (None, None, 'nieznane', 'timeout')
    except Exception as e:
        logger.error(f"PTV API wyjątek: {str(e)} dla zapytania '{search_text}'")
        return (None, None, 'nieznane', str(e))


# ============================================================================
# FUNKCJE POMOCNICZE - Obsługa tras z punktami pośrednimi
# ============================================================================

def calculate_multi_waypoint_route(route_request: RouteRequest):
    """
    Oblicza trasę z punktami pośrednimi.
    
    Args:
        route_request: Obiekt RouteRequest z wszystkimi danymi trasy
    
    Returns:
        Dict z wynikami:
        {
            'success': bool,
            'distance': float (km),
            'legs': List[Dict],
            'segments_info': List[str],  # Opisy segmentów
            'toll_cost': float,
            'road_toll': float,
            'other_toll': float,
            'polyline': str,
            'toll_details': Dict,
            'special_systems': List,
            'error_message': str (jeśli success=False),
            'failed_points': List[WaypointData] (jeśli success=False)
        }
    """
    logger.info(f"Rozpoczynam obliczanie trasy z waypoints: {route_request}")
    
    # KROK 1: Geokodowanie wszystkich punktów
    all_points = route_request.get_all_points_ordered()
    geocoded_coords = []
    failed_points = []
    
    for i, point in enumerate(all_points):
        coords = get_coordinates(point.country, point.postal_code, point.city)
        
        if coords and len(coords) >= 2 and None not in coords[:2]:
            point.coordinates = (coords[0], coords[1])
            geocoded_coords.append((coords[0], coords[1]))
            logger.debug(f"Punkt {i+1}/{len(all_points)} OK: {point}")
        else:
            failed_points.append(point)
            logger.warning(f"Punkt {i+1}/{len(all_points)} FAILED: {point}")
    
    if failed_points:
        failed_str = ", ".join(str(p) for p in failed_points)
        error_msg = (
            f"Nie można geokodować {len(failed_points)} punkt(ów): {failed_str}. "
            f"Sprawdź poprawność kodów pocztowych i nazw miast."
        )
        logger.error(error_msg)
        return {
            'success': False,
            'error_message': error_msg,
            'failed_points': failed_points
        }
    
    logger.info(f"Geokodowanie zakończone sukcesem: {len(geocoded_coords)} punktów")
    
    # KROK 2: Wywołanie PTV API z wieloma waypoints
    ptv_result = ptv_manager.get_route_with_waypoints(
        waypoints=geocoded_coords,
        avoid_switzerland=route_request.avoid_switzerland,
        routing_mode=route_request.routing_mode
    )
    
    if not ptv_result:
        error_msg = "PTV API nie zwróciło wyników dla trasy"
        logger.error(error_msg)
        return {
            'success': False,
            'error_message': error_msg
        }
    
    # KROK 3: Analiza segmentów
    segments_info = []
    legs = ptv_result.get('legs', [])
    
    for i, leg in enumerate(legs):
        if i + 1 < len(all_points):
            from_point = all_points[i]
            to_point = all_points[i + 1]
            distance = leg.get('distance', 0) / 1000  # m → km
            duration = leg.get('duration', 0) / 3600  # s → h
            
            segment_desc = (
                f"Segment {i+1}: {from_point} → {to_point} "
                f"({distance:.1f}km, {duration:.1f}h)"
            )
            segments_info.append(segment_desc)
            logger.debug(segment_desc)
    
    # KROK 4: Zwróć wyniki
    result = {
        'success': True,
        'distance': ptv_result.get('distance', 0.0),
        'legs': legs,
        'segments_info': segments_info,
        'toll_cost': ptv_result.get('toll_cost', 0.0),
        'road_toll': ptv_result.get('road_toll', 0.0),
        'other_toll': ptv_result.get('other_toll', 0.0),
        'polyline': ptv_result.get('polyline', ''),
        'toll_details': ptv_result.get('toll_details', {}),
        'special_systems': ptv_result.get('special_systems', [])
    }
    
    logger.info(
        f"Trasa obliczona pomyślnie: {result['distance']:.1f}km, "
        f"{len(segments_info)} segmentów, {result['toll_cost']:.2f}€ opłat"
    )
    
    return result


def parse_waypoints_from_form(form_data):
    """
    Parsuje punkty pośrednie z danych formularza Flask.
    
    Args:
        form_data: request.form z Flask
    
    Returns:
        List[WaypointData] - lista punktów pośrednich (może być pusta)
    """
    waypoints = []
    
    for i in range(1, 6):  # max 5 waypoints
        wp_country = form_data.get(f'waypoint_{i}_country', '').strip().upper()
        wp_postal = form_data.get(f'waypoint_{i}_postal', '').strip()
        wp_city = form_data.get(f'waypoint_{i}_city', '').strip() or None
        
        if wp_country and wp_postal:
            waypoints.append(WaypointData(
                country=wp_country,
                postal_code=wp_postal,
                city=wp_city
            ))
    
    return waypoints


def parse_waypoints_from_excel_row(row, column_name='Punkty_posrednie'):
    """
    Parsuje punkty pośrednie z wiersza Excel (format Compact).
    
    Format obsługiwany:
    - "CZ:11000;AT:1010" - bez miast
    - "CZ:11000:Praha;AT:1010:Wien" - z miastami
    - "" lub NaN - brak punktów
    
    Args:
        row: pandas Series (wiersz DataFrame)
        column_name: nazwa kolumny z waypointami (default: 'Punkty_posrednie')
    
    Returns:
        List[WaypointData] - lista punktów pośrednich (może być pusta)
    
    Examples:
        >>> row = pd.Series({'Punkty_posrednie': 'CZ:11000;AT:1010'})
        >>> waypoints = parse_waypoints_from_excel_row(row)
        >>> len(waypoints)
        2
    """
    waypoints = []
    
    # Sprawdź czy kolumna istnieje
    if column_name not in row:
        return waypoints
    
    waypoints_str = row[column_name]
    
    # Sprawdź czy wartość jest pusta lub NaN
    if pd.isna(waypoints_str) or not waypoints_str:
        return waypoints
    
    waypoints_str = str(waypoints_str).strip()
    if not waypoints_str:
        return waypoints
    
    # Parse każdy punkt oddzielony średnikiem
    parts = waypoints_str.split(';')
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        # Split po dwukropku: KRAJ:KOD[:MIASTO]
        tokens = part.split(':')
        
        if len(tokens) < 2:
            logger.warning(f"Nieprawidłowy format waypoint: '{part}' - oczekiwano KRAJ:KOD[:MIASTO]")
            continue
        
        country = tokens[0].strip().upper()
        postal_code = tokens[1].strip()
        city = tokens[2].strip() if len(tokens) > 2 else None
        
        if country and postal_code:
            waypoints.append(WaypointData(
                country=country,
                postal_code=postal_code,
                city=city
            ))
            logger.debug(f"Sparsowano waypoint: {country} {postal_code}" + (f" ({city})" if city else ""))
        else:
            logger.warning(f"Pominięto waypoint z pustym krajem lub kodem: '{part}'")
    
    return waypoints


def get_coordinates(country, postal_code, city=None):
    global GEOCODING_CURRENT, GEOCODING_TOTAL
    
    norm_postal = str(postal_code).strip()
    # Upewnij się, że city jest ciągiem znaków – jeśli nie, ustaw pusty ciąg
    if (city is None or 
        not hasattr(city, "strip") or 
        city.strip() == "" or
        pd.isna(city) or
        str(city).lower().strip() in ['nan', 'none', 'null']):
        city = ""
    else:
        try:
            city = str(city).strip()
        except Exception:
            city = ""
    print(">>> get_coordinates wywołane dla:", country, norm_postal, city)

    norm_country = normalize_country(country)
    standard_key = f"{norm_country}_{norm_postal}"
    
    # Pobierz kod ISO kraju
    iso_code = ISO_CODES.get(norm_country, "")
    
    # Sprawdź czy mamy kod ISO dla danego kraju
    if not iso_code:
        print(f"UWAGA: Brak kodu ISO dla kraju '{norm_country}'. Geokodowanie może być nieprecyzyjne!")
        print(f"Dostępne kraje w ISO_CODES: {list(ISO_CODES.keys())}")
    else:
        print(f"Znaleziono kod ISO dla kraju '{norm_country}': '{iso_code}'")



    # Sprawdź czy lokalizacja jest już w cache
    if standard_key in geo_cache:
        cached = geo_cache[standard_key]
        if cached[0] is not None:
            print(f"Znaleziono wynik w cache dla klucza: {standard_key}: {cached}")
            return cached

    # NOWA STRATEGIA GEOKODOWANIA:
    # 1. PTV API (structured) -> 2. PTV API (text) -> 3. LOOKUP_DICT -> 4. Nominatim
    
    # 1. Próba geokodowania przez PTV API (structured) - NAJWYŻSZY PRIORYTET
    print(f"1. Rozpoczynam geokodowanie przez PTV API (structured)...")
    print(f"Kraj: {norm_country}, Kod ISO: '{iso_code}', Kod pocztowy: {norm_postal}, Miasto: {city}")
    
    ptv_structured_result = ptv_geocode_by_address(
        country=norm_country,
        postal_code=norm_postal if norm_postal else None,
        city=city if city and city.strip() else None,
        api_key=PTV_API_KEY,
        language="pl"
    )
    
    if ptv_structured_result[0] is not None:
        print(f"PTV API (structured) - zwrócił wynik: {ptv_structured_result}")
        geo_cache[standard_key] = ptv_structured_result
        print(f"Zapisano do geo_cache: {standard_key} -> {ptv_structured_result}")
        return ptv_structured_result

    # Różne ścieżki dla dwucyfrowych i nie-dwucyfrowych kodów (fallback)
    if len(norm_postal) == 2:
        # DLA KODÓW DWUCYFROWYCH - KOLEJNOŚĆ FALLBACK:
        # 2. PTV API (text) -> 3. LOOKUP_DICT -> 4. Nominatim
        print("2. PTV API (structured) nie zwróciło wyniku, próbuję PTV API (text)...")

        # Generujemy warianty zapytań
        query_variants = [(f"{norm_country} postal code {norm_postal}", standard_key)]
        if city and city.strip():
            norm_city = city.strip().title()
            clean_city = clean_text(norm_city)
            query_variants.append((f"{norm_country} postal code {norm_postal}, {norm_city}",
                                   f"{norm_country}_postal_{norm_postal}_{clean_city}"))

        for query_string, variant_key in query_variants:
            print(f"PTV API (text) - wysyłam zapytanie: '{query_string}' z country_code='{iso_code}'")
            ptv_result = ptv_geocode_by_text(query_string, PTV_API_KEY, language="pl", country_code=iso_code)
            if ptv_result[0] is not None:
                print(f"PTV API (text) - wariant '{query_string}' zwrócił wynik: {ptv_result}")
                geo_cache[variant_key] = ptv_result
                # Zapisz wynik również pod standardowym kluczem
                key = f"{norm_country}_{norm_postal}"
                if key != variant_key:
                    geo_cache[key] = ptv_result
                print(f"Zapisano do geo_cache: {variant_key} -> {ptv_result}")
                return ptv_result

        # 3. Sprawdzamy LOOKUP_DICT
        key = standard_key
        print(f"3. PTV API (text) nie zwróciło wyniku, sprawdzam klucz '{key}' w LOOKUP_DICT")
        if key in LOOKUP_DICT:
            lat, lon = LOOKUP_DICT[key]
            result = (lat, lon, "lookup", "lookup")
            print(f"LOOKUP: Znaleziono współrzędne dla {key}: {result}")
            geo_cache[key] = result
            return result

        # 4. Jeśli wszystkie metody PTV i LOOKUP_DICT nie znalazły lokalizacji, próbujemy Nominatim
        print("4. LOOKUP_DICT nie zwrócił wyniku, wywołuję Nominatim...")
        for query_string, variant_key in query_variants:
            print(f"Nominatim - próba zapytania: '{query_string}' (klucz: {variant_key}) z country_codes={iso_code}")
            try:
                extra_params = {}  # Dla kodów nie-dwucyfrowych nie potrzebujemy polygon_geojson
                location = geolocator.geocode(query_string, exactly_one=True, country_codes=iso_code, **extra_params)
                time.sleep(0.1)
                if location:
                    # Sprawdź czy Nominatim zwrócił kod pocztowy
                    returned_postal = location.raw.get('address', {}).get('postcode', '')
                    if not returned_postal:
                        print(f"Nominatim: Brak kodu pocztowego w odpowiedzi")
                        continue
                    
                    # Sprawdź czy pierwsza cyfra kodu pocztowego się zgadza
                    requested_first_digit = ''.join(c for c in norm_postal if c.isdigit())[0]
                    returned_first_digit = ''.join(c for c in returned_postal if c.isdigit())[0]
                    
                    if requested_first_digit != returned_first_digit:
                        print(f"Nominatim: Pierwsza cyfra kodu ({returned_first_digit}) nie zgadza się z szukaną ({requested_first_digit})")
                        continue
                    
                    if "geojson" in location.raw and location.raw["geojson"]:
                        geo = location.raw["geojson"]
                        if geo.get("type") == "Polygon":
                            coords = geo.get("coordinates")[0]
                            lat = sum(point[1] for point in coords) / len(coords)
                            lon = sum(point[0] for point in coords) / len(coords)
                            quality = "centroid (polygon)"
                        else:
                            bb = location.raw.get("boundingbox")
                            if bb:
                                lat = (float(bb[0]) + float(bb[1])) / 2
                                lon = (float(bb[2]) + float(bb[3])) / 2
                                quality = "przybliżone (bounding box)"
                            else:
                                lat, lon = location.latitude, location.longitude
                                quality = "dokładne"
                    else:
                        if hasattr(location, 'raw') and 'boundingbox' in location.raw:
                            bb = location.raw['boundingbox']
                            lat = (float(bb[0]) + float(bb[1])) / 2
                            lon = (float(bb[2]) + float(bb[3])) / 2
                            quality = 'przybliżone (bounding box)'
                        else:
                            lat, lon = location.latitude, location.longitude
                            quality = 'dokładne'
                    source = location.raw.get('osm_type', 'API Nominatim')
                    display = location.raw.get('display_name', "").lower()
                    if "poland" in display or "polska" in display:
                        source = 'Polska (Nominatim)'
                    elif "italy" in display or "italia" in display:
                        source = 'Italy (Nominatim)'
                    result = (lat, lon, quality, source)
                    print(f"Nominatim - znaleziono wynik: {result}")
                    geo_cache[variant_key] = result
                    # Zapisz wynik również pod standardowym kluczem
                    key = f"{norm_country}_{norm_postal}"
                    if key != variant_key:
                        geo_cache[key] = result
                    print(f"Zapisano do geo_cache: {variant_key} -> {result}")
                    return result
            except Exception as e:
                print(f"Błąd Nominatim przy zapytaniu '{query_string}': {e}")

        # 4. Dla kodów nie-dwucyfrowych, na końcu sprawdzamy LOOKUP_DICT
        key = f"{norm_country}_{norm_postal}"
        if key in LOOKUP_DICT:
            lat, lon = LOOKUP_DICT[key]
            result = (lat, lon, "lookup (ostatnia opcja)", "lookup")
            print(f"LOOKUP (ostatnia opcja): Znaleziono współrzędne dla {key}: {result}")
            geo_cache[key] = result
            return result
    else:
        # DLA KODÓW NIE-DWUCYFROWYCH - PRIORYTET MA PTV, POTEM NOMINATIM
        print("Priorytetowe użycie PTV API, potem Nominatim dla kodu nie-dwucyfrowego")

        # Generujemy warianty zapytań
        query_variants = generate_query_variants(country, norm_postal, city)
    iso_code = ISO_CODES.get(norm_country, "")

    # Sprawdzenie w cache dla wszystkich wariantów
    for query_string, variant_key in query_variants:
        if variant_key in geo_cache:
            cached = geo_cache[variant_key]
            if cached[0] is not None:
                print(f"Znaleziono wynik w cache dla klucza: {variant_key}: {cached}")
                # Zapisz wynik również pod standardowym kluczem
                key = f"{norm_country}_{norm_postal}"
                if key != variant_key:
                    geo_cache[key] = cached
                return cached

    # Różne ścieżki geokodowania w zależności od długości kodu pocztowego
    if len(norm_postal) != 2:
        # DLA KODÓW NIE-DWUCYFROWYCH - KOLEJNOŚĆ FALLBACK:
        # 2. PTV API (text) -> 3. LOOKUP_DICT -> 4. Nominatim

        # 2. Próba geokodowania przez PTV API (text) - fallback
        print("2. PTV API (structured) nie zwróciło wyniku, próbuję PTV API (text)...")
        query_variants = generate_query_variants(norm_country, norm_postal, city)
        for query_string, variant_key in query_variants:
            ptv_result = ptv_geocode_by_text(query_string, PTV_API_KEY, language="pl", country_code=iso_code)
            if ptv_result[0] is not None:
                print(f"PTV API (text) - wariant '{query_string}' zwrócił wynik: {ptv_result}")
                geo_cache[variant_key] = ptv_result
                # Zapisz wynik również pod standardowym kluczem
                key = f"{norm_country}_{norm_postal}"
                if key != variant_key:
                    geo_cache[key] = ptv_result
                print(f"Zapisano do geo_cache: {variant_key} -> {ptv_result}")
                return ptv_result

        # 3. Sprawdzamy LOOKUP_DICT
        key = f"{norm_country}_{norm_postal}"
        print(f"3. PTV API (text) nie zwróciło wyniku, sprawdzam klucz '{key}' w LOOKUP_DICT")
        if key in LOOKUP_DICT:
            lat, lon = LOOKUP_DICT[key]
            result = (lat, lon, "lookup", "lookup")
            print(f"LOOKUP: Znaleziono współrzędne dla {key}: {result}")
            geo_cache[key] = result
            return result

        # 4. Jeśli PTV i LOOKUP_DICT nie zwróciły wyników, próbujemy przez Nominatim
        print("4. LOOKUP_DICT nie zwrócił wyniku, wywołuję Nominatim...")
        for query_string, variant_key in query_variants:
            print(f"Nominatim - próba zapytania: '{query_string}' (klucz: {variant_key}) z country_codes={iso_code}")
            try:
                extra_params = {}  # Dla kodów nie-dwucyfrowych nie potrzebujemy polygon_geojson
                location = geolocator.geocode(query_string, exactly_one=True, country_codes=iso_code, **extra_params)
                time.sleep(0.1)
                if location:
                    # Sprawdź czy Nominatim zwrócił kod pocztowy
                    returned_postal = location.raw.get('address', {}).get('postcode', '')
                    if not returned_postal:
                        print(f"Nominatim: Brak kodu pocztowego w odpowiedzi")
                        continue
                    
                    # Sprawdź czy pierwsza cyfra kodu pocztowego się zgadza
                    requested_first_digit = ''.join(c for c in norm_postal if c.isdigit())[0]
                    returned_first_digit = ''.join(c for c in returned_postal if c.isdigit())[0]
                    
                    if requested_first_digit != returned_first_digit:
                        print(f"Nominatim: Pierwsza cyfra kodu ({returned_first_digit}) nie zgadza się z szukaną ({requested_first_digit})")
                        continue
                    
                    if "geojson" in location.raw and location.raw["geojson"]:
                        geo = location.raw["geojson"]
                        if geo.get("type") == "Polygon":
                            coords = geo.get("coordinates")[0]
                            lat = sum(point[1] for point in coords) / len(coords)
                            lon = sum(point[0] for point in coords) / len(coords)
                            quality = "centroid (polygon)"
                        else:
                            bb = location.raw.get("boundingbox")
                            if bb:
                                lat = (float(bb[0]) + float(bb[1])) / 2
                                lon = (float(bb[2]) + float(bb[3])) / 2
                                quality = "przybliżone (bounding box)"
                            else:
                                lat, lon = location.latitude, location.longitude
                                quality = "dokładne"
                    else:
                        if hasattr(location, 'raw') and 'boundingbox' in location.raw:
                            bb = location.raw['boundingbox']
                            lat = (float(bb[0]) + float(bb[1])) / 2
                            lon = (float(bb[2]) + float(bb[3])) / 2
                            quality = 'przybliżone (bounding box)'
                        else:
                            lat, lon = location.latitude, location.longitude
                            quality = 'dokładne'
                    source = location.raw.get('osm_type', 'API Nominatim')
                    display = location.raw.get('display_name', "").lower()
                    if "poland" in display or "polska" in display:
                        source = 'Polska (Nominatim)'
                    elif "italy" in display or "italia" in display:
                        source = 'Italy (Nominatim)'
                    result = (lat, lon, quality, source)
                    print(f"Nominatim - znaleziono wynik: {result}")
                    geo_cache[variant_key] = result
                    # Zapisz wynik również pod standardowym kluczem
                    key = f"{norm_country}_{norm_postal}"
                    if key != variant_key:
                        geo_cache[key] = result
                    print(f"Zapisano do geo_cache: {variant_key} -> {result}")
                    return result
            except Exception as e:
                print(f"Błąd Nominatim przy zapytaniu '{query_string}': {e}")


    else:
        # DLA KODÓW DWUCYFROWYCH - NOWA KOLEJNOŚĆ:
        # PTV API -> Nominatim (po LOOKUP_DICT, który był sprawdzany wcześniej)

        # Próba geokodowania przez PTV API
        print("Rozpoczynam geokodowanie przez PTV API dla kodów dwucyfrowych...")
        for query_string, variant_key in query_variants:
            ptv_result = ptv_geocode_by_text(query_string, PTV_API_KEY, language="pl", country_code=iso_code)
            if ptv_result[0] is not None:
                print(f"PTV API - wariant '{query_string}' zwrócił wynik: {ptv_result}")
                geo_cache[variant_key] = ptv_result
                # Zapisz wynik również pod standardowym kluczem
                key = f"{norm_country}_{norm_postal}"
                if key != variant_key:
                    geo_cache[key] = ptv_result
                print(f"Zapisano do geo_cache: {variant_key} -> {ptv_result}")
                return ptv_result

        # Jeśli PTV API nie znalazło lokalizacji, próbujemy Nominatim
        print("PTV API nie zwróciło odpowiedniego wyniku, wywołuję Nominatim...")
        for query_string, variant_key in query_variants:
            print(f"Nominatim - próba zapytania: '{query_string}' (klucz: {variant_key}) z country_codes={iso_code}")
            try:
                extra_params = {"polygon_geojson": 1}  # Dla kodów dwucyfrowych używamy polygon_geojson
                location = geolocator.geocode(query_string, exactly_one=True, country_codes=iso_code, **extra_params)
                time.sleep(0.1)
                if location:
                    if "geojson" in location.raw and location.raw["geojson"]:
                        geo = location.raw["geojson"]
                        if geo.get("type") == "Polygon":
                            coords = geo.get("coordinates")[0]
                            lat = sum(point[1] for point in coords) / len(coords)
                            lon = sum(point[0] for point in coords) / len(coords)
                            quality = "centroid (polygon)"
                        else:
                            bb = location.raw.get("boundingbox")
                            if bb:
                                lat = (float(bb[0]) + float(bb[1])) / 2
                                lon = (float(bb[2]) + float(bb[3])) / 2
                                quality = "przybliżone (bounding box)"
                            else:
                                lat, lon = location.latitude, location.longitude
                                quality = "dokładne"
                    else:
                        if hasattr(location, 'raw') and 'boundingbox' in location.raw:
                            bb = location.raw['boundingbox']
                            lat = (float(bb[0]) + float(bb[1])) / 2
                            lon = (float(bb[2]) + float(bb[3])) / 2
                            quality = 'przybliżone (bounding box)'
                        else:
                            lat, lon = location.latitude, location.longitude
                            quality = 'dokładne'
                    source = location.raw.get('osm_type', 'API Nominatim')
                    display = location.raw.get('display_name', "").lower()
                    if "poland" in display or "polska" in display:
                        source = 'Polska (Nominatim)'
                    elif "italy" in display or "italia" in display:
                        source = 'Italy (Nominatim)'
                    result = (lat, lon, quality, source)
                    print(f"Nominatim - znaleziono wynik: {result}")
                    geo_cache[variant_key] = result
                    # Zapisz wynik również pod standardowym kluczem
                    key = f"{norm_country}_{norm_postal}"
                    if key != variant_key:
                        geo_cache[key] = result
                    print(f"Zapisano do geo_cache: {variant_key} -> {result}")
                    return result
            except Exception as e:
                print(f"Błąd Nominatim przy zapytaniu '{query_string}': {e}")

    # Jeśli żadna metoda nie znalazła lokalizacji
    if query_variants:
        _, variant_key = query_variants[-1]
        result = (None, None, 'nieznane', 'brak danych')
        geo_cache[variant_key] = result
        # Zapisz wynik również pod standardowym kluczem
        key = f"{norm_country}_{norm_postal}"
        if key != variant_key:
            geo_cache[key] = result
        print(f"Brak wyników, zapisano do geo_cache: {variant_key} -> {result}")
        return result

    return (None, None, 'nieznane', 'brak danych')


def get_ungeocoded_locations(df):
    ungeocoded_locations = []
    unique_locations = set()

    print("Zbieranie unikalnych lokalizacji z pliku...")
    for _, row in df.iterrows():
        try:
            # Sprawdź czy to nie jest wiersz nagłówka
            kraj_zal = str(row.get('kraj zaladunku', '')).strip()
            if kraj_zal.lower() in ['kraj zaladunku', 'kraj załadunku', '']:
                continue
                
            # Zachowaj dokładnie taki kod pocztowy, jaki jest w pliku
            kod_zal = row.get('kod pocztowy zaladunku', '')
            if kod_zal is not None:
                kod_zal = str(kod_zal)
            else:
                kod_zal = ''
            miasto_zal = str(row.get('miasto zaladunku', '')).strip()

            kraj_rozl = str(row.get('kraj rozladunku', '')).strip()
            # Zachowaj dokładnie taki kod pocztowy, jaki jest w pliku
            kod_rozl = row.get('kod pocztowy rozladunku', '')
            if kod_rozl is not None:
                kod_rozl = str(kod_rozl)
            else:
                kod_rozl = ''
            miasto_rozl = str(row.get('miasto rozladunku', '')).strip()

            # Dodaj tylko niepuste lokalizacje
            if kraj_zal and kod_zal:
                unique_locations.add((kraj_zal, kod_zal, miasto_zal))
            if kraj_rozl and kod_rozl:
                unique_locations.add((kraj_rozl, kod_rozl, miasto_rozl))
        except Exception as e:
            print(f"Błąd podczas zbierania lokalizacji z wiersza: {str(e)}")
            continue

    print(f"Zebrano {len(unique_locations)} unikalnych lokalizacji do sprawdzenia")

    for loc in unique_locations:
        country, postal_code, city = loc
        # Ujednolicamy city – jeśli None, ustaw pusty ciąg
        if not isinstance(city, str):
            city = ""

        # Normalizacja danych
        norm_country = normalize_country(country)
        # Zachowaj dokładnie taki kod pocztowy, jaki jest w pliku
        norm_postal = str(postal_code) if postal_code is not None else ''

        print(f"Sprawdzanie geokodowania dla: {norm_country}, {norm_postal}, {city}")

        # Sprawdzamy czy lokalizacja jest już zgeokodowana
        is_geocoded = False

        # Standardowy klucz
        std_key = f"{norm_country}_{norm_postal}"

        # Sprawdź w geo_cache
        if std_key in geo_cache:
            cached = geo_cache[std_key]
            if cached[0] is not None and cached[1] is not None:
                print(f"Znaleziono w geo_cache: {std_key} -> {cached}")
                is_geocoded = True

        # Sprawdź w LOOKUP_DICT dla dwucyfrowych kodów
        if not is_geocoded and len(norm_postal) == 2 and std_key in LOOKUP_DICT:
            print(f"Znaleziono w LOOKUP_DICT: {std_key}")
            is_geocoded = True

        # Sprawdź inne warianty kluczy w geo_cache
        if not is_geocoded:
            variants = generate_query_variants(country, postal_code, city)
            for _, variant_key in variants:
                if variant_key in geo_cache:
                    cached = geo_cache[variant_key]
                    if cached[0] is not None and cached[1] is not None:
                        print(f"Znaleziono w geo_cache dla wariantu: {variant_key} -> {cached}")
                        # Synchronizuj ze standardowym kluczem
                        geo_cache[std_key] = cached
                        is_geocoded = True
                        break

        # Jeśli nadal nie znaleziono, spróbuj zgeokodować
        if not is_geocoded:
            # Użyj get_coordinates do próby geokodowania
            try:
                result = get_coordinates(country, postal_code, city)
                if result[0] is not None and result[1] is not None:
                    print(f"Udało się zgeokodować: {country}, {postal_code}, {city} -> {result}")
                    is_geocoded = True
            except Exception as e:
                print(f"Błąd przy próbie geokodowania: {e}")

        # Jeśli nadal nie udało się zgeokodować, dodaj do nierozpoznanych
        if not is_geocoded:
            variants = generate_query_variants(country, postal_code, city)
            ungeocoded_locations.append({
                'country': country,
                'postal_code': postal_code,
                'city': city,
                'key': std_key,
                'query_variants': variants
            })

    print(f"Znaleziono {len(ungeocoded_locations)} nierozpoznanych lokalizacji")

    return ungeocoded_locations


def get_all_locations_status(df):
    """
    Sprawdza status geokodowania dla wszystkich lokalizacji w pliku.
    Zwraca dwie listy: poprawnie i niepoprawnie zgeokodowane lokalizacje.
    """
    global GEOCODING_CURRENT, GEOCODING_TOTAL
    
    ungeocoded_locations = []
    geocoded_locations = []
    unique_locations = set()

    print("Zbieranie unikalnych lokalizacji z pliku...")
    for _, row in df.iterrows():
        try:
            # Sprawdź czy to nie jest wiersz nagłówka
            kraj_zal = str(row.get('kraj zaladunku', '')).strip()
            if kraj_zal.lower() in ['kraj zaladunku', 'kraj załadunku', '']:
                continue
                
            kod_zal = str(row.get('kod pocztowy zaladunku', '')).strip()
            miasto_zal = str(row.get('miasto zaladunku', '')).strip()

            kraj_rozl = str(row.get('kraj rozladunku', '')).strip()
            kod_rozl = str(row.get('kod pocztowy rozladunku', '')).strip()
            miasto_rozl = str(row.get('miasto rozladunku', '')).strip()

            # Dodaj tylko niepuste lokalizacje
            if kraj_zal and kod_zal:
                unique_locations.add((kraj_zal, kod_zal, miasto_zal))
            if kraj_rozl and kod_rozl:
                unique_locations.add((kraj_rozl, kod_rozl, miasto_rozl))
        except Exception as e:
            print(f"Błąd podczas zbierania lokalizacji z wiersza: {str(e)}")
            continue

    print(f"Zebrano {len(unique_locations)} unikalnych lokalizacji do sprawdzenia")

    # Inicjalizuj zmienne postępu
    GEOCODING_TOTAL = len(unique_locations)
    GEOCODING_CURRENT = 0

    location_id = 1  # Dodajemy licznik dla ID lokalizacji
    for loc in unique_locations:
        country, postal_code, city = loc
        # Ujednolicamy city – jeśli None, ustaw pusty ciąg
        if not isinstance(city, str):
            city = ""

        # Normalizacja danych
        norm_country = normalize_country(country)
        norm_postal = str(postal_code).strip()

        print(f"Sprawdzanie geokodowania dla: {norm_country}, {norm_postal}, {city}")

        # Sprawdzamy czy lokalizacja jest już zgeokodowana
        is_geocoded = False
        coords = None

        # Standardowy klucz
        std_key = f"{norm_country}_{norm_postal}"

        # Sprawdź w geo_cache
        if std_key in geo_cache:
            cached = geo_cache[std_key]
            if cached[0] is not None and cached[1] is not None:
                print(f"Znaleziono w geo_cache: {std_key} -> {cached}")
                is_geocoded = True
                coords = f"{cached[0]},{cached[1]}"

        # Sprawdź inne warianty kluczy w geo_cache
        if not is_geocoded:
            variants = generate_query_variants(country, postal_code, city)
            for _, variant_key in variants:
                if variant_key in geo_cache:
                    cached = geo_cache[variant_key]
                    if cached[0] is not None and cached[1] is not None:
                        print(f"Znaleziono w geo_cache dla wariantu: {variant_key} -> {cached}")
                        # Synchronizuj ze standardowym kluczem
                        geo_cache[std_key] = cached
                        is_geocoded = True
                        coords = f"{cached[0]},{cached[1]}"
                        break

        # Jeśli nadal nie znaleziono, spróbuj zgeokodować
        if not is_geocoded:
            # Użyj get_coordinates do próby geokodowania
            try:
                result = get_coordinates(country, postal_code, city)
                if result[0] is not None and result[1] is not None:
                    print(f"Udało się zgeokodować: {country}, {postal_code}, {city} -> {result}")
                    is_geocoded = True
                    coords = f"{result[0]},{result[1]}"
            except Exception as e:
                print(f"Błąd przy próbie geokodowania: {e}")

        # Sprawdź weryfikację zgodności miasta i kodu pocztowego
        verification_status = "N/A"
        # Sprawdź czy miasto jest prawidłowe (nie puste, nie "nan" itp.)
        is_city_valid = (city and 
                        isinstance(city, str) and 
                        city.strip() and 
                        not pd.isna(city) and
                        str(city).lower().strip() not in ['nan', 'none', 'null'])
        
        if is_geocoded and is_city_valid:
            try:
                verification_result = verify_city_postal_code_match(country, postal_code, city)
                verification_status = "TAK" if verification_result.get('is_match', True) else "NIE"
                print(f"Weryfikacja dla {country}, {postal_code}, {city}: {verification_status}")
            except Exception as e:
                print(f"Błąd weryfikacji dla {country}, {postal_code}, {city}: {e}")
                verification_status = "BŁĄD"
        elif is_geocoded and not is_city_valid:
            # Lokalizacja jest zgeokodowana ale nie ma prawidłowego miasta
            verification_status = "N/A"
            print(f"Lokalizacja {country}, {postal_code} jest zgeokodowana, ale miasto ('{city}') nie jest prawidłowe - pomijam weryfikację")

        # Dodaj lokalizację do odpowiedniej listy
        location_data = {
            'id': location_id,  # Dodajemy ID dla każdej lokalizacji
            'country': country,
            'postal_code': postal_code,
            'city': city,
            'key': std_key,
            'verification_status': verification_status,  # Dodajemy status weryfikacji
        }
        location_id += 1  # Inkrementujemy ID

        if is_geocoded:
            location_data['coords'] = coords
            geocoded_locations.append(location_data)
        else:
            variants = generate_query_variants(country, postal_code, city)
            location_data['query_variants'] = variants
            ungeocoded_locations.append(location_data)

        # Aktualizuj postęp
        GEOCODING_CURRENT += 1

    print(f"Znaleziono {len(ungeocoded_locations)} nierozpoznanych lokalizacji")
    print(f"Znaleziono {len(geocoded_locations)} rozpoznanych lokalizacji")

    return {
        'correct_locations': geocoded_locations,
        'locations': ungeocoded_locations
    }


def haversine(coord1, coord2):
    if None in coord1 or None in coord2:
        return None
    R = 6371
    lat1, lon1 = math.radians(coord1[0]), math.radians(coord1[1])
    lat2, lon2 = math.radians(coord2[0]), math.radians(coord2[1])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


def safe_float(value):
    if pd.isna(value) or str(value).strip().lower() in ['nan', 'none', '']:
        return None
    try:
        return float(str(value).strip().replace(',', '.'))
    except (ValueError, TypeError):
        return None


def format_currency(value):
    if value is None or pd.isna(value):
        return None
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return None


def select_best_rate(row, rate_columns):
    for col in rate_columns:
        rate = safe_float(row.get(col))
        if rate is not None:
            # Zwróć zarówno stawkę jak i okres
            period = col.split('_')[-1] if '_' in col else '3m'  # Domyślnie 3m
            return {'rate': rate, 'period': period}
    return None


def calculate_fracht(distance, rate):
    if distance is None or rate is None:
        return None
    try:
        return distance * rate
    except TypeError:
        return None

def get_podlot(rates, region_rates=None):
    """
    Pobiera podlot z danych historycznych z fallbackiem do regionów lub zwraca wartość domyślną 100 km
    
    Hierarchia:
    1. rates.get('podlot_historyczny') - konkretne dane historyczne
    2. rates.get('podlot_sredni_wazony') - średni ważony podlot
    3. region_rates.get('region_podlot') - regionalny podlot
    4. 100 km - wartość domyślna
    
    Returns:
        tuple: (podlot_value, source_info)
    """
    # Sprawdź najpierw konkretny podlot historyczny
    podlot = rates.get('podlot_historyczny')
    if podlot is not None:
        return podlot, 'historyczny'
    
    # Sprawdź średni ważony podlot
    podlot_sredni = rates.get('podlot_sredni_wazony')
    if podlot_sredni is not None:
        return podlot_sredni, 'średni ważony'
    
    # Sprawdź regionalny podlot jako fallback
    if region_rates and region_rates.get('region_podlot') is not None:
        return region_rates.get('region_podlot'), 'regionalny'
    
    # Wartość domyślna
    return 100, 'domyślny'


# Globalna zmienna dla macierzy marży
MARGIN_MATRIX = None
CURRENT_MATRIX_FILE = 'Matrix.xlsx'  # Domyślny plik matrixa


def load_margin_matrix(matrix_file='Matrix.xlsx'):
    """
    Wczytuje macierz marży z podanego pliku
    
    Args:
        matrix_file (str): Nazwa pliku Excel z macierzą marży
    
    Returns:
        pandas.DataFrame: Macierz marży z regionami jako indeksy (załadunek) i kolumny (rozładunek)
    """
    global MARGIN_MATRIX, CURRENT_MATRIX_FILE
    
    try:
        # Wczytaj plik Excel bez nagłówków
        df = pd.read_excel(matrix_file, sheet_name='Sheet1', header=None)
        
        if df.empty or df.shape[0] < 3 or df.shape[1] < 3:
            logger.error(f"Macierz marży z pliku {matrix_file} jest pusta lub ma nieprawidłowe wymiary")
            return None
            
        # Wyodrębnij nazwy regionów z pierwszego wiersza (rozładunek - kolumny)
        regions_columns = df.iloc[0, 2:].dropna().tolist()
        
        # Wyodrębnij nazwy regionów z pierwszej kolumny (załadunek - wiersze)  
        regions_rows = df.iloc[2:, 0].dropna().tolist()
        
        # Utwórz macierz marży z danymi liczbowymi
        margin_data = df.iloc[2:len(regions_rows)+2, 2:len(regions_columns)+2]
        
        # Ustaw indeksy i kolumny
        MARGIN_MATRIX = pd.DataFrame(
            margin_data.values,
            index=regions_rows,
            columns=regions_columns
        )
        
        # Konwertuj wartości na liczby, zastępując NaN przez None
        MARGIN_MATRIX = MARGIN_MATRIX.apply(pd.to_numeric, errors='coerce')
        
        # Zaktualizuj aktualny plik
        CURRENT_MATRIX_FILE = matrix_file
        
        logger.info(f"Wczytano macierz marży z {matrix_file}: {MARGIN_MATRIX.shape[0]}x{MARGIN_MATRIX.shape[1]} regionów")
        return MARGIN_MATRIX
        
    except Exception as e:
        logger.error(f"Błąd podczas wczytywania macierzy marży z {matrix_file}: {e}")
        return None


def set_margin_matrix(matrix_type='klient'):
    """
    Ustawia macierz marży na podstawie typu
    
    Args:
        matrix_type (str): 'klient' dla Matrix.xlsx lub 'targi' dla Matrix_Targi.xlsx
    """
    matrix_file = 'Matrix.xlsx' if matrix_type == 'klient' else 'Matrix_Targi.xlsx'
    return load_margin_matrix(matrix_file)


def get_margin_matrix_info():
    """
    Zwraca informacje o aktualnie załadowanej macierzy marży
    """
    global CURRENT_MATRIX_FILE
    matrix_name = "Matrix Klient" if CURRENT_MATRIX_FILE == 'Matrix.xlsx' else "Matrix Targi"
    return matrix_name, CURRENT_MATRIX_FILE


def get_margin_for_route(loading_region, unloading_region):
    """
    Pobiera marżę dla konkretnej relacji z macierzy marży
    
    Args:
        loading_region (str): Region załadunku
        unloading_region (str): Region rozładunku
        
    Returns:
        float or None: Marża w Euro lub None jeśli nie znaleziono
    """
    global MARGIN_MATRIX
    
    # Sprawdź czy macierz została wczytana
    if MARGIN_MATRIX is None:
        load_margin_matrix()
        
    if MARGIN_MATRIX is None:
        return None
        
    try:
        # Sprawdź czy regiony istnieją w macierzy
        if loading_region not in MARGIN_MATRIX.index:
            logger.warning(f"Region załadunku '{loading_region}' nie znajduje się w macierzy marży")
            return None
            
        if unloading_region not in MARGIN_MATRIX.columns:
            logger.warning(f"Region rozładunku '{unloading_region}' nie znajduje się w macierzy marży")
            return None
            
        # Pobierz marżę
        margin = MARGIN_MATRIX.loc[loading_region, unloading_region]
        
        # Sprawdź czy wartość nie jest NaN
        if pd.isna(margin):
            return None
            
        return float(margin)
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania marży dla relacji {loading_region} -> {unloading_region}: {e}")
        return None


def get_transit_time_from_row(row):
    """
    Pobiera wartość transit time z wiersza danych
    
    Args:
        row: Pandas Series - wiersz danych
        
    Returns:
        float or None: Wartość transit time lub None jeśli nie ma lub nieprawidłowa
    """
    # Kolumna 'transit time' zawsze będzie obecna, ale może być pusta
    if 'transit time' not in row.index:
        return None
        
    value = row['transit time']
    
    # Sprawdź czy wartość nie jest pusta ani NaN
    if pd.notna(value) and value != '' and str(value).strip() != '':
        try:
            # Konwertuj na float
            transit_time = float(value)
            # Sprawdź czy wartość jest rozsądna (między 0.1 a 30 dni)
            if 0.1 <= transit_time <= 30:
                return transit_time
            else:
                print(f"UWAGA: Wartość transit time poza zakresem (0.1-30): {transit_time}")
                return None
        except (ValueError, TypeError):
            print(f"UWAGA: Nie można skonwertować transit time na liczbę: {value}")
            return None
    
    return None


def calculate_driver_days(distance):
    """
    Oblicza liczbę dni kierowcy na podstawie dystansu
    
    Args:
        distance (float): Dystans w kilometrach
        
    Returns:
        float or None: Liczba dni kierowcy lub None jeśli dystans jest None
    """
    if distance is None:
        return None
    
    if distance <= 350:
        return 1
    elif 351 <= distance <= 500:
        return 1.25
    elif 501 <= distance <= 700:
        return 1.5
    elif 701 <= distance <= 1100:
        return 2
    elif 1101 <= distance <= 1700:
        return 3
    elif 1701 <= distance <= 2300:
        return 4
    elif 2301 <= distance <= 2900:
        return 5
    elif 2901 <= distance <= 3500:
        return 6
    else:
        # Dla dystansów powyżej 3500 km, oblicz dni proporcjonalnie
        return math.ceil(distance / 600)  # Średnio 600 km dziennie dla długich tras


def calculate_expected_profit(loading_region, unloading_region, driver_days):
    """
    Oblicza oczekiwany zysk na podstawie marży z macierzy i liczby dni kierowcy
    
    Args:
        loading_region (str): Region załadunku
        unloading_region (str): Region rozładunku
        driver_days (float): Liczba dni kierowcy
        
    Returns:
        tuple: (oczekiwany_zysk, marża_jednostkowa, źródło_info)
    """
    if not loading_region or not unloading_region or driver_days is None:
        return None, None, "Brak danych regionalnych lub dni kierowcy"
        
    # Pobierz marżę jednostkową
    unit_margin = get_margin_for_route(loading_region, unloading_region)
    
    if unit_margin is None:
        return None, None, f"Brak marży dla relacji {loading_region} -> {unloading_region}"
        
    # Oblicz oczekiwany zysk
    expected_profit = unit_margin * driver_days
    
    # Dodaj informację o typie matrixa
    matrix_name, _ = get_margin_matrix_info()
    
    return expected_profit, unit_margin, f"{matrix_name}: {unit_margin}€ × {driver_days} dni"

def calculate_podlot_toll(podlot, toll_rate_per_km=0.30, fuel_cost_per_km=None):
    """
    Oblicza całkowity koszt podlotu (opłaty drogowe + paliwo)
    
    Args:
        podlot: Dystans podlotu w km
        toll_rate_per_km: Stawka opłat drogowych za km (domyślnie 0.30 EUR/km)
        fuel_cost_per_km: Stawka paliwa za km (jeśli None, używa DEFAULT_FUEL_COST)
    
    Returns:
        float: Całkowity koszt podlotu (opłaty + paliwo)
    """
    if podlot is None:
        return None
    
    # Użyj przekazanej stawki paliwa lub domyślnej
    fuel_rate = fuel_cost_per_km if fuel_cost_per_km is not None else DEFAULT_FUEL_COST
    
    # Całkowity koszt = opłaty drogowe + paliwo
    total_rate_per_km = toll_rate_per_km + fuel_rate
    
    return podlot * total_rate_per_km

def calculate_toll_per_km(road_toll, distance):
    """Oblicza opłaty drogowe na kilometr"""
    return road_toll / distance if distance and distance > 0 else 0

def calculate_total_costs(cost_list):
    """Oblicza sumę kosztów, pomijając wartości None"""
    total = 0
    for cost in cost_list:
        if cost is not None:
            total += cost
    return total


def format_toll_details(toll_details, road_toll, other_toll, special_systems=None):
    """
    Formatuje szczegóły opłat drogowych jako tekst
    
    Args:
        toll_details (dict): Słownik z kosztami według krajów
        road_toll (float): Standardowe opłaty drogowe
        other_toll (float): Opłaty specjalne (tunele/mosty/promy)
        special_systems (list): Lista systemów specjalnych z nazwami
        
    Returns:
        str: Sformatowany tekst z szczegółami opłat
    """
    print(f"DEBUG format_toll_details: toll_details={toll_details}, road_toll={road_toll}, other_toll={other_toll}, special_systems={special_systems}")
    
    ##toll_text = "Opłaty drogowe według krajów:\n"
    toll_text = "\n".join([f"{country}: {format_currency(cost)}€" for country, cost in toll_details.items()])
    ##toll_text += f"\n\nOpłaty drogowe: {format_currency(road_toll)}€"
    

    ##print(f"DEBUG format_toll_details: initial toll_text='{toll_text}'")

    print(f"DEBUG format_toll_details: initial toll_text='{toll_text}'")

    
    if other_toll > 0:
        print(f"DEBUG format_toll_details: other_toll > 0, checking special_systems")
        if special_systems and len(special_systems) > 0:
            print(f"DEBUG format_toll_details: special_systems not empty: {special_systems}")
            # Wyświetl konkretne nazwy systemów
            systems_text = []
            for system in special_systems:
                name = system.get('name')
                cost = system.get('cost', 0)
                print(f"DEBUG format_toll_details: processing system: name='{name}', cost={cost}")
                # Dodaj tylko systemy z rzeczywistymi nazwami
                if name:
                    systems_text.append(f"\n{name}: {format_currency(cost)}€")
            
            print(f"DEBUG format_toll_details: systems_text={systems_text}")
            if systems_text:
                toll_text += f"\n" + "\n".join(systems_text)
            else:
                toll_text += f"\nDodatkowe opłaty (tunele/mosty/promy): {format_currency(other_toll)}€"
        else:
            print(f"DEBUG format_toll_details: special_systems empty or None, using fallback")
            # Fallback do oryginalnego tekstu jeśli brak szczegółów
            toll_text += f"\nDodatkowe opłaty (tunele/mosty/promy): {format_currency(other_toll)}€"
    
    print(f"DEBUG format_toll_details: final toll_text='{toll_text}'")
    return toll_text

def get_best_rates(rates, region_rates):
    """Pobiera najlepsze stawki giełdowe i klienckie z fallbackiem do regionów"""
    gielda_rate_info = select_best_rate(rates, ['gielda_stawka_3m', 'gielda_stawka_6m', 'gielda_stawka_12m'])
    hist_rate_info = select_best_rate(rates, ['hist_stawka_3m', 'hist_stawka_6m', 'hist_stawka_12m'])
    
    if gielda_rate_info is None:
        gielda_rate_info = select_best_rate(region_rates, ['region_gielda_stawka_3m',
                                                      'region_gielda_stawka_6m',
                                                      'region_gielda_stawka_12m'])
    if hist_rate_info is None:
        hist_rate_info = select_best_rate(region_rates, ['region_klient_stawka_3m',
                                                    'region_klient_stawka_6m',
                                                    'region_klient_stawka_12m'])
    
    return {
        'gielda_rate': gielda_rate_info['rate'] if gielda_rate_info else None,
        'hist_rate': hist_rate_info['rate'] if hist_rate_info else None,
        'gielda_period': gielda_rate_info['period'] if gielda_rate_info else None,
        'hist_period': hist_rate_info['period'] if hist_rate_info else None,
        'gielda_rate_info': gielda_rate_info,
        'hist_rate_info': hist_rate_info
    }

def format_coordinates(lat, lon, default_text="Brak danych"):
    """Formatuje współrzędne jako string"""
    return f"{lat}, {lon}" if None not in (lat, lon) else default_text


def get_toll_cost(
    coord_from, coord_to,
    loading_country=None, unloading_country=None,
    start_time="2023-08-29T10:00:00.000Z",
    routing_mode=DEFAULT_ROUTING_MODE,
    avoid_switzerland=True
):
    """
    Pobiera koszty drogowe dla trasy między dwoma punktami.
    Używa PTVRouteManager do pobrania i przetworzenia danych.
    """
    # Jeśli trasa jest do/ze Szwajcarii, nie unikamy Szwajcarii
    if loading_country and unloading_country:
        if is_route_to_or_from_switzerland(loading_country, unloading_country):
            avoid_switzerland = False
    
    result = ptv_manager.get_route_distance(coord_from, coord_to, avoid_switzerland, routing_mode)
    if result and 'toll_cost' in result:
        return result['toll_cost']
    return None


def get_route_distance(coord_from, coord_to, avoid_switzerland=True, routing_mode=DEFAULT_ROUTING_MODE):
    result = ptv_manager.get_route_distance(coord_from, coord_to, avoid_switzerland, routing_mode)
    return result


def verify_city_postal_code_match(country, postal_code, city, threshold_km=100):
    print(f"Wywołano verify_city_postal_code_match z parametrami: kraj={country}, kod={postal_code}, miasto={city}")

    result = {
        'is_match': True,
        'distance_km': None,
        'postal_coords': None,
        'city_coords': None,
        'error': None,
        'lookup_coords': None,  # Nowe pole dla współrzędnych z LOOKUP_DICT
        'suggested_coords': None,  # Nowe pole dla sugerowanych współrzędnych
        'city_name': city,  # Dodaję nazwę miasta
        'postal_name': postal_code,  # Dodaję kod pocztowy
        'distance': None  # Dodaję pole distance dla odległości
    }

    # Bardziej dokładne sprawdzenie, czy miasto jest puste
    if (city is None or 
        not isinstance(city, str) or 
        city.strip() == "" or 
        pd.isna(city) or 
        str(city).lower().strip() in ['nan', 'none', 'null']):
        print(f"Miasto jest puste lub nieprawidłowe ('{city}'), zwracam wynik bez geokodowania")
        result['error'] = "Brak miasta"
        return result

    if postal_code is None or str(postal_code).strip() == "" or pd.isna(postal_code):
        print(f"Kod pocztowy jest pusty, zwracam wynik bez geokodowania")
        result['error'] = "Brak kodu pocztowego"
        return result

    norm_country = normalize_country(country)

    # Dodaj mapowanie kodów ISO dla każdego kraju
    ISO_CODES = {
        "Poland": "pl",
        "Germany": "de",
        "France": "fr",
        "Italy": "it",
        "Spain": "es",
        "Netherlands": "nl",
        "Belgium": "be",
        "Czech Republic": "cz",
        "Austria": "at",
        "Slovakia": "sk",
        "Slovenia": "si",
        "Hungary": "hu",
        "Portugal": "pt",
        "Greece": "gr",
        "Switzerland": "ch",
        "Sweden": "se",
        "Finland": "fi",
        "Norway": "no",
        "Denmark": "dk",
        "Bulgaria": "bg",
        "Estonia": "ee",
        "Croatia": "hr",
        "Ireland": "ie",
        "Lithuania": "lt",
        "Latvia": "lv",
        "Romania": "ro",
        "United Kingdom": "gb"
    }

    iso_code = ISO_CODES.get(norm_country, "")

    # ETAP 1: Geokodowanie kodu pocztowego
    # A. Najpierw sprawdzamy w geo_cache
    postal_key = f"{norm_country}_{postal_code}"
    if postal_key in geo_cache:
        postal_coords = geo_cache[postal_key][:2]
        postal_quality = geo_cache[postal_key][2] if len(geo_cache[postal_key]) > 2 else 'nieznane'
        postal_source = geo_cache[postal_key][3] if len(geo_cache[postal_key]) > 3 else 'cache'
        print(f"Znaleziono współrzędne dla kodu {postal_code} w cache: {postal_coords}")
    else:
        # B. Sprawdzamy w LOOKUP_DICT
        if postal_key in LOOKUP_DICT:
            postal_coords = LOOKUP_DICT[postal_key]
            postal_quality = 'lookup'
            postal_source = 'LOOKUP_DICT'
            # Zapisz do cache dla przyszłych zapytań
            geo_cache[postal_key] = (*postal_coords, postal_quality, postal_source)
            print(f"Znaleziono współrzędne dla kodu {postal_code} w LOOKUP_DICT: {postal_coords}")
        else:
            # Sprawdź dla dwucyfrowego prefiksu
            postal_prefix = str(postal_code).strip()[:2].zfill(2)
            prefix_key = f"{norm_country}_{postal_prefix}"
            if prefix_key in LOOKUP_DICT:
                postal_coords = LOOKUP_DICT[prefix_key]
                postal_quality = 'lookup (prefiks)'
                postal_source = 'LOOKUP_DICT'
                # Zapisz do cache dla przyszłych zapytań
                geo_cache[postal_key] = (*postal_coords, postal_quality, postal_source)
                print(f"Znaleziono współrzędne dla prefiksu kodu {postal_prefix} w LOOKUP_DICT: {postal_coords}")
            else:
                # C. Próbujemy PTV API
                try:
                    query_string = f"{postal_code}, {norm_country}"
                    print(f"Zapytanie PTV API o kod pocztowy: {query_string}")
                    ptv_result = ptv_geocode_by_text(query_string, PTV_API_KEY, language="pl", country_code=iso_code)
                    if ptv_result[0] is not None:
                        postal_coords = ptv_result[:2]
                        postal_quality = ptv_result[2] if len(ptv_result) > 2 else 'PTV API'
                        postal_source = ptv_result[3] if len(ptv_result) > 3 else 'PTV API'
                        # Zapisz do cache
                        geo_cache[postal_key] = (*postal_coords, postal_quality, postal_source)
                        print(f"Znaleziono współrzędne dla kodu {postal_code} w PTV API: {postal_coords}")
                    else:
                        # D. Ostatnia próba - Nominatim
                        query_string = f"{postal_code}, {norm_country}"
                        print(f"Zapytanie Nominatim o kod pocztowy: {query_string}")
                        location = geolocator.geocode(query_string, exactly_one=True, country_codes=iso_code)
                        time.sleep(0.1)
                        if location:
                            postal_coords = (location.latitude, location.longitude)
                            postal_quality = 'Nominatim'
                            postal_source = 'Nominatim'
                            geo_cache[postal_key] = (*postal_coords, postal_quality, postal_source)
                            print(f"Znaleziono współrzędne dla kodu {postal_code} w Nominatim: {postal_coords}")
                        else:
                            print(f"Nie znaleziono lokalizacji dla kodu pocztowego {postal_code}")
                            result['error'] = f"Nie znaleziono lokalizacji dla kodu pocztowego {postal_code}"
                            return result
                except Exception as e:
                    print(f"Błąd geokodowania kodu pocztowego: {str(e)}")
                    result['error'] = f"Błąd geokodowania kodu pocztowego: {str(e)}"
                    return result

    # ETAP 2: Geokodowanie miasta
    # A. Najpierw sprawdzamy w geo_cache
    city_key = f"{norm_country}_{city}"
    if city_key in geo_cache:
        city_coords = geo_cache[city_key][:2]
        city_quality = geo_cache[city_key][2] if len(geo_cache[city_key]) > 2 else 'nieznane'
        city_source = geo_cache[city_key][3] if len(geo_cache[city_key]) > 3 else 'cache'
        print(f"Znaleziono współrzędne dla miasta {city} w cache: {city_coords}")
    else:
        # B. Sprawdzamy w LOOKUP_DICT (pełna nazwa miasta jako klucz)
        city_lookup_key = f"{norm_country}_{city.strip().lower()}"
        # Sprawdź również z wyczyszczoną nazwą miasta
        clean_city = clean_text(city)
        city_clean_lookup_key = f"{norm_country}_{clean_city}"

        found_in_lookup = False
        # Sprawdź dokładne dopasowanie
        for lookup_key in [city_lookup_key, city_clean_lookup_key]:
            if lookup_key in LOOKUP_DICT:
                city_coords = LOOKUP_DICT[lookup_key]
                city_quality = 'lookup'
                city_source = 'LOOKUP_DICT'
                # Zapisz do cache dla przyszłych zapytań
                geo_cache[city_key] = (*city_coords, city_quality, city_source)
                print(f"Znaleziono współrzędne dla miasta {city} w LOOKUP_DICT: {city_coords}")
                found_in_lookup = True
                break

        # Jeśli nie znaleziono dokładnego dopasowania, sprawdź częściowe
        if not found_in_lookup:
            # Przeszukaj klucze zawierające nazwę miasta
            for db_key in LOOKUP_DICT:
                if db_key.startswith(f"{norm_country}_") and (clean_city in db_key or city.lower() in db_key.lower()):
                    city_coords = LOOKUP_DICT[db_key]
                    city_quality = 'lookup (częściowe dopasowanie)'
                    city_source = 'LOOKUP_DICT'
                    # Zapisz do cache dla przyszłych zapytań
                    geo_cache[city_key] = (*city_coords, city_quality, city_source)
                    print(f"Znaleziono częściowe dopasowanie dla miasta {city} w LOOKUP_DICT: {city_coords}")
                    found_in_lookup = True
                    break

        if not found_in_lookup:
            # C. Próbujemy PTV API
            try:
                query_string = f"{city}, {norm_country}"
                #print(f"Zapytanie PTV API o miasto: {query_string} z region={iso_code}")
                ptv_result = ptv_geocode_by_text(query_string, PTV_API_KEY, language="pl", country_code=iso_code)
                if ptv_result[0] is not None:
                    city_coords = ptv_result[:2]
                    city_quality = ptv_result[2] if len(ptv_result) > 2 else 'PTV API'
                    city_source = ptv_result[3] if len(ptv_result) > 3 else 'PTV API'
                    # Zapisz do cache
                    geo_cache[city_key] = (*city_coords, city_quality, city_source)
                    print(f"Znaleziono współrzędne dla miasta {city} w PTV API: {city_coords}")
                else:
                    # D. Ostatnia próba - Nominatim
                    query_string = f"{city}, {norm_country}"
                    print(f"Zapytanie Nominatim o miasto: {query_string} z country_codes={iso_code}")
                    location = geolocator.geocode(query_string, exactly_one=True, country_codes=iso_code)
                    time.sleep(0.1)
                    if location:
                        city_coords = (location.latitude, location.longitude)
                        city_quality = 'Nominatim'
                        city_source = 'Nominatim'
                        geo_cache[city_key] = (*city_coords, city_quality, city_source)
                        print(f"Znaleziono współrzędne dla miasta {city} w Nominatim: {city_coords}")
                    else:
                        print(f"Nie znaleziono lokalizacji dla miasta {city}")
                        result['error'] = f"Nie znaleziono lokalizacji dla miasta {city}"
                        return result
            except Exception as e:
                print(f"Błąd geokodowania miasta: {str(e)}")
                result['error'] = f"Błąd geokodowania miasta: {str(e)}"
                return result

    # Obliczenie odległości między kodem pocztowym a miastem
    distance_km = haversine(postal_coords, city_coords)
    result['postal_coords'] = postal_coords
    result['city_coords'] = city_coords
    result['distance_km'] = distance_km
    result['distance'] = distance_km  # Ustawiam również pole distance
    print(f"Obliczona odległość między kodem pocztowym a miastem: {distance_km} km")

    # DODATKOWE SPRAWDZENIE: Czy miasto z kodu pocztowego zgadza się z podanym miastem
    postal_city_name = None
    
    # Spróbuj uzyskać nazwę miasta z geokodowania kodu pocztowego
    if postal_source == 'PTV API':
        # Dla PTV API, spróbuj ponownie geokodować kod pocztowy i sprawdź nazwę miasta
        try:
            query_string = f"{postal_code}, {norm_country}"
            ptv_result = ptv_geocode_by_text(query_string, PTV_API_KEY, language="pl", country_code=iso_code)
            if ptv_result[0] is not None:
                # Tutaj moglibyśmy sprawdzić nazwę miasta z odpowiedzi PTV, ale to wymagałoby dodatkowych kroków
                pass
        except:
            pass
    
    # Sprawdź odległość i zgodność nazw
    city_name_mismatch = False
    
    # Sprawdź czy odległość przekracza próg
    if distance_km is not None and distance_km > threshold_km:
        result['is_match'] = False
        print(f"Odległość przekracza próg {threshold_km} km - weryfikacja negatywna")
        city_name_mismatch = True
    else:
        print(f"Odległość {distance_km} km jest w granicach progu {threshold_km} km")
        
        # DODATKOWE SPRAWDZENIE: Czy to są różne miasta w tym samym kraju
        # Jeśli odległość jest mała (< 50km), ale to mogą być różne miejscowości
        if distance_km > 5:  # Jeśli odległość > 5km, sprawdź czy to mogą być różne miasta
            print(f"Odległość {distance_km} km > 5km - sprawdzam czy to różne miejscowości")
            
            # Sprawdź czy nazwisko miasta pasuje do geokodowania kodu pocztowego
            # To jest uproszczone sprawdzenie - w przyszłości można to rozszerzyć
            # Na razie oznaczamy jako potencjalny problem jeśli odległość > 20km
            if distance_km > 20:
                print(f"Odległość {distance_km} km > 20km - prawdopodobnie różne miejscowości")
                result['is_match'] = False
                city_name_mismatch = True
            else:
                print(f"Odległość {distance_km} km <= 20km - prawdopodobnie ta sama okolica")
        else:
            print(f"Odległość {distance_km} km <= 5km - bardzo blisko, prawdopodobnie zgodne")

    # Jeśli wystąpił problem z dopasowaniem, wykonaj dodatkowe sprawdzenie
    if not result['is_match'] or city_name_mismatch:
        print(f"Wykonuję dodatkowe sprawdzenie wiarygodności")

        # Ocena jakości i wiarygodności geokodowania
        postal_reliability = evaluate_geocoding_reliability(postal_quality, postal_source)
        city_reliability = evaluate_geocoding_reliability(city_quality, city_source)

        print(f"Wiarygodność kodu pocztowego: {postal_reliability}, Wiarygodność miasta: {city_reliability}")

        # Wybór najbardziej wiarygodnych współrzędnych
        if postal_reliability >= city_reliability:
            result['suggested_coords'] = postal_coords
            print(f"Wybieram współrzędne kodu pocztowego jako sugerowane: {postal_coords}")
        else:
            result['suggested_coords'] = city_coords
            print(f"Wybieram współrzędne miasta jako sugerowane: {city_coords}")
    else:
        print(f"Weryfikacja pozytywna - miasto i kod pocztowy są zgodne")

    return result


# Funkcja pomocnicza do oceny wiarygodności geokodowania
def evaluate_geocoding_reliability(quality, source):
    reliability = 50  # Wyjściowa średnia ocena

    # Ocena na podstawie jakości
    if quality in ['lookup', 'LOOKUP_DICT']:
        reliability += 30  # Najwyższa wiarygodność - dane z ręcznie zweryfikowanej bazy
    elif quality in ['lookup (prefiks)', 'lookup (częściowe dopasowanie)']:
        reliability += 20  # Wysoka wiarygodność, ale częściowe dopasowanie
    elif 'PTV API' in quality:
        reliability += 15  # Dobra wiarygodność - komercyjne API
    elif 'Nominatim' in quality:
        reliability += 5  # Niższa wiarygodność - darmowe API

    # Ocena na podstawie źródła
    if source in ['LOOKUP_DICT', 'lookup']:
        reliability += 20  # Najlepsze źródło
    elif source == 'PTV API':
        reliability += 15  # Dobre źródło
    elif source == 'Nominatim':
        reliability += 5  # Mniej wiarygodne źródło

    # Dodatkowe czynniki można dodać w przyszłości

    return min(100, max(0, reliability))  # Wartość między 0 a 100


def get_all_rates(lc, lp, uc, up, lc_coords, uc_coords):
    try:
        historical_rates_df = pd.read_excel("historical_rates.xlsx",
                                            dtype={'kod pocztowy zaladunku': str, 'kod pocztowy rozladunku': str})
        historical_rates_gielda_df = pd.read_excel("historical_rates_gielda.xlsx",
                                                   dtype={'kod pocztowy zaladunku': str, 'kod pocztowy rozladunku': str})
    except Exception as e:
        print(f"Błąd wczytywania danych historycznych: {e}")
        historical_rates_df = pd.DataFrame()
        historical_rates_gielda_df = pd.DataFrame()

    norm_lc = normalize_country(lc).strip()
    norm_uc = normalize_country(uc).strip()

    # Określamy, czy kod pocztowy ma conajmniej dwie cyfry
    lp_has_two_digits = len(str(lp).strip()) >= 2
    up_has_two_digits = len(str(up).strip()) >= 2

    # Przygotowujemy kody do dopasowania
    norm_lp = str(lp).strip()[:2].zfill(2) if lp_has_two_digits else str(lp).strip()[0]
    norm_up = str(up).strip()[:2].zfill(2) if up_has_two_digits else str(up).strip()[0]

    # Pierwsza cyfra kodu (na wszelki wypadek zachowujemy)
    norm_lp_1 = str(lp).strip()[0]
    norm_up_1 = str(up).strip()[0]

    # Inicjalizujemy informacje o dopasowaniu
    if lp_has_two_digits and up_has_two_digits:
        dopasowanie_hist = "2 cyfry"
        dopasowanie_gielda = "2 cyfry"
    elif not lp_has_two_digits and up_has_two_digits:
        dopasowanie_hist = "Załadunek: 1 cyfra, Rozładunek: 2 cyfry"
        dopasowanie_gielda = "Załadunek: 1 cyfra, Rozładunek: 2 cyfry"
    elif lp_has_two_digits and not up_has_two_digits:
        dopasowanie_hist = "Załadunek: 2 cyfry, Rozładunek: 1 cyfra"
        dopasowanie_gielda = "Załadunek: 2 cyfry, Rozładunek: 1 cyfra"
    else:
        dopasowanie_hist = "Obie lokalizacje: 1 cyfra"
        dopasowanie_gielda = "Obie lokalizacje: 1 cyfra"

    # Dla dwóch cyfr w obu kodach pocztowych
    if lp_has_two_digits and up_has_two_digits:
        exact_hist = historical_rates_df[
            (historical_rates_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_df['kod pocztowy zaladunku'].astype(str).str.zfill(2) == norm_lp) &
            (historical_rates_df['kod pocztowy rozladunku'].astype(str).str.zfill(2) == norm_up)
            ]

        exact_gielda = historical_rates_gielda_df[
            (historical_rates_gielda_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_gielda_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_gielda_df['kod pocztowy zaladunku'].astype(str).str.zfill(2) == norm_lp) &
            (historical_rates_gielda_df['kod pocztowy rozladunku'].astype(str).str.zfill(2) == norm_up)
            ]
    # Dla jednej cyfry w kodzie załadunku, dwóch cyfr w kodzie rozładunku
    elif not lp_has_two_digits and up_has_two_digits:
        exact_hist = historical_rates_df[
            (historical_rates_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_df['kod pocztowy zaladunku'].astype(str).str[0] == norm_lp) &
            (historical_rates_df['kod pocztowy rozladunku'].astype(str).str.zfill(2) == norm_up)
            ]

        exact_gielda = historical_rates_gielda_df[
            (historical_rates_gielda_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_gielda_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_gielda_df['kod pocztowy zaladunku'].astype(str).str[0] == norm_lp) &
            (historical_rates_gielda_df['kod pocztowy rozladunku'].astype(str).str.zfill(2) == norm_up)
            ]
    # Dla dwóch cyfr w kodzie załadunku, jednej cyfry w kodzie rozładunku
    elif lp_has_two_digits and not up_has_two_digits:
        exact_hist = historical_rates_df[
            (historical_rates_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_df['kod pocztowy zaladunku'].astype(str).str.zfill(2) == norm_lp) &
            (historical_rates_df['kod pocztowy rozladunku'].astype(str).str[0] == norm_up)
            ]

        exact_gielda = historical_rates_gielda_df[
            (historical_rates_gielda_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_gielda_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_gielda_df['kod pocztowy zaladunku'].astype(str).str.zfill(2) == norm_lp) &
            (historical_rates_gielda_df['kod pocztowy rozladunku'].astype(str).str[0] == norm_up)
            ]
    # Dla jednej cyfry w obu kodach
    else:
        exact_hist = historical_rates_df[
            (historical_rates_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_df['kod pocztowy zaladunku'].astype(str).str[0] == norm_lp) &
            (historical_rates_df['kod pocztowy rozladunku'].astype(str).str[0] == norm_up)
            ]

        exact_gielda = historical_rates_gielda_df[
            (historical_rates_gielda_df['kraj zaladunku'].apply(normalize_country) == norm_lc) &
            (historical_rates_gielda_df['kraj rozladunku'].apply(normalize_country) == norm_uc) &
            (historical_rates_gielda_df['kod pocztowy zaladunku'].astype(str).str[0] == norm_lp) &
            (historical_rates_gielda_df['kod pocztowy rozladunku'].astype(str).str[0] == norm_up)
            ]

    # Inicjalizacja wyników
    podlot_hist, z_hist = None, 0
    podlot_gielda, z_gielda = None, 0
    result = {
        'hist_stawka_3m': None,
        'hist_stawka_6m': None,
        'hist_stawka_12m': None,
        'hist_stawka_48m': None,
        'hist_fracht_3m': None,
        'gielda_stawka_3m': None,
        'gielda_stawka_6m': None,
        'gielda_stawka_12m': None,
        'gielda_stawka_48m': None,
        'gielda_fracht_3m': None,
        'podlot_historyczny': None,  # Zmieniona nazwa z 'dystans' na 'podlot_historyczny'
        'relacja': f"{norm_lc} {norm_lp} - {norm_uc} {norm_up}",
        'dopasowanie_hist': dopasowanie_hist,
        'dopasowanie_gielda': dopasowanie_gielda
    }

    # Przetwarzanie wyników historycznych (średnia ważona, jeśli jest wiele rekordów)
    if not exact_hist.empty:
        try:
            if len(exact_hist) > 1:
                # Filtrujemy tylko rekordy z niepustymi stawkami dla każdego okresu
                column_mappings = {
                    'hist_stawka_3m': 'stawka_3m',
                    'hist_stawka_6m': 'stawka_6m',
                    'hist_stawka_12m': 'stawka_12m',
                    'hist_fracht_3m': 'fracht_3m'
                }

                # Przetwarzanie każdej kolumny stawek oddzielnie
                used_records_count = {}

                for result_key, df_key in column_mappings.items():
                    if df_key in exact_hist.columns:
                        # Filtruj tylko rekordy z niepustymi wartościami
                        valid_records = exact_hist[~pd.isna(exact_hist[df_key])]

                        if not valid_records.empty:
                            # Zapisz informację o liczbie użytych rekordów
                            used_records_count[result_key] = len(valid_records)

                            # Oblicz sumę zleceń dla tych rekordów
                            total_zlecen = valid_records[
                                'Liczba zlecen'].sum() if 'Liczba zlecen' in valid_records.columns else len(
                                valid_records)

                            if total_zlecen > 0:
                                # Oblicz wagi na podstawie liczby zleceń
                                if 'Liczba zlecen' in valid_records.columns:
                                    weights = valid_records['Liczba zlecen'] / total_zlecen
                                else:
                                    weights = pd.Series([1 / len(valid_records)] * len(valid_records),
                                                        index=valid_records.index)

                                # Oblicz średnią ważoną
                                result[result_key] = (valid_records[df_key] * weights).sum()

                                # Aktualizuj liczbę zleceń (dla stawki 3m)
                                if result_key == 'hist_stawka_3m':
                                    z_hist = total_zlecen

                # Podlot (dawniej dystans) - użyj centralnej funkcji
                podlot_hist = calculate_podlot_from_data(exact_hist, "podlot historyczny")
                if podlot_hist is not None:
                    result['podlot_historyczny'] = podlot_hist
                    # Liczba rekordów dla podlotu
                    valid_dist_records = exact_hist.dropna(subset=['dystans'])
                    if not valid_dist_records.empty:
                        used_records_count['podlot_historyczny'] = len(valid_dist_records)

                # Informacja o dopasowaniu - zaktualizuj, jeśli część rekordów pominięto
                if any(count < len(exact_hist) for count in used_records_count.values()):
                    non_empty_counts = ", ".join([f"{key.split('_')[1]}: {count}/{len(exact_hist)}"
                                                  for key, count in used_records_count.items()])

                    # Zaktualizuj dopasowanie, zachowując informację o rodzaju dopasowania
                    if "średnia ważona" in dopasowanie_hist:
                        dopasowanie_hist = dopasowanie_hist.split(" (")[0] + f" (tylko niepuste: {non_empty_counts})"
                    else:
                        dopasowanie_hist += f" (tylko niepuste: {non_empty_counts})"
                else:
                    # Dodaj informację o średniej ważonej
                    dopasowanie_hist += f" (średnia ważona z {len(exact_hist)} dopasowań)"

                # Aktualizacja po modyfikacji
                result['dopasowanie_hist'] = dopasowanie_hist
            else:
                # Jeśli jest tylko jeden rekord, użyj go bezpośrednio
                row = exact_hist.iloc[0]
                
                # Oblicz podlot za pomocą centralnej funkcji
                podlot_hist = calculate_podlot_from_data(exact_hist, "podlot historyczny")
                
                try:
                    z_hist = float(row.get('Liczba zlecen', 0)) if not pd.isna(row.get('Liczba zlecen', 0)) else 0
                except:
                    z_hist = 0

                # Używaj tylko niepustych wartości z pojedynczego rekordu
                result.update({
                    'hist_stawka_3m': row.get('stawka_3m') if not pd.isna(row.get('stawka_3m')) else None,
                    'hist_stawka_6m': row.get('stawka_6m') if not pd.isna(row.get('stawka_6m')) else None,
                    'hist_stawka_12m': row.get('stawka_12m') if not pd.isna(row.get('stawka_12m')) else None,
                    'hist_fracht_3m': row.get('fracht_3m') if not pd.isna(row.get('fracht_3m')) else None,
                    'podlot_historyczny': podlot_hist
                })
        except Exception as e:
            print(f"Błąd przetwarzania danych historycznych: {e}")

    # Przetwarzanie wyników z giełdy (średnia ważona, jeśli jest wiele rekordów)
    if not exact_gielda.empty:
        try:
            if len(exact_gielda) > 1:
                # Filtrujemy tylko rekordy z niepustymi stawkami dla każdego okresu
                column_mappings = {
                    'gielda_stawka_3m': 'stawka_3m',
                    'gielda_stawka_6m': 'stawka_6m',
                    'gielda_stawka_12m': 'stawka_12m',
                    'gielda_fracht_3m': 'fracht_3m'
                }

                # Przetwarzanie każdej kolumny stawek oddzielnie
                used_records_count = {}

                for result_key, df_key in column_mappings.items():
                    if df_key in exact_gielda.columns:
                        # Filtruj tylko rekordy z niepustymi wartościami
                        valid_records = exact_gielda[~pd.isna(exact_gielda[df_key])]

                        if not valid_records.empty:
                            # Zapisz informację o liczbie użytych rekordów
                            used_records_count[result_key] = len(valid_records)

                            # Oblicz sumę zleceń dla tych rekordów
                            total_zlecen = valid_records[
                                'Liczba zlecen'].sum() if 'Liczba zlecen' in valid_records.columns else len(
                                valid_records)

                            if total_zlecen > 0:
                                # Oblicz wagi na podstawie liczby zleceń
                                if 'Liczba zlecen' in valid_records.columns:
                                    weights = valid_records['Liczba zlecen'] / total_zlecen
                                else:
                                    weights = pd.Series([1 / len(valid_records)] * len(valid_records),
                                                        index=valid_records.index)

                                # Oblicz średnią ważoną
                                result[result_key] = (valid_records[df_key] * weights).sum()

                                # Aktualizuj liczbę zleceń (dla stawki 3m)
                                if result_key == 'gielda_stawka_3m':
                                    z_gielda = total_zlecen

                # Podlot - tylko jeśli nie mamy jeszcze podlotu z danych historycznych
                if result['podlot_historyczny'] is None:
                    podlot_gielda = calculate_podlot_from_data(exact_gielda, "podlot giełda")
                    if podlot_gielda is not None:
                        result['podlot_historyczny'] = podlot_gielda
                        # Liczba rekordów dla podlotu
                        valid_dist_records = exact_gielda.dropna(subset=['dystans'])
                        if not valid_dist_records.empty:
                            used_records_count['podlot_historyczny'] = len(valid_dist_records)

                # Informacja o dopasowaniu - zaktualizuj, jeśli część rekordów pominięto
                if any(count < len(exact_gielda) for count in used_records_count.values()):
                    non_empty_counts = ", ".join([f"{key.split('_')[1]}: {count}/{len(exact_gielda)}"
                                                  for key, count in used_records_count.items()])

                    # Zaktualizuj dopasowanie, zachowując informację o rodzaju dopasowania
                    if "średnia ważona" in dopasowanie_gielda:
                        dopasowanie_gielda = dopasowanie_gielda.split(" (")[
                                                 0] + f" (tylko niepuste: {non_empty_counts})"
                    else:
                        dopasowanie_gielda += f" (tylko niepuste: {non_empty_counts})"
                else:
                    # Dodaj informację o średniej ważonej
                    dopasowanie_gielda += f" (średnia ważona z {len(exact_gielda)} dopasowań)"

                # Aktualizacja po modyfikacji
                result['dopasowanie_gielda'] = dopasowanie_gielda
            else:
                # Jeśli jest tylko jeden rekord, użyj go bezpośrednio
                row = exact_gielda.iloc[0]
                
                # Oblicz podlot za pomocą centralnej funkcji
                podlot_gielda = calculate_podlot_from_data(exact_gielda, "podlot giełda")
                
                try:
                    z_gielda = float(row.get('Liczba zlecen', 0)) if not pd.isna(row.get('Liczba zlecen', 0)) else 0
                except:
                    z_gielda = 0

                # Używaj tylko niepustych wartości z pojedynczego rekordu
                gielda_values = {
                    'gielda_stawka_3m': row.get('stawka_3m') if not pd.isna(row.get('stawka_3m')) else None,
                    'gielda_stawka_6m': row.get('stawka_6m') if not pd.isna(row.get('stawka_6m')) else None,
                    'gielda_stawka_12m': row.get('stawka_12m') if not pd.isna(row.get('stawka_12m')) else None,
                    'gielda_fracht_3m': row.get('fracht_3m') if not pd.isna(row.get('fracht_3m')) else None,
                }

                # Aktualizuj wyniki, ale nie nadpisuj podlotu jeśli już jest ustawiony
                result.update(gielda_values)
                if result['podlot_historyczny'] is None:
                    result['podlot_historyczny'] = podlot_gielda
        except Exception as e:
            print(f"Błąd przetwarzania danych z giełdy: {e}")

    # Obliczenie średniego podlotu ważonego liczbą zleceń
    weighted_podlot = calculate_weighted_podlot(podlot_hist, z_hist, podlot_gielda, z_gielda)
    if weighted_podlot is not None:
        result['podlot_sredni_wazony'] = weighted_podlot

    return result


def modify_process_przetargi(process_func):
    @wraps(process_func)
    def wrapper(*args, **kwargs):
        # Pobierz session_id z kwargs lub None dla legacy mode
        session_id = kwargs.get('session_id', None)
        
        # Pobierz user_data jeśli session_id istnieje
        if session_id:
            user_data = session_manager.get_session(session_id, create_if_missing=False)
            session_id_short = session_id[:8]
        else:
            # Legacy mode - użyj globalnych zmiennych
            user_data = None
            session_id_short = "LEGACY"
            global GEOCODING_CURRENT, GEOCODING_TOTAL
        
        df = args[0]
        unique_locations = set()
        for _, row in df.iterrows():
            city_load = row.get('miasto_zaladunku', '')
            if not isinstance(city_load, str):
                city_load = ''
            city_unload = row.get('miasto_rozladunku', '')
            if not isinstance(city_unload, str):
                city_unload = ''
            unique_locations.add((
                row['kraj_zaladunku'],
                str(row['kod_pocztowy_zaladunku']).strip(),
                city_load.strip()
            ))
            unique_locations.add((
                row['kraj_rozladunku'],
                str(row['kod_pocztowy_rozladunku']).strip(),
                city_unload.strip()
            ))
        
        logger.info(f"[{session_id_short}] Geokodowanie {len(unique_locations)} unikalnych lokalizacji")
        
        # Ustaw geocoding total
        if user_data:
            user_data.geocoding_total = len(unique_locations)
            user_data.geocoding_current = 0
        else:
            # Legacy mode
            with progress_lock:
                GEOCODING_TOTAL = len(unique_locations)
                GEOCODING_CURRENT = 0
        
        # Geokoduj każdą lokalizację
        for loc in unique_locations:
            logger.debug(f"[{session_id_short}] Geokodowanie: {loc}")
            get_coordinates(*loc)
            
            # Aktualizuj postęp
            if user_data:
                user_data.geocoding_current += 1
            else:
                # Legacy mode
                with progress_lock:
                    GEOCODING_CURRENT += 1
        
        # Upewnij się, że postęp wynosi 100%
        if user_data:
            user_data.geocoding_current = user_data.geocoding_total
        else:
            # Legacy mode
            with progress_lock:
                GEOCODING_CURRENT = GEOCODING_TOTAL
            
        # Sprawdź czy są niegeokodowane lokalizacje
        ungeocoded = get_ungeocoded_locations(df)
        if ungeocoded:
            logger.warning(f"[{session_id_short}] Znaleziono {len(ungeocoded)} nierozpoznanych lokalizacji")
            raise GeocodeException(ungeocoded)
        else:
            logger.info(f"[{session_id_short}] Wszystkie lokalizacje zgeokodowane pomyślnie")
        
        return process_func(*args, **kwargs)
    return wrapper


def sample_route_points(points, num_points=15):
    """
    Wybiera reprezentatywne punkty z trasy
    """
    if not points or len(points) <= num_points:
        return points
    
    # Wybierz punkty w równych odstępach
    step = len(points) // (num_points - 1)  # -1 bo chcemy zachować punkt końcowy
    sampled_points = points[::step]
    
    # Dodaj ostatni punkt jeśli nie został uwzględniony
    if points[-1] not in sampled_points:
        sampled_points = sampled_points[:num_points-1] + [points[-1]]
    
    return sampled_points[:num_points]

def decode_polyline(polyline_str):
    """
    Dekoduje polyline z API PTV do listy punktów [lat, lng]
    """
    try:
        if not polyline_str or not isinstance(polyline_str, str):
            return []
            
        # Sprawdź czy mamy do czynienia z formatem GeoJSON
        if 'coordinates' in polyline_str:
            try:
                # Znajdź początek i koniec koordynatów
                start = polyline_str.find('[[') + 2
                end = polyline_str.rfind(']]')
                if start == -1 or end == -1:
                    return []
                    
                # Wyciągnij tylko koordynaty
                coords_str = polyline_str[start:end]
                points = []
                
                # Podziel na pary współrzędnych
                coords = coords_str.split('],[')
                for coord in coords:
                    coord = coord.strip('[]')
                    if ',' in coord:
                        try:
                            lng, lat = map(float, coord.split(','))
                            points.append([lat, lng])
                        except (ValueError, TypeError):
                            continue
                
                return points
            except Exception:
                return []
        else:
            # Spróbuj standardowego dekodowania polyline
            try:
                return polyline.decode(polyline_str)
            except Exception:
                return []
            
    except Exception:
        return []

def create_google_maps_link(coord_from, coord_to, polyline_str):
    """
    Tworzy link do Google Maps z trasą używając punktów pośrednich
    """
    if not polyline_str:
        route_logger.info("Brak polyline - tworzę prosty link punkt-punkt")
        return f"https://www.google.com/maps/dir/{coord_from[0]},{coord_from[1]}/{coord_to[0]},{coord_to[1]}"
    
    try:
        # Dekoduj polyline do listy punktów
        route_points = decode_polyline(polyline_str)
        
                 # Wybierz 20 reprezentatywnych punktów
        sampled_points = sample_route_points(route_points, 15)
        
        if not sampled_points:
            route_logger.info("Nie udało się wybrać punktów pośrednich - tworzę prosty link punkt-punkt")
            return f"https://www.google.com/maps/dir/{coord_from[0]},{coord_from[1]}/{coord_to[0]},{coord_to[1]}"
        
        # Wyświetl punkty pośrednie
        print("\nPunkty pośrednie na trasie:")
        for i, point in enumerate(sampled_points):
            print(f"Punkt {i+1}: {point[0]}, {point[1]}")
        
        # Twórz link z punktami pośrednimi
        waypoints = "/".join(f"{lat},{lng}" for lat, lng in sampled_points[1:-1])
        link = f"https://www.google.com/maps/dir/{coord_from[0]},{coord_from[1]}/{waypoints}/{coord_to[0]},{coord_to[1]}"
        print(f"Wygenerowany link z {len(sampled_points)-2} punktami pośrednimi")
        return link
        
    except Exception as e:
        route_logger.warning(f"Błąd podczas generowania linku: {str(e)} - tworzę prosty link punkt-punkt")
        return f"https://www.google.com/maps/dir/{coord_from[0]},{coord_from[1]}/{coord_to[0]},{coord_to[1]}"

@modify_process_przetargi
def process_przetargi(df, fuel_cost=DEFAULT_FUEL_COST, driver_cost=DEFAULT_DRIVER_COST, session_id=None):
    """
    Główna funkcja przetwarzająca dane z pliku Excel.
    
    Args:
        df: DataFrame pandas z danymi do przetworzenia
        fuel_cost: Koszt paliwa EUR/km
        driver_cost: Koszt kierowcy EUR/dzień
        session_id: ID sesji użytkownika (None dla kompatybilności wstecznej)
    """
    # Pobierz dane sesji jeśli podano session_id
    if session_id:
        user_data = session_manager.get_session(session_id, create_if_missing=False)
        if user_data is None:
            logger.error(f"[{session_id[:8]}] Nie znaleziono sesji w process_przetargi!")
            return
        session_id_short = session_id[:8]
    else:
        # Tryb kompatybilności wstecznej - użyj globalnych zmiennych (deprecated)
        user_data = None
        session_id_short = "LEGACY"
    
    results = []

    # Inicjalizacja danych sesji
    if user_data:
        user_data.total_rows = len(df)
        user_data.progress = 0
        user_data.preview_data['total_count'] = user_data.total_rows
    else:
        # Legacy mode - użyj globalnych zmiennych
        global PROGRESS, TOTAL_ROWS, PREVIEW_DATA, CURRENT_ROW, GEOCODING_TOTAL, RESULT_EXCEL
        with progress_lock:
            TOTAL_ROWS = len(df)
            PROGRESS = 0
            PREVIEW_DATA = {
                'headers': [
                    'Kraj załadunku',
                    'Kod pocztowy załadunku',
                    'Kraj rozładunku',
                    'Kod pocztowy rozładunku',
                    'Dystans (km)',
                    'Podlot (km)',
                    'Koszt paliwa',
                    'Opłaty drogowe',
                    'Koszt kierowcy + leasing',
                    'Koszt podlotu (opłaty + paliwo)',
                    'Opłaty/km',
                    'Opłaty drogowe (szczegóły)',
                    'Suma kosztów',
                    'Link do mapy',
                    'Sugerowany fracht wg historycznych stawek',
                    'Sugerowany fracht wg matrixa',
                    'Oczekiwany zysk',
                    'Transit time (dni)'
                ],
                'rows': [],
                'total_count': TOTAL_ROWS
            }

        # Przypisanie nazw kolumn - oczekujemy 7 kolumn (6 standardowych + transit time)
    expected_columns = ['Kraj zaladunku', 'Kod zaladunku', 'Miasto zaladunku',
                       'Kraj rozladunku', 'Kod rozładunku', 'Miasto rozładunku', 'transit time']
    
    if len(df.columns) >= 7:
        # Przypisz nazwy do pierwszych 7 kolumn
        df.columns = expected_columns[:len(df.columns)]
        print(f"Przypisano nazwy do {len(df.columns)} kolumn, w tym kolumna 'transit time'")
    elif len(df.columns) == 6:
        # Stary format - tylko 6 kolumn bez transit time
        df.columns = expected_columns[:6]
        print(f"Stary format pliku - {len(df.columns)} kolumn bez 'transit time'")
    else:
        print(f"UWAGA: Nieoczekiwana liczba kolumn: {len(df.columns)}")
        df.columns = expected_columns[:len(df.columns)]
    
    print(f"Nazwy kolumn: {list(df.columns)}")

    for i, row in df.iterrows():
        try:
            # Aktualizacja postępu
            if user_data:
                user_data.current_row = i + 1
                user_data.progress = int((user_data.current_row / user_data.total_rows) * 100)
            else:
                # Legacy mode
                with progress_lock:
                    CURRENT_ROW = i + 1
                    PROGRESS = int((CURRENT_ROW / TOTAL_ROWS) * 100)
            
            lc = normalize_country(row["Kraj zaladunku"])
            lp = row["Kod zaladunku"]
            lc_city = row["Miasto zaladunku"]
            uc = normalize_country(row["Kraj rozladunku"])
            up = row["Kod rozładunku"]
            uc_city = row["Miasto rozładunku"]

            # Weryfikacja lokalizacji
            verify_load = verify_city_postal_code_match(lc, lp, lc_city)
            verify_unload = verify_city_postal_code_match(uc, up, uc_city)

            # ========== NOWE: Parsuj punkty pośrednie z Excel ==========
            waypoints = parse_waypoints_from_excel_row(row)
            
            # Decyzja: routing z waypoints czy bez
            if waypoints:
                logger.info(f"[{session_id_short}] Wiersz {i}: Wykryto {len(waypoints)} punktów pośrednich")
                
                # Buduj RouteRequest
                route_req = RouteRequest(
                    start=WaypointData(lc, lp, lc_city),
                    end=WaypointData(uc, up, uc_city),
                    waypoints=waypoints,
                    fuel_cost=fuel_cost,
                    driver_cost=driver_cost
                )
                
                # Użyj nowej funkcji dla tras z waypoints
                route_result_wp = calculate_multi_waypoint_route(route_req)
                
                if route_result_wp['success']:
                    dist_ptv = route_result_wp['distance']
                    polyline = route_result_wp['polyline']
                    road_toll = route_result_wp['road_toll']
                    other_toll = route_result_wp['other_toll']
                    toll_cost = road_toll + other_toll
                    toll_details = route_result_wp['toll_details']
                    special_systems = route_result_wp['special_systems']
                else:
                    # Geokodowanie się nie powiodło
                    logger.warning(f"[{session_id_short}] Wiersz {i}: Błąd waypoints - {route_result_wp.get('error_message')}")
                    dist_ptv = None
                    polyline = ''
                    road_toll = 0
                    other_toll = 0
                    toll_cost = 0
                    toll_details = {}
                    special_systems = []
                
                # Współrzędne dla mapy (start i koniec)
                coords_zl = route_req.start.coordinates if route_req.start.is_geocoded() else None
                coords_roz = route_req.end.coordinates if route_req.end.is_geocoded() else None
                
            else:
                # ========== STARY FLOW: bez waypoints (backward compatibility) ==========
                # Pobierz współrzędne
                coords_zl = get_coordinates(lc, lp, lc_city)
                coords_roz = get_coordinates(uc, up, uc_city)

                if coords_zl and coords_roz and None not in coords_zl[:2] and None not in coords_roz[:2]:
                    route_result = get_route_distance(coords_zl[:2], coords_roz[:2], 
                                              loading_country=lc, unloading_country=uc,
                                              avoid_switzerland=True, routing_mode=DEFAULT_ROUTING_MODE)
                    if isinstance(route_result, dict):
                        dist_ptv = route_result.get('distance')
                        polyline = route_result.get('polyline', '')
                        road_toll = route_result.get('road_toll', 0)  # Standardowe opłaty drogowe
                        other_toll = route_result.get('other_toll', 0)  # Opłaty za tunele/mosty/promy
                        toll_cost = road_toll + other_toll  # Całkowity koszt opłat
                        toll_details = route_result.get('toll_details', {})  # Szczegóły kosztów według krajów
                        special_systems = route_result.get('special_systems', [])  # Szczegóły systemów specjalnych
                    else:
                        dist_ptv = route_result
                        polyline = ''
                        road_toll = 0
                        other_toll = 0
                        toll_cost = 0
                        toll_details = {}
                        special_systems = []
                else:
                    dist_ptv = None
                    polyline = ''
                    road_toll = 0
                    other_toll = 0
                    toll_cost = 0
                    toll_details = {}
                    special_systems = []

            # Tworzenie linku do mapy
            if coords_zl and coords_roz and None not in coords_zl[:2] and None not in coords_roz[:2] and polyline:
                map_link = create_google_maps_link(coords_zl[:2], coords_roz[:2], polyline)
            else:
                if coords_zl and coords_roz and None not in coords_zl[:2] and None not in coords_roz[:2]:
                    # Jeśli mamy współrzędne, ale nie mamy polyline, stwórz prosty link
                    map_link = f"https://www.google.com/maps/dir/{coords_zl[0]},{coords_zl[1]}/{coords_roz[0]},{coords_roz[1]}"
                else:
                    map_link = None

            fuel_cost_value = dist_ptv * fuel_cost if dist_ptv is not None else None

            # Sprawdź czy w wierszu jest zdefiniowana wartość transit time
            transit_time_from_file = get_transit_time_from_row(row)
            
            if transit_time_from_file is not None:
                # Użyj wartości z pliku
                driver_days = transit_time_from_file
                print(f"Wiersz {i+1}: Użyto transit time z pliku: {driver_days} dni")
            else:
                # Oblicz standardowo na podstawie dystansu
                driver_days = calculate_driver_days(dist_ptv)
                if 'transit time' in row.index:
                    print(f"Wiersz {i+1}: Kolumna transit time pusta - obliczono z dystansu: {driver_days} dni")
                else:
                    print(f"Wiersz {i+1}: Brak kolumny transit time - obliczono z dystansu: {driver_days} dni")

            driver_cost_value = driver_days * driver_cost if driver_days is not None else None

            # Pobierz standardowe stawki
            rates = get_all_rates(lc, lp, uc, up, coords_zl, coords_roz)

            # Pobierz stawki bazujące na regionach
            region_rates = get_region_based_rates(lc, lp, uc, up)

            # Zmiana: podlot jest już bezpośrednio dostępny z historycznych danych z fallbackiem do regionów
            podlot, podlot_source = get_podlot(rates, region_rates)

            # Oblicz całkowity dystans (PTV + podlot)
            if dist_ptv is not None:
                total_distance = dist_ptv + podlot
            else:
                total_distance = None

            # Obliczenie całkowitego kosztu podlotu (opłaty drogowe + paliwo)
            # Stawka: 0.30 EUR/km (opłaty) + fuel_cost EUR/km (paliwo)
            oplaty_drogowe_podlot = calculate_podlot_toll(podlot, fuel_cost_per_km=fuel_cost)
            print(f"Koszt podlotu: {podlot} km × (0.30 + {fuel_cost}) EUR/km = {oplaty_drogowe_podlot} EUR")

            suma_kosztow = calculate_total_costs([road_toll, fuel_cost_value, driver_cost_value, oplaty_drogowe_podlot, other_toll])

            # Pobierz regiony dla obliczenia oczekiwanego zysku
            loading_region = get_region(lc, lp)
            unloading_region = get_region(uc, up)
            
            # Oblicz oczekiwany zysk na podstawie macierzy marży
            expected_profit, unit_margin, margin_source = calculate_expected_profit(
                loading_region, unloading_region, driver_days
            )
            
            # Oblicz sugerowany fracht wg matrixa: suma kosztów + oczekiwany zysk
            suggested_fracht_matrix = None
            if suma_kosztow is not None and expected_profit is not None and driver_days is not None:
                suggested_fracht_matrix = suma_kosztow + expected_profit

            lc_lat, lc_lon, lc_jakosc, lc_zrodlo = coords_zl if coords_zl else (None, None, 'nieznane', 'brak danych')
            uc_lat, uc_lon, uc_jakosc, uc_zrodlo = coords_roz if coords_roz else (None, None, 'nieznane', 'brak danych')
            lc_coords_str = format_coordinates(lc_lat, lc_lon)
            uc_coords_str = format_coordinates(uc_lat, uc_lon)

            dist_haversine = haversine(coords_zl[:2] if coords_zl else (None, None),
                                       coords_roz[:2] if coords_roz else (None, None))

            best_rates = get_best_rates(rates, region_rates)
            gielda_rate = best_rates['gielda_rate']
            hist_rate = best_rates['hist_rate']
            gielda_period = best_rates['gielda_period']
            hist_period = best_rates['hist_period']

            # Obliczenie frachtów na podstawie dystansu PTV i dystansu całkowitego
            gielda_fracht_km_ptv = calculate_fracht(dist_ptv, gielda_rate)
            gielda_fracht_km_total = calculate_fracht(total_distance, gielda_rate)
            klient_fracht_km_ptv = calculate_fracht(dist_ptv, hist_rate)
            klient_fracht_km_total = calculate_fracht(total_distance, hist_rate)

            gielda_fracht_3m = rates.get('gielda_fracht_3m')
            hist_fracht_3m = rates.get('hist_fracht_3m')

            # Pobierz współrzędne z verify_load i verify_unload
            city_load_coords = None
            postal_load_coords = None
            if verify_load.get('city_coords'):
                city_load_coords = f"{verify_load['city_coords'][0]}, {verify_load['city_coords'][1]}"
            if verify_load.get('postal_coords'):
                postal_load_coords = f"{verify_load['postal_coords'][0]}, {verify_load['postal_coords'][1]}"

            city_unload_coords = None
            postal_unload_coords = None
            if verify_unload.get('city_coords'):
                city_unload_coords = f"{verify_unload['city_coords'][0]}, {verify_unload['city_coords'][1]}"
            if verify_unload.get('postal_coords'):
                postal_unload_coords = f"{verify_unload['postal_coords'][0]}, {verify_unload['postal_coords'][1]}"

            # Dodaj informację o użyciu sugerowanych współrzędnych
            suggested_coords_info = ""
            if lc_city and not verify_load.get('is_match') and verify_load.get('suggested_coords'):
                suggested_coords_info += "Użyto sugerowanych współrzędnych dla załadunku. "
            if uc_city and not verify_unload.get('is_match') and verify_unload.get('suggested_coords'):
                suggested_coords_info += "Użyto sugerowanych współrzędnych dla rozładunku."

            # Przygotuj informacje o opłatach drogowych dla poszczególnych krajów
            # Przygotuj tekst z opisem opłat drogowych
            print(f"\nDEBUG special_systems dla wiersza {i+1}: {special_systems}")
            print(f"DEBUG other_toll dla wiersza {i+1}: {other_toll}")
            toll_text = format_toll_details(toll_details, road_toll, other_toll, special_systems)
            print(f"DEBUG toll_text dla wiersza {i+1}: {toll_text}")

            # Oblicz koszt opłat na kilometr (tylko standardowe opłaty drogowe)
            toll_per_km = calculate_toll_per_km(road_toll, dist_ptv)
            print(f"\nDEBUG - Wartości przed sumowaniem dla wiersza:")
            print(f"fuel_cost_value: {fuel_cost_value} ({type(fuel_cost_value).__name__})")
            print(f"driver_cost_value: {driver_cost_value} ({type(driver_cost_value).__name__})")
            print(f"road_toll: {road_toll} ({type(road_toll).__name__})")
            print(f"other_toll: {other_toll} ({type(other_toll).__name__})")
            print(f"oplaty_drogowe_podlot: {oplaty_drogowe_podlot} ({type(oplaty_drogowe_podlot)})")
    
            # Oblicz sumę kosztów (dodaj opłaty specjalne)
            print("\nDEBUG - Wartości kosztów:")
            print(f"fuel_cost_value: {fuel_cost_value} ({type(fuel_cost_value).__name__})")
            print(f"driver_cost_value: {driver_cost_value} ({type(driver_cost_value).__name__})")
            print(f"road_toll: {road_toll} ({type(road_toll).__name__})")
            print(f"other_toll: {other_toll} ({type(other_toll).__name__})")
            print(f"oplaty_drogowe_podlot: {oplaty_drogowe_podlot} ({type(oplaty_drogowe_podlot).__name__})")
            
            suma_kosztow = calculate_total_costs([fuel_cost_value, driver_cost_value, road_toll, other_toll, oplaty_drogowe_podlot])

            # Tworzenie preview_row dla udanego przetwarzania
            # Konwertuj wartości NaN na None przed utworzeniem słownika
            preview_row = {
                'Kraj załadunku': None if pd.isna(lc) else str(lc),
                'Kod pocztowy załadunku': None if pd.isna(lp) else str(lp),
                'Kraj rozładunku': None if pd.isna(uc) else str(uc),
                'Kod pocztowy rozładunku': None if pd.isna(up) else str(up),
                'Dystans (km)': dist_ptv,
                'Podlot (km)': podlot,
                'Koszt paliwa': fuel_cost_value,
                'Opłaty drogowe': road_toll,
                'Koszt kierowcy + leasing': driver_cost_value,
                'Koszt podlotu (opłaty + paliwo)': oplaty_drogowe_podlot,
                'Opłaty/km': toll_per_km,
                'Opłaty drogowe (szczegóły)': toll_text,
                'Opłaty specjalne': other_toll,
                'Suma kosztów': suma_kosztow,
                'Link do mapy': map_link if map_link else "-",
                'Sugerowany fracht wg historycznych stawek': klient_fracht_km_total if klient_fracht_km_total is not None else gielda_fracht_km_total,
                'Sugerowany fracht źródło': 'klient' if klient_fracht_km_total is not None else 'gielda',
                'Sugerowany fracht okres': hist_period if klient_fracht_km_total is not None else gielda_period,
                'Sugerowany fracht wg matrixa': suggested_fracht_matrix,
                'Region - Klient stawka 3m': region_rates.get('region_klient_stawka_3m'),
                'Region - Klient stawka 6m': region_rates.get('region_klient_stawka_6m'),
                'Region - Klient stawka 12m': region_rates.get('region_klient_stawka_12m'),
                'Region - Giełda stawka 3m': region_rates.get('region_gielda_stawka_3m'),
                'Region - Giełda stawka 6m': region_rates.get('region_gielda_stawka_6m'),
                'Region - Giełda stawka 12m': region_rates.get('region_gielda_stawka_12m'),
                'Region - Podlot (km)': region_rates.get('region_podlot'),
                'Podlot - źródło': podlot_source,
                'Oczekiwany zysk': expected_profit,
                'Transit time (dni)': driver_days
            }

            # Dodawanie do podglądu niezależnie od tego czy był błąd czy nie
            if user_data:
                user_data.preview_data['rows'].append(preview_row)
                if len(user_data.preview_data['rows']) > 1000:
                    user_data.preview_data['rows'].pop(0)
            else:
                # Legacy mode
                PREVIEW_DATA['rows'].append(preview_row)
                if len(PREVIEW_DATA['rows']) > 1000:
                    PREVIEW_DATA['rows'].pop(0)

            # Przygotuj wyniki z dodanymi kolumnami regionalnymi
            result_dict = {
                "Kraj zaladunku": None if pd.isna(lc) else str(lc),
                "Kod zaladunku": None if pd.isna(lp) else str(lp),
                "Miasto zaladunku": None if pd.isna(lc_city) else str(lc_city),
                "Region załadunku": get_region(lc, lp),
                "Współrzędne zaladunku": lc_coords_str,
                "Kraj rozladunku": uc,
                "Kod rozładunku": up,
                "Miasto rozładunku": uc_city,
                "Region rozładunku": get_region(uc, up),
                "Współrzędne rozładunku": uc_coords_str,
                "km PTV (tylko ładowne)": format_currency(dist_ptv),
                "km całkowite z podlotem": format_currency(total_distance),
                "podlot": format_currency(podlot),
                "Transit time (dni)": driver_days,
                # Zamień link na tekst "Mapa" już na etapie tworzenia DataFrame
                "Link do mapy": "Mapa" if map_link and isinstance(map_link, str) and map_link.startswith('http') else map_link,
                # Zapisz oryginalny link w ukrytej kolumnie
                "_original_map_link": map_link,
                "Sugerowany fracht wg matrixa": format_currency(suggested_fracht_matrix),
                "Klient sugerowany fracht/km z podlotem": format_currency(klient_fracht_km_total),
                "Giełda sugerowany fracht/km z podlotem": format_currency(gielda_fracht_km_total),
                "Koszt paliwa": format_currency(fuel_cost_value),
                "Koszt kierowcy + leasing": format_currency(driver_cost_value),
                "Opłaty drogowe": format_currency(road_toll),
                "Opłaty specjalne": format_currency(other_toll),
                "Koszt podlotu (opłaty + paliwo)": format_currency(oplaty_drogowe_podlot),
                "Opłaty drogowe/km": format_currency(toll_per_km),
                "Suma kosztów": format_currency(suma_kosztow),
                "Szczegóły opłat drogowych": toll_text,
                "Dopasowanie giełda": rates.get('gielda_dopasowanie'),
                "Giełda stawka 3m": format_currency(rates.get('gielda_stawka_3m')),
                "Giełda stawka 6m": format_currency(rates.get('gielda_stawka_6m')),
                "Giełda stawka 12m": format_currency(rates.get('gielda_stawka_12m')),
                "Giełda fracht 3m": format_currency(gielda_fracht_3m),
                "Giełda sugerowany fracht/km PTV": format_currency(gielda_fracht_km_ptv),
                "Dopasowanie klient": rates.get('hist_dopasowanie'),
                "Klient stawka 3m": format_currency(rates.get('hist_stawka_3m')),
                "Klient stawka 6m": format_currency(rates.get('hist_stawka_6m')),
                "Klient stawka 12m": format_currency(rates.get('hist_stawka_12m')),
                "Klient fracht 3m": format_currency(hist_fracht_3m),
                "Klient sugerowany fracht/km PTV": format_currency(klient_fracht_km_ptv),
                "Weryfikacja załadunku - miasto": verify_load.get('city_name', ''),
                "Weryfikacja załadunku - kod pocztowy": verify_load.get('postal_name', ''),
                "Weryfikacja załadunku - współrzędne miasta": city_load_coords,
                "Weryfikacja załadunku - współrzędne kodu": postal_load_coords,
                "Weryfikacja załadunku - odległość (km)": format_currency(verify_load.get('distance_km')),
                "Weryfikacja załadunku - poprawna": "TAK" if verify_load.get('is_match', True) else "NIE",
                "Weryfikacja rozładunku - miasto": verify_unload.get('city_name', ''),
                "Weryfikacja rozładunku - kod pocztowy": verify_unload.get('postal_name', ''),
                "Weryfikacja rozładunku - współrzędne miasta": city_unload_coords,
                "Weryfikacja rozładunku - współrzędne kodu": postal_unload_coords,
                "Weryfikacja rozładunku - odległość (km)": format_currency(verify_unload.get('distance_km')),
                "Weryfikacja rozładunku - poprawna": "TAK" if verify_unload.get('is_match', True) else "NIE",
                "Uwagi do geokodowania": suggested_coords_info,
                "Region - Dopasowanie giełda": region_rates.get('region_gielda_dopasowanie'),
                "Region - Giełda stawka 3m": format_currency(region_rates.get('region_gielda_stawka_3m')),
                "Region - Giełda stawka 6m": format_currency(region_rates.get('region_gielda_stawka_6m')),
                "Region - Giełda stawka 12m": format_currency(region_rates.get('region_gielda_stawka_12m')),
                "Region - Dopasowanie klient": region_rates.get('region_klient_dopasowanie'),
                "Region - Klient stawka 3m": format_currency(region_rates.get('region_klient_stawka_3m')),
                "Region - Klient stawka 6m": format_currency(region_rates.get('region_klient_stawka_6m')),
                "Region - Klient stawka 12m": format_currency(region_rates.get('region_klient_stawka_12m')),
                "Region - Podlot (km)": format_currency(region_rates.get('region_podlot')),
                "Podlot - źródło": podlot_source,
                "Oczekiwany zysk": format_currency(expected_profit),
                "Jakość geokodowania (zał.)": lc_jakosc,
                "Źródło geokodowania (zał.)": lc_zrodlo,
                "Jakość geokodowania (rozł.)": uc_jakosc,
                "Źródło geokodowania (rozł.)": uc_zrodlo,
                "km w linii prostej": format_currency(dist_haversine),
                "Źródło marży": margin_source
            }

            results.append(result_dict)

        except Exception as e:
            current_row_num = user_data.current_row if user_data else (CURRENT_ROW if 'CURRENT_ROW' in globals() else i+1)
            print(f"\n❌ BŁĄD w wierszu {current_row_num}: {str(e)}")
            print(f"Szczegóły wiersza: {dict(row)}")
            print("Traceback:")
            import traceback
            traceback.print_exc()
            
            # Tworzenie basic_result dla pliku Excel
            basic_result = {
               "Kraj zaladunku": lc,
               "Kod zaladunku": lp,
               "Miasto zaladunku": lc_city,
               "Kraj rozladunku": uc,
               "Kod rozładunku": up,
               "Miasto rozładunku": uc_city,
               "Błąd przetwarzania": str(e)
            }
            results.append(basic_result)
            
            # Dodanie wiersza do podglądu z informacją o błędzie
            preview_row = {
                'Kraj załadunku': lc,
                'Kod pocztowy załadunku': lp,
                'Kraj rozładunku': uc,
                'Kod pocztowy rozładunku': up,
                'Dystans (km)': None,
                'Podlot (km)': None,
                'Koszt paliwa': None,
                'Opłaty drogowe': None,
                'Koszt kierowcy + leasing': None,
                'Koszt podlotu (opłaty + paliwo)': None,
                'Opłaty/km': None,
                'Opłaty drogowe (szczegóły)': None,
                'Opłaty specjalne': None,
                'Suma kosztów': None,
                'Link do mapy': "-",
                'Sugerowany fracht wg historycznych stawek': None,
                'Sugerowany fracht wg matrixa': None,
                'Region - Klient stawka 3m': None,
                'Region - Klient stawka 6m': None,
                'Region - Klient stawka 12m': None,
                'Region - Giełda stawka 3m': None,
                'Region - Giełda stawka 6m': None,
                'Region - Giełda stawka 12m': None,
                'Oczekiwany zysk': None,
                'Transit time (dni)': None
            }
            
            # Dodawanie do podglądu niezależnie od tego czy był błąd czy nie
            if user_data:
                user_data.preview_data['rows'].append(preview_row)
                if len(user_data.preview_data['rows']) > 1000:
                    user_data.preview_data['rows'].pop(0)
            else:
                # Legacy mode
                PREVIEW_DATA['rows'].append(preview_row)
                if len(PREVIEW_DATA['rows']) > 1000:
                    PREVIEW_DATA['rows'].pop(0)

        finally:
            pass

    # Log końcowy przetwarzania
    if user_data:
        logger.info(f"[{session_id_short}] Przetworzono {user_data.current_row} z {user_data.total_rows} wierszy")
    else:
        print(f"\nPrzetworzono {CURRENT_ROW} z {TOTAL_ROWS} wierszy")
    
    print("\nGenerowanie pliku Excel...")
    try:
        # Tworzenie DataFrame z wyników
        result_df = pd.DataFrame(results)
        
        # Usuń ukrytą kolumnę przed zapisaniem do Excela
        if '_original_map_link' in result_df.columns:
            # Zapisz wartości do zmiennej tymczasowej, aby można było ich użyć później
            original_map_links = result_df['_original_map_link'].copy()
            # Usuń kolumnę z DataFrame
            result_df = result_df.drop(columns=['_original_map_link'])
        else:
            original_map_links = None
        
        # Zapisz do bufora w pamięci
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Wycena')
            
            # Pobierz worksheet do formatowania
            worksheet = writer.sheets['Wycena']
            
            # Funkcja do konwersji numeru kolumny na nazwę kolumny Excel
            def get_column_letter(n):
                string = ""
                while n > 0:
                    n, remainder = divmod(n - 1, 26)
                    string = chr(65 + remainder) + string
                return string

            # Definicje kolorów dla różnych typów kolumn
            COLORS = {
                'podstawowe': 'E6E6E6',  # Szary
                'geokodowanie': 'FFE699',  # Jasny żółty
                'dystans': 'BDD7EE',  # Jasny niebieski
                'gielda': 'C6E0B4',  # Jasny zielony
                'klient': 'F8CBAD',  # Jasny pomarańczowy
                'koszty': 'D9D9D9',  # Jaśniejszy szary
                'region': 'E2EFDA',  # Bardzo jasny zielony
                'weryfikacja': 'FCE4D6',  # Bardzo jasny pomarańczowy
                'marza': 'FFD700'  # Złoty dla kolumn z marżą
            }

            # Mapowanie kolumn do typów (indeksowane od 1)
            COLUMN_TYPES = {
                'podstawowe': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],  # A, B, C, G, H, I
                'geokodowanie': [80],  # D, E, F, J, K, L
                'dystans': [11, 12, 13, 14, 60, 67],  # M, N, O, P, Q
                'gielda': [18, 27, 28, 29, 30, 31, 32, 52, 53, 54, 55],  # R-X
                'klient': [16, 17, 33, 34, 35, 36, 37, 38, 56, 57, 58, 59],  # Y-AE
                'koszty': [19, 20, 21, 22, 23, 24, 25, 26],  # AF-AL
                'region': [53, 54, 55],  # AW-BE
                'weryfikacja': [39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51],  # AM-AV
                'marza': [62]  # Nowe kolumny z marżą: Oczekiwany zysk, Marża jednostkowa, Dni kierowcy, Źródło marży
            }
            
            # Zamień długie linki do map na krótki tekst "Mapa" przed ustawieniem szerokości kolumn
            if "Link do mapy" in result_df.columns:
                # Dodajemy więcej informacji diagnostycznych
                print(f"Zamieniam linki na tekst 'Mapa'. Liczba linków do zamiany: {result_df['Link do mapy'].str.startswith('http', na=False).sum()}")
                # Nie musimy już tworzyć tymczasowej kolumny, bo mamy już _original_map_link
                # Upewnij się, że wszystkie linki są zamienione na "Mapa"
                result_df.loc[result_df['Link do mapy'].str.startswith('http', na=False), 'Link do mapy'] = "Mapa"
                print(f"Po zamianie, liczba komórek z tekstem 'Mapa': {(result_df['Link do mapy'] == 'Mapa').sum()}")
                
            # Ustaw szerokość kolumn i formatowanie
            for idx, col in enumerate(result_df.columns, start=1):
                column_letter = get_column_letter(idx)
                
                # Ustaw szerokość
                max_length = max(
                    result_df.iloc[:, idx-1].astype(str).apply(len).max(),
                    len(str(col))
                )
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

                # Ustaw kolor tła dla całej kolumny
                for col_type, columns in COLUMN_TYPES.items():
                    if idx in columns:
                        for cell in worksheet[column_letter]:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=COLORS[col_type],
                                end_color=COLORS[col_type],
                                fill_type='solid'
                            )
            
            # Formatowanie nagłówków
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Lista kolumn z różnymi formatami
            distance_columns = [11, 12, 13, 14, 60, 67]  # M, N, O, P, AQ, AU - dystans i odległości weryfikacji
            percentage_columns = [51]  # R, Y, AY - dopasowanie
            currency_columns = [16,	17,	18,	19,	20,	21, 22,	23,	24,	25,
                                28,	29,	30,	31,	32,34, 35, 36,	37,	38,
                                53,	54,	55,57,	58,	59,62]  # Nowe kolumny: Sugerowany fracht wg matrixa i Oczekiwany zysk
            
            # Formatowanie komórek z danymi
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    
                    # Formatowanie liczb
                    if isinstance(cell.value, (int, float)):
                        if cell.column in distance_columns:  # Kolumny z kilometrami
                            cell.number_format = '#,##0'
                        elif cell.column in percentage_columns:  # Kolumny z procentami
                            cell.number_format = '0.0%'
                        elif cell.column in currency_columns:  # Kolumny z walutą
                            cell.number_format = '#,##0.00 €'
                        else:  # Pozostałe kolumny liczbowe
                            cell.number_format = '#,##0'
                    
                    # Dodaj hiperłącza do komórek z tekstem "Mapa" lub linkami
                    if cell.column == 15:  # Kolumna z linkami do map (Q)
                        column_name = list(result_df.columns)[cell.column-1]  # Sprawdź nazwę kolumny
                        if column_name == "Link do mapy":
                            # Sprawdź czy komórka zawiera link lub tekst "Mapa"
                            if cell.value == "Mapa" and original_map_links is not None:
                                # Pobierz oryginalny link z zapisanej zmiennej
                                original_link = original_map_links.iloc[row_idx] if row_idx < len(original_map_links) else None
                                if isinstance(original_link, str) and original_link.startswith('http'):
                                    # Dodaj hiperłącze
                                    cell.hyperlink = original_link
                                    # Ustaw styl hiperłącza
                                    cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                            # Dodatkowe sprawdzenie dla linków, które nie zostały zamienione
                            elif isinstance(cell.value, str) and cell.value.startswith('http'):
                                # Zapisz oryginalny link
                                link_url = cell.value
                                # Zamień wartość komórki na krótki tekst
                                cell.value = "Mapa"
                                # Dodaj hiperłącze
                                cell.hyperlink = link_url
                                # Ustaw styl hiperłącza
                                cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                                print(f"Zamieniono link na 'Mapa' w komórce Excel (wiersz {row_idx+1}, kolumna {cell.column})")
            
            # Dodaj obramowanie do wszystkich komórek
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Zamrożenie pierwszego wiersza
            worksheet.freeze_panes = 'A2'
        
        # Pobierz zawartość bufora
        excel_buffer.seek(0)
        excel_data = excel_buffer.getvalue()
        
        # Aktualizuj dane sesji lub zmienną globalną
        if user_data:
            logger.info(f"[{session_id_short}] Zapisuję wynik, rozmiar: {len(excel_data)} bajtów")
            user_data.result_excel = io.BytesIO(excel_data)
        else:
            # Legacy mode
            with progress_lock:
                print(f"[process_przetargi] Ustawiam RESULT_EXCEL, rozmiar danych: {len(excel_data)} bajtów")
                RESULT_EXCEL = excel_data
        
        logger.info(f"[{session_id_short}] Plik Excel został wygenerowany pomyślnie")
        
    except Exception as e:
        logger.error(f"[{session_id_short}] Błąd podczas generowania pliku Excel: {e}", exc_info=True)
        if user_data:
            user_data.result_excel = None
            user_data.progress = -1
        else:
            # Legacy mode
            with progress_lock:
                print("[process_przetargi] Błąd, ustawiam RESULT_EXCEL=None")
                RESULT_EXCEL = None
                PROGRESS = -1

    return results


def save_caches():
    try:
        geo_dict = {key: geo_cache[key] for key in geo_cache}
        route_dict = {key: route_cache[key] for key in route_cache}
        joblib.dump(geo_dict, 'geo_cache_backup.joblib')
        joblib.dump(route_dict, 'route_cache_backup.joblib')
        print("Zapisano pamięć podręczną.")
    except Exception as e:
        print(f"Błąd zapisywania pamięci podręcznej: {e}")


def load_caches():
    try:
        if os.path.exists('geo_cache_backup.joblib'):
            geo_cache_data = joblib.load('geo_cache_backup.joblib')
            for k, v in geo_cache_data.items():
                geo_cache[k] = v
            print(f"Wczytano {len(geo_cache_data)} elementów geo_cache.")
        if os.path.exists('route_cache_backup.joblib'):
            route_cache_data = joblib.load('route_cache_backup.joblib')
            for k, v in route_cache_data.items():
                route_cache[k] = v
            print(f"Wczytano {len(route_cache_data)} elementów route_cache.")
    except Exception as e:
        print(f"Błąd wczytywania pamięci podręcznej: {e}")

def clear_luxembourg_cache():
    """Czyści błędne wpisy cache dla Luksemburga"""
    keys_to_remove = []
    
    # Znajdź wszystkie klucze związane z Luksemburgiem
    for key in geo_cache:
        if 'LU_' in str(key) or 'Luxembourg' in str(key) or 'Luksemburg' in str(key):
            keys_to_remove.append(key)
    
    # Usuń znalezione klucze
    for key in keys_to_remove:
        print(f"Usuwam błędny cache dla klucza: {key} -> {geo_cache[key]}")
        del geo_cache[key]
    
    print(f"Wyczyszczono {len(keys_to_remove)} wpisów cache dla Luksemburga")
    return len(keys_to_remove)


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max-limit

# Konfiguracja sesji Flask dla wielu użytkowników
# W produkcji SECRET_KEY powinien pochodzić ze zmiennych środowiskowych
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_COOKIE_SECURE'] = False  # Ustaw True w produkcji z HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

# Konfiguracja logowania
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)

# Dodanie filtra postępu do loggera
class FilterProgress(logging.Filter):
    def filter(self, record):
        return not record.getMessage().startswith("Progress:")


# === DUPLIKACJA ZMIENNYCH GLOBALNYCH USUNIĘTA ===
# Wszystkie dane sesji znajdują się teraz w SessionManager
# === KONIEC ===


# ============================================================================
# FUNKCJE POMOCNICZE DO ZARZĄDZANIA SESJAMI UŻYTKOWNIKÓW
# ============================================================================

def get_or_create_session_id() -> str:
    """
    Pobiera lub tworzy unikalny identyfikator sesji dla użytkownika.
    Używa Flask session do przechowania session_id.
    
    Returns:
        Unikalny identyfikator sesji
    """
    if 'session_id' not in session:
        session['session_id'] = session_manager.generate_session_id()
        session.permanent = True  # Sesja przetrwa zamknięcie przeglądarki
        logger.info(f"Utworzono nową sesję użytkownika: {session['session_id'][:8]}...")
    return session['session_id']


def get_user_session() -> UserSessionData:
    """
    Pobiera dane sesji dla aktualnego użytkownika.
    
    Returns:
        Obiekt UserSessionData dla aktualnego użytkownika
    """
    session_id = get_or_create_session_id()
    return session_manager.get_session(session_id)


# Inicjalizacja i uruchomienie schedulera czyszczenia sesji
cleanup_scheduler = SessionCleanupScheduler(session_manager, interval_hours=1)
cleanup_scheduler.start()

# Zatrzymaj scheduler przy zamykaniu aplikacji
atexit.register(lambda: cleanup_scheduler.stop())

logger.info("System zarządzania sesjami zainicjalizowany pomyślnie")


# ============================================================================
# ENDPOINTY FLASK
# ============================================================================

@app.route("/show_cache")
def show_cache():
    geo_cache_info = {k: geo_cache[k] for k in geo_cache}
    route_cache_info = {k: str(route_cache[k])[:100] + "..." if len(str(route_cache[k])) > 100 else route_cache[k] for k in route_cache}
    locations_cache_info = {k: f"Locations data ({len(locations_cache[k].get('locations', []))} ungeocoded, {len(locations_cache[k].get('correct_locations', []))} geocoded)" if locations_cache[k] else "None" for k in locations_cache}
    
    return jsonify({
        'geo_cache': geo_cache_info,
        'route_cache': route_cache_info,
        'locations_cache': locations_cache_info,
        'cache_sizes': {
            'geo_cache': len(geo_cache_info),
            'route_cache': len(route_cache_info),
            'locations_cache': len(locations_cache_info)
        }
    })


@app.route("/", methods=["GET", "POST"])
def upload_file():
    """
    Endpoint do uploadu pliku Excel i rozpoczęcia przetwarzania.
    Każdy użytkownik ma swoją izolowaną sesję.
    """
    if request.method == "POST":
        logger.info("="*80)
        logger.info("POST / - Otrzymano request uploadu pliku")
        logger.info(f"Request files: {list(request.files.keys())}")
        logger.info(f"Request form: {dict(request.form)}")
        logger.info("="*80)
        
        try:
            file = request.files.get("file")
            if not file:
                logger.error("Brak pliku w requeście!")
                return render_template("error.html", message="Nie wybrano pliku")

            # Pobierz lub utwórz sesję użytkownika
            user_data = get_user_session()
            
            # Reset danych użytkownika dla nowego przetwarzania
            user_data.reset_progress()
            
            # Pobierz parametry z formularza
            user_data.fuel_cost = float(request.form.get("fuel_cost", DEFAULT_FUEL_COST))
            user_data.driver_cost = float(request.form.get("driver_cost", DEFAULT_DRIVER_COST))
            user_data.matrix_type = request.form.get("matrix_type", "klient")

            # Ustaw odpowiednią macierz marży
            set_margin_matrix(user_data.matrix_type)

            # Czytaj plik w kontekście żądania
            user_data.file_bytes = file.read()
            
            logger.info(f"[{user_data.session_id[:8]}] Rozpoczynam przetwarzanie (fuel={user_data.fuel_cost}, driver={user_data.driver_cost}, matrix={user_data.matrix_type})")

            # Uruchom przetwarzanie w tle z session_id
            thread = threading.Thread(
                target=background_processing,
                args=(user_data.session_id,),
                daemon=True,
                name=f"BgProcess-{user_data.session_id[:8]}"
            )
            thread.start()
            user_data.thread = thread
            
            return render_template("processing.html")
            
        except Exception as e:
            logger.error(f"Błąd podczas uploadu pliku: {e}", exc_info=True)
            return render_template("error.html", message=str(e))
            
    return render_template("upload.html", 
                         default_fuel_cost=DEFAULT_FUEL_COST,
                         default_driver_cost=DEFAULT_DRIVER_COST)


@app.route("/download")
def download():
    """
    Endpoint do pobierania wyników przetwarzania.
    Każdy użytkownik pobiera tylko swoje wyniki.
    """
    try:
        user_data = get_user_session()
        session_id_short = user_data.session_id[:8]
        
        logger.info(
            f"[{session_id_short}] /download - "
            f"complete={user_data.processing_complete}, "
            f"has_result={user_data.result_excel is not None}"
        )
        
        if not user_data.processing_complete:
            logger.warning(f"[{session_id_short}] Próba pobrania - przetwarzanie w toku")
            return "Przetwarzanie jeszcze nie zostało zakończone.", 400
        
        if user_data.result_excel is None:
            logger.warning(f"[{session_id_short}] Próba pobrania - brak wyników")
            return "Brak wyników do pobrania.", 404
        
        # Sprawdź czy plik nie jest pusty
        user_data.result_excel.seek(0, 2)  # Przejdź na koniec
        file_size = user_data.result_excel.tell()
        user_data.result_excel.seek(0)  # Wróć na początek
        
        if file_size <= 100:
            logger.error(f"[{session_id_short}] Plik wynikowy jest zbyt mały: {file_size} bajtów")
            return "Plik Excel jest nieprawidłowy.", 500
        
        logger.info(f"[{session_id_short}] Wysyłam plik wynikowy ({file_size} bajtów)")
        
        return send_file(
            user_data.result_excel,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'zlecenia_{session_id_short}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Błąd podczas pobierania pliku: {e}", exc_info=True)
        return "Błąd podczas pobierania pliku.", 500


@app.route("/progress")
def progress():
    """
    Endpoint zwracający postęp przetwarzania dla aktualnego użytkownika.
    Każdy użytkownik widzi tylko swój postęp.
    """
    try:
        user_data = get_user_session()
        
        geocoding_progress = 0
        if user_data.geocoding_total > 0:
            geocoding_progress = min(
                int((user_data.geocoding_current / user_data.geocoding_total) * 100),
                100
            )
        
        # Dodaj informację o używanej macierzy
        matrix_name, matrix_file = get_margin_matrix_info()
        
        response_data = {
            'progress': user_data.progress,
            'current': user_data.current_row,
            'total': user_data.total_rows,
            'geocoding_progress': geocoding_progress,
            'error': user_data.progress == -1 or user_data.progress == -2,
            'preview_data': user_data.preview_data,
            'processing_complete': user_data.processing_complete,
            'matrix_name': matrix_name,
            'matrix_file': matrix_file,
            'session_id': user_data.session_id[:8]  # Dla debugowania
        }
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Błąd w /progress: {e}", exc_info=True)
        return jsonify({
            'error': str(e),
            'progress': -1,
            'processing_complete': False
        }), 500


@app.route("/admin/sessions")
def admin_sessions():
    """
    Endpoint administracyjny do monitorowania aktywnych sesji.
    Pokazuje statystyki i listę wszystkich aktywnych sesji użytkowników.
    """
    try:
        stats = session_manager.get_session_statistics()
        sessions_info = session_manager.get_all_sessions_info()
        
        return jsonify({
            'statistics': stats,
            'active_sessions': sessions_info,
            'timestamp': time.time()
        })
    except Exception as e:
        logger.error(f"Błąd w /admin/sessions: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route("/admin/cleanup_sessions")
def admin_cleanup_sessions():
    """
    Endpoint administracyjny do wymuszenia natychmiastowego czyszczenia sesji.
    """
    try:
        deleted_count = cleanup_scheduler.cleanup_now()
        stats = session_manager.get_session_statistics()
        
        return jsonify({
            'deleted_sessions': deleted_count,
            'remaining_sessions': stats['total_sessions'],
            'message': f'Usunięto {deleted_count} wygasłych sesji'
        })
    except Exception as e:
        logger.error(f"Błąd w /admin/cleanup_sessions: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route("/save_cache")
def save_cache_endpoint():
    save_caches()
    return "Zapisano pamięć podręczną."

@app.route("/clear_luxembourg_cache")
def clear_luxembourg_cache_endpoint():
    """Endpoint do czyszczenia błędnego cache dla Luksemburga"""
    count = clear_luxembourg_cache()
    return f"Wyczyszczono {count} błędnych wpisów cache dla Luksemburga."

@app.route("/clear_locations_cache")
def clear_locations_cache_endpoint():
    """Endpoint do czyszczenia cache lokalizacji"""
    try:
        locations_cache.clear()
        # Usuń także klucz z sesji
        session.pop('locations_cache_key', None)
        return "Wyczyszczono cache lokalizacji."
    except Exception as e:
        return f"Błąd podczas czyszczenia cache lokalizacji: {e}"

@app.route("/ptv_stats")
def ptv_stats():
    stats = ptv_manager.get_stats()
    return jsonify(stats)


@app.route("/upload_for_geocoding")
def upload_for_geocoding():
    return render_template("upload_for_geocoding.html")

def background_geocoding_processing(file_bytes, cache_key):
    """Funkcja do przetwarzania geokodowania w tle"""
    global GEOCODING_CURRENT, GEOCODING_TOTAL
    
    try:
        # Wczytaj dane z Excela, zachowując kody pocztowe jako tekst
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            dtype={
                'kod pocztowy zaladunku': str,
                'kod pocztowy rozladunku': str,
                'kod_pocztowy_zaladunku': str,
                'kod_pocztowy_rozladunku': str
            }
        )
        print("Kolumny w pliku Excel:", df.columns.tolist())
        
        locations_data = get_all_locations_status(df)
        print("Znalezione lokalizacje:")
        print(f"Poprawne: {len(locations_data['correct_locations'])}")
        print(f"Do weryfikacji: {len(locations_data['locations'])}")
        
        # Zapisz dane w cache z TTL 1 godzina (3600 sekund)
        locations_cache.set(cache_key, locations_data, expire=3600)
        print(f"Zapisano dane w cache z kluczem: {cache_key}")
        
    except Exception as e:
        print(f"Błąd podczas przetwarzania pliku: {str(e)}")
        # Zapisz błąd w cache
        locations_cache.set(cache_key, {'error': str(e)}, expire=3600)

@app.route("/ungeocoded_locations", methods=["GET", "POST"])
def ungeocoded_locations():
    if request.method == "POST":
        try:
            file = request.files.get("file")
            if not file:
                return render_template("error.html", message="Nie wybrano pliku")

            # Wygeneruj unikalny klucz dla danych
            data_timestamp = str(time.time())
            file_bytes = file.read()
            data_size = str(len(file_bytes))
            cache_key = hashlib.md5(f"{data_timestamp}_{data_size}".encode()).hexdigest()
            
            # Zapisz klucz w sesji
            session['locations_cache_key'] = cache_key
            session['processing_status'] = 'running'
            
            # Uruchom przetwarzanie w tle
            thread = threading.Thread(target=background_geocoding_processing, args=(file_bytes, cache_key))
            thread.daemon = True
            thread.start()
            
            # Zwróć stronę z paskiem postępu
            return render_template("processing.html", processing_type="geocoding")
            
        except Exception as e:
            print(f"Błąd podczas przetwarzania pliku: {str(e)}")
            return render_template("error.html", message=f"Błąd podczas przetwarzania pliku: {str(e)}")
    
    # Dla GET, spróbuj pobrać dane z cache za pomocą klucza z sesji
    cache_key = session.get('locations_cache_key')
    locations_data = None
    
    if cache_key:
        try:
            locations_data = locations_cache.get(cache_key)
            if locations_data:
                if 'error' in locations_data:
                    return render_template("error.html", message=f"Błąd podczas przetwarzania: {locations_data['error']}")
                    
                print(f"Pobrano dane z cache z kluczem: {cache_key}")
                session['processing_status'] = 'completed'
            else:
                print(f"Brak danych w cache dla klucza: {cache_key}")
                # Usuń nieważny klucz z sesji
                session.pop('locations_cache_key', None)
        except Exception as e:
            print(f"Błąd pobierania z cache: {e}")
            session.pop('locations_cache_key', None)
    
    return render_template("ungeocoded_locations.html", locations_data=locations_data)


@app.route("/save_manual_coordinates", methods=['POST'])
def save_manual_coordinates():
    data = request.json
    try:
        key = data.get('key')
        lat = data.get('latitude')
        lon = data.get('longitude')
        if not key or lat is None or lon is None:
            return jsonify({'success': False, 'message': 'Nieprawidłowe dane'})
        geo_cache[key] = (lat, lon, 'ręczne', 'użytkownik')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route("/test_route_form", methods=['GET', 'POST'])
def test_route_form():
    if request.method == 'POST':
        load_country = request.form.get('load_country', '').strip().upper()
        load_postal = request.form.get('load_postal', '').strip()
        load_city = request.form.get('load_city', '').strip() or None
        
        unload_country = request.form.get('unload_country', '').strip().upper()
        unload_postal = request.form.get('unload_postal', '').strip()
        unload_city = request.form.get('unload_city', '').strip() or None
        
        matrix_type = request.form.get('matrix_type', 'klient')
        
        # Pobierz opcjonalny transit time
        transit_time_str = request.form.get('transit_time', '').strip()
        transit_time = None
        if transit_time_str:
            try:
                transit_time = float(transit_time_str)
                if not (0.1 <= transit_time <= 30):
                    transit_time = None
            except (ValueError, TypeError):
                transit_time = None

        # Pobierz koszty paliwa i kierowcy
        fuel_cost_str = request.form.get('fuel_cost', '').strip()
        driver_cost_str = request.form.get('driver_cost', '').strip()
        
        try:
            fuel_cost = float(fuel_cost_str) if fuel_cost_str else DEFAULT_FUEL_COST
            if not (0 <= fuel_cost <= 5):
                fuel_cost = DEFAULT_FUEL_COST
        except (ValueError, TypeError):
            fuel_cost = DEFAULT_FUEL_COST
            
        try:
            driver_cost = float(driver_cost_str) if driver_cost_str else DEFAULT_DRIVER_COST
            if not (0 <= driver_cost <= 1000):
                driver_cost = DEFAULT_DRIVER_COST
        except (ValueError, TypeError):
            driver_cost = DEFAULT_DRIVER_COST

        # Ustaw odpowiednią macierz marży
        set_margin_matrix(matrix_type)
        
        # ========== NOWE: Parsowanie punktów pośrednich ==========
        waypoints = parse_waypoints_from_form(request.form)
        
        # Decyzja: czy używać nowego flow z waypoints?
        if waypoints:
            logger.info(f"Wykryto {len(waypoints)} punktów pośrednich - używam nowego flow")
            
            try:
                # Budowanie żądania trasy
                route_req = RouteRequest(
                    start=WaypointData(load_country, load_postal, load_city),
                    end=WaypointData(unload_country, unload_postal, unload_city),
                    waypoints=waypoints,
                    fuel_cost=fuel_cost,
                    driver_cost=driver_cost,
                    matrix_type=matrix_type
                )
                
                # Oblicz trasę z waypoints
                result = calculate_multi_waypoint_route(route_req)
                
                if not result['success']:
                    error_msg = result.get('error_message', 'Unknown error')
                    return render_template("error.html", message=error_msg)
                
                # Dodatkowe obliczenia (koszt paliwa, kierowcy, marża)
                fuel_cost_total = result['distance'] * fuel_cost
                
                # Transit time
                if transit_time is not None:
                    driver_days = transit_time
                else:
                    driver_days = calculate_driver_days(result['distance'])
                
                driver_cost_total = driver_days * driver_cost if driver_days else 0
                
                # Podlot (tylko dla pierwszego segmentu)
                podlot_cost = calculate_podlot(load_country, load_postal)
                
                total_cost = (
                    fuel_cost_total + 
                    result['toll_cost'] + 
                    driver_cost_total + 
                    podlot_cost
                )
                
                # Renderuj specjalny template dla waypoints
                return render_template(
                    "test_route_form.html",  # użyjemy tego samego template z danymi
                    waypoints_result=True,
                    route_request=route_req,
                    route_result=result,
                    fuel_cost_total=fuel_cost_total,
                    driver_cost_total=driver_cost_total,
                    driver_days=driver_days,
                    podlot_cost=podlot_cost,
                    total_cost=total_cost,
                    matrix_name=matrix_type.title()
                )
                
            except ValueError as e:
                return render_template("error.html", message=f"Błąd walidacji: {str(e)}")
            except Exception as e:
                logger.error(f"Błąd obliczania trasy z waypoints: {e}", exc_info=True)
                return render_template("error.html", message=f"Nieoczekiwany błąd: {str(e)}")
        
        else:
            # ========== STARY FLOW: bez waypoints (backward compatibility) ==========
            try:
                # Pobierz współrzędne dla punktu załadunku
                load_coords = get_coordinates(load_country, load_postal, load_city)
                if not load_coords:
                    return render_template("error.html", message=f"Nie można znaleźć współrzędnych dla: {load_country} {load_postal}")
                
                # Pobierz współrzędne dla punktu rozładunku
                unload_coords = get_coordinates(unload_country, unload_postal, unload_city)
                if not unload_coords:
                    return render_template("error.html", message=f"Nie można znaleźć współrzędnych dla: {unload_country} {unload_postal}")

                # Oblicz trasę
                load_coord_str = f"{load_coords[0]},{load_coords[1]}"
                unload_coord_str = f"{unload_coords[0]},{unload_coords[1]}"
                
                # Przekieruj do wyników z parametrami
                return redirect(url_for('test_route_result', 
                                      load=[load_country, load_postal, load_coord_str],
                                      unload=[unload_country, unload_postal, unload_coord_str],
                                      matrix_type=matrix_type,
                                      transit_time=transit_time,
                                      fuel_cost=fuel_cost,
                                      driver_cost=driver_cost))

            except Exception as e:
                return render_template("error.html", message=str(e))

    return render_template("test_route_form.html")

@app.route("/test_route_result")
def test_route_result():
    # Ustaw macierz na podstawie parametru (jeśli przekazany)
    matrix_type = request.args.get('matrix_type', 'klient')
    set_margin_matrix(matrix_type)
    matrix_name, _ = get_margin_matrix_info()
    
    # Pobierz opcjonalny transit time
    transit_time_str = request.args.get('transit_time')
    transit_time = None
    if transit_time_str and transit_time_str != 'None':
        try:
            transit_time = float(transit_time_str)
        except (ValueError, TypeError):
            transit_time = None
    
    # Pobierz koszty paliwa i kierowcy
    fuel_cost_str = request.args.get('fuel_cost')
    driver_cost_str = request.args.get('driver_cost')
    
    try:
        fuel_cost = float(fuel_cost_str) if fuel_cost_str else DEFAULT_FUEL_COST
    except (ValueError, TypeError):
        fuel_cost = DEFAULT_FUEL_COST
        
    try:
        driver_cost = float(driver_cost_str) if driver_cost_str else DEFAULT_DRIVER_COST
    except (ValueError, TypeError):
        driver_cost = DEFAULT_DRIVER_COST
    
    return render_template("test_route_result.html", 
                         load=request.args.getlist('load'),
                         unload=request.args.getlist('unload'),
                         matrix_name=matrix_name,
                         transit_time=transit_time,
                         fuel_cost=fuel_cost,
                         driver_cost=driver_cost)

@app.route("/test_truck_route")
def test_truck_route():
    coord_from = request.args.get('coord_from')
    coord_to = request.args.get('coord_to')
    
    # Pobierz kody pocztowe i kraje z parametrów URL
    load_country = request.args.get('load_country')
    load_postal = request.args.get('load_postal')
    unload_country = request.args.get('unload_country')
    unload_postal = request.args.get('unload_postal')
    
    # Pobierz opcjonalny transit time
    transit_time_str = request.args.get('transit_time')
    transit_time = None
    if transit_time_str and transit_time_str != 'None':
        try:
            transit_time = float(transit_time_str)
        except (ValueError, TypeError):
            transit_time = None
    
    # Pobierz koszty paliwa i kierowcy z parametrów URL
    fuel_cost_str = request.args.get('fuel_cost')
    driver_cost_str = request.args.get('driver_cost')
    
    try:
        fuel_cost = float(fuel_cost_str) if fuel_cost_str else DEFAULT_FUEL_COST
    except (ValueError, TypeError):
        fuel_cost = DEFAULT_FUEL_COST
        
    try:
        driver_cost = float(driver_cost_str) if driver_cost_str else DEFAULT_DRIVER_COST
    except (ValueError, TypeError):
        driver_cost = DEFAULT_DRIVER_COST
    
    if not coord_from or not coord_to:
        return jsonify({'error': 'Brak współrzędnych'})
    
    try:
        coord_from = tuple(map(float, coord_from.split(',')))
        coord_to = tuple(map(float, coord_to.split(',')))
        
        result = get_route_distance(coord_from, coord_to, 
                                loading_country=load_country, unloading_country=unload_country)
        if not result:
            return jsonify({'error': 'Nie udało się wyznaczyć trasy'})
        
        dist = result.get('distance', 0)
        polyline = result.get('polyline', '')
        road_toll = result.get('road_toll', 0)  # Standardowe opłaty drogowe
        other_toll = result.get('other_toll', 0)  # Opłaty za tunele/mosty/promy
        total_toll = road_toll + other_toll
        toll_details = result.get('toll_details', {})
        special_systems = result.get('special_systems', [])
        
        # Obliczanie kosztu paliwa (używa przekazanej wartości)
        fuel_cost_value = dist * fuel_cost if dist is not None else None
        
        # Użyj transit_time jeśli został przekazany, w przeciwnym razie oblicz na podstawie dystansu
        if transit_time is not None:
            driver_days = transit_time
        elif dist is not None:
            if dist <= 350:
                driver_days = 1
            elif 351 <= dist <= 500:
                driver_days = 1.25
            elif 501 <= dist <= 700:
                driver_days = 1.5
            elif 701 <= dist <= 1100:
                driver_days = 2
            elif 1101 <= dist <= 1700:
                driver_days = 3
            elif 1701 <= dist <= 2300:
                driver_days = 4
            elif 2301 <= dist <= 2900:
                driver_days = 5
            elif 2901 <= dist <= 3500:
                driver_days = 6
            else:
                driver_days = None
        else:
            driver_days = None
            
        driver_cost_value = driver_days * driver_cost if driver_days is not None else None
        
        # Przygotuj tekst z opisem opłat drogowych
        toll_text = format_toll_details(toll_details, road_toll, other_toll, special_systems)
        
        # Oblicz koszt opłat na kilometr (tylko standardowe opłaty)
        toll_per_km = calculate_toll_per_km(road_toll, dist)
        
        # Przygotuj listę krajów
        countries = list(toll_details.keys()) if toll_details else ["Brak informacji o krajach"]
            
        # Stwórz link do mapy
        map_link = create_google_maps_link(coord_from, coord_to, polyline)
        
        # Pobierz stawki regionalne najpierw
        # Jeśli mamy przekazane kody pocztowe i kraje, użyj ich
        if load_country and load_postal and unload_country and unload_postal:
            # Użyj przekazanych danych
            region_rates = get_region_based_rates(load_country, load_postal, unload_country, unload_postal)
        else:
            # Fallback - użyj krajów z trasy i przykładowych kodów pocztowych
            load_country_fallback = "Poland"
            load_postal_fallback = "00-001"  # Przykładowy kod pocztowy
            unload_country_fallback = "Germany"
            unload_postal_fallback = "10115"  # Przykładowy kod pocztowy
            region_rates = get_region_based_rates(load_country_fallback, load_postal_fallback, unload_country_fallback, unload_postal_fallback)

        # Pobierz standardowe stawki i sprawdź czy jest historyczny podlot
        rates = get_all_rates(load_country, load_postal, unload_country, unload_postal, coord_from, coord_to)
        
        # Pobierz podlot z danych historycznych z fallbackiem do regionów lub użyj wartości domyślnej  
        podlot, podlot_source = get_podlot(rates, region_rates)
        
        oplaty_drogowe_podlot = calculate_podlot_toll(podlot, fuel_cost_per_km=fuel_cost)
        print(f"Koszt podlotu: {podlot} km × (0.30 + {fuel_cost}) EUR/km = {oplaty_drogowe_podlot} EUR")
        
        # Oblicz sumę kosztów
        suma_kosztow = calculate_total_costs([road_toll, fuel_cost_value, driver_cost_value, other_toll, oplaty_drogowe_podlot])
        
        # Wykorzystaj istniejącą logikę z process_przetargi - stwórz pojedynczy wiersz danych
        test_row = {
            'Kraj załadunku': load_country,
            'Kod pocztowy załadunku': load_postal,
            'Kraj rozładunku': unload_country,
            'Kod pocztowy rozładunku': unload_postal,
            'Dystans PTV (km)': dist,
            'Podlot (km)': podlot,
            'Opłaty drogowe (€)': road_toll,
            'Koszt paliwa (€)': fuel_cost_value,
            'Koszt kierowcy (€)': driver_cost_value,
            'Opłaty drogowe podlot (€)': oplaty_drogowe_podlot,
            'Inne opłaty (€)': other_toll,
            'Suma kosztów (€)': suma_kosztow
        }
        
        # Użyj tej samej logiki co w process_przetargi dla pojedynczego wiersza
        processed_data = []
        
        # Pobierz regiony (z normalizacją krajów)
        loading_region = get_region(normalize_country(test_row['Kraj załadunku']), test_row['Kod pocztowy załadunku'])
        unloading_region = get_region(normalize_country(test_row['Kraj rozładunku']), test_row['Kod pocztowy rozładunku'])
        
        # Oblicz dni kierowcy - użyj przekazanego transit_time lub oblicz na podstawie dystansu
        distance = test_row.get('Dystans PTV (km)', 0) or 0
        if transit_time is not None:
            driver_days_for_profit = transit_time
        else:
            driver_days_for_profit = calculate_driver_days(distance)
        
        # Oblicz oczekiwany zysk
        expected_profit, unit_margin, margin_source = calculate_expected_profit(
            loading_region, unloading_region, driver_days_for_profit
        )
        
        # Oblicz sugerowane frachty używając tej samej logiki co w process_przetargi
        best_rates = get_best_rates(rates, region_rates)
        hist_rate = best_rates['hist_rate']
        gielda_rate = best_rates['gielda_rate']
        
        total_distance = distance + (test_row.get('Podlot (km)', 0) or 0)
        
        # Sugerowany fracht wg historycznych stawek
        if hist_rate is not None and total_distance > 0:
            suggested_fracht_historical = total_distance * hist_rate
        elif gielda_rate is not None and total_distance > 0:
            suggested_fracht_historical = total_distance * gielda_rate
        else:
            suggested_fracht_historical = None
            
        # Sugerowany fracht wg matrixa
        if suma_kosztow is not None and expected_profit is not None:
            suggested_fracht_matrix = suma_kosztow + expected_profit
        else:
            suggested_fracht_matrix = None
        
        # Funkcja pomocnicza do konwersji None na 0
        def safe_value(value):
            return value if value is not None else 0
        
        # Przygotuj odpowiedź JSON
        response = {
            'distance': round(dist, 2),
            'polyline': polyline,
            'podlot': safe_value(podlot),  # Dodajemy wartość podlotu
            'podlot_source': podlot_source,  # Dodajemy źródło podlotu
            'oplaty_drogowe': safe_value(format_currency(road_toll)),
            'oplaty_dodatkowe': safe_value(format_currency(other_toll)),
            'oplaty_na_km': safe_value(format_currency(toll_per_km)),
            'koszt_paliwa': safe_value(format_currency(fuel_cost_value)),
            'koszt_kierowcy': safe_value(format_currency(driver_cost_value)),
            'oplaty_drogowe_podlot': safe_value(format_currency(oplaty_drogowe_podlot)),
            'suma_kosztow': safe_value(format_currency(suma_kosztow)),
            'sugerowany_fracht_historyczny': safe_value(format_currency(suggested_fracht_historical)),
            'sugerowany_fracht_matrix': safe_value(format_currency(suggested_fracht_matrix)),
            'oczekiwany_zysk': safe_value(format_currency(expected_profit)),
            'marga_jednostkowa': safe_value(format_currency(unit_margin)),
            'transit_time_dni': safe_value(driver_days),
            'loading_region': loading_region or 'Nieznany',
            'unloading_region': unloading_region or 'Nieznany',
            'kraje': countries,
            'szczegoly_oplat': toll_text,
            'map_link': map_link,
            'region_klient_stawka_3m': safe_value(format_currency(region_rates.get('region_klient_stawka_3m'))),
            'region_klient_stawka_6m': safe_value(format_currency(region_rates.get('region_klient_stawka_6m'))),
            'region_klient_stawka_12m': safe_value(format_currency(region_rates.get('region_klient_stawka_12m'))),
            'region_gielda_stawka_3m': safe_value(format_currency(region_rates.get('region_gielda_stawka_3m'))),
            'region_gielda_stawka_6m': safe_value(format_currency(region_rates.get('region_gielda_stawka_6m'))),
            'region_gielda_stawka_12m': safe_value(format_currency(region_rates.get('region_gielda_stawka_12m'))),
            'region_relacja': region_rates.get('region_relacja', ''),
            'region_dopasowanie': region_rates.get('region_dopasowanie', '')
        }
        
        return jsonify(response)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route("/test_truck_route_form")
def test_truck_route_form():
    return render_template("test_truck_route.html")

@app.route("/test_truck_route_map")
def test_truck_route_map():
    """
    Endpoint zwracający link do mapy Google z trasą
    """
    from_coords = request.args.get('from')
    to_coords = request.args.get('to')
    
    if not from_coords or not to_coords:
        return "Brak wymaganych parametrów", 400
    
    try:
        # Konwertuj współrzędne na listy [lat, lon]
        from_coords = [float(x) for x in from_coords.split(',')]
        to_coords = [float(x) for x in to_coords.split(',')]
        
        # Pobierz trasę z PTV
        # Pobierz kraje z parametrów URL
        load_country = request.args.get('load_country')
        unload_country = request.args.get('unload_country')
        
        route_result = get_route_distance(from_coords, to_coords, 
                                        loading_country=load_country, unloading_country=unload_country,
                                        avoid_switzerland=True, routing_mode=DEFAULT_ROUTING_MODE)
        
        # Wyciągnij polyline z wyniku
        polyline = route_result.get('polyline', '') if isinstance(route_result, dict) else ''
        
        # Stwórz link do mapy
        map_link = create_google_maps_link(from_coords, to_coords, polyline)
        
        # Przekieruj do Google Maps
        return redirect(map_link)
        
    except Exception as e:
        route_logger.error(f"Błąd podczas generowania linku do mapy: {str(e)}")
        return f"Błąd: {str(e)}", 500

def background_processing(session_id: str):
    """
    Przetwarza plik Excel w tle dla danej sesji użytkownika.
    
    Args:
        session_id: Unikalny identyfikator sesji użytkownika
    """
    user_data = session_manager.get_session(session_id, create_if_missing=False)
    
    if user_data is None:
        print(f"[{session_id[:8]}] BŁĄD: Nie znaleziono sesji użytkownika!")
        logger.error(f"[{session_id[:8]}] Nie znaleziono sesji użytkownika!")
        return
    
    session_id_short = session_id[:8]
    
    try:
        logger.info(f"[{session_id_short}] Rozpoczynam przetwarzanie pliku...")
        
        # Użyj bezpośrednio bajtów pliku z sesji użytkownika
        file_stream = io.BytesIO(user_data.file_bytes)
        df = pd.read_excel(file_stream, dtype=str)
        df.columns = df.columns.str.lower().str.replace(" ", "_").str.strip()
        logger.info(f"[{session_id_short}] Przekształcone kolumny: {list(df.columns)}")
        
        try:
            # Wywołaj process_przetargi z session_id
            process_przetargi(df, user_data.fuel_cost, user_data.driver_cost, session_id)
            save_caches()
            
            logger.info(f"[{session_id_short}] Przetwarzanie zakończone pomyślnie")
            user_data.processing_complete = True
            if user_data.progress not in [-1, -2]:  # Nie zmieniaj progress jeśli wystąpił błąd
                user_data.progress = 100
        
        except GeocodeException as ge:
            logger.warning(
                f"[{session_id_short}] GeocodeException: "
                f"Znaleziono {len(ge.ungeocoded_locations)} nierozpoznanych lokalizacji"
            )
            user_data.progress = -2
            user_data.processing_complete = True
            user_data.locations_to_verify = ge.ungeocoded_locations
        
        except Exception as e:
            logger.error(f"[{session_id_short}] Błąd w process_przetargi: {e}", exc_info=True)
            user_data.progress = -1
            user_data.processing_complete = True
            user_data.result_excel = None
            
    except Exception as e:
        logger.error(f"[{session_id_short}] Krytyczny błąd przetwarzania: {e}", exc_info=True)
        user_data.progress = -1
        user_data.processing_complete = True
        user_data.result_excel = None

@app.route("/check_locations", methods=["POST"])
def check_locations():
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "Nie wybrano pliku"})

        # Wczytaj plik Excel
        df = pd.read_excel(file)
        
        # Zbierz unikalne lokalizacje
        unique_locations = set()
        for _, row in df.iterrows():
            city_load = row.get('miasto zaladunku', '')
            if not isinstance(city_load, str):
                city_load = ''
            city_unload = row.get('miasto rozladunku', '')
            if not isinstance(city_unload, str):
                city_unload = ''
            unique_locations.add((
                row['kraj zaladunku'],
                str(row['kod pocztowy zaladunku']).strip(),
                city_load.strip()
            ))
            unique_locations.add((
                row['kraj rozladunku'],
                str(row['kod pocztowy rozladunku']).strip(),
                city_unload.strip()
            ))

        # Sprawdź każdą lokalizację
        correct_locations = []
        incorrect_locations = []
        
        for country, postal_code, city in unique_locations:
            coords = get_coordinates(country, postal_code, city)
            if coords and coords[0] is not None and coords[1] is not None:
                correct_locations.append({
                    'country': country,
                    'postal_code': postal_code,
                    'city': city
                })
            else:
                incorrect_locations.append({
                    'country': country,
                    'postal_code': postal_code,
                    'city': city
                })

        return jsonify({
            'total_locations': len(unique_locations),
            'correct_locations': len(correct_locations),
            'incorrect_locations': len(incorrect_locations),
            'incorrect_details': incorrect_locations
        })

    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/update_coordinates", methods=["POST"])
def update_coordinates():
    try:
        data = request.get_json()
        location_id = data.get('location_id')
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        
        if not all([location_id, latitude, longitude]):
            return jsonify({'success': False, 'error': 'Brak wymaganych danych'})
            
        # Pobierz dane z cache
        cache_key = session.get('locations_cache_key')
        if not cache_key:
            return jsonify({'success': False, 'error': 'Brak danych lokalizacji w sesji'})
        
        locations_data = locations_cache.get(cache_key)
        if not locations_data:
            return jsonify({'success': False, 'error': 'Nie znaleziono danych lokalizacji w cache'})
        
        ungeocoded = locations_data.get('locations', [])
        geocoded = locations_data.get('correct_locations', [])
        
        # Szukaj lokalizacji do aktualizacji
        location = None
        for loc in ungeocoded:
            if str(loc['id']) == str(location_id):
                location = loc
                break
                
        if not location:
            return jsonify({'success': False, 'error': 'Nie znaleziono lokalizacji'})
            
        # Aktualizuj współrzędne w geo_cache
        key = location['key']
        coords = (float(latitude), float(longitude))
        geo_cache[key] = (*coords, 'manual', 'manual verification')
        
        # Przenieś lokalizację z ungeocoded do geocoded
        location['coords'] = f"{latitude},{longitude}"
        ungeocoded.remove(location)
        geocoded.append(location)
        
        # Aktualizuj dane w cache
        locations_data['locations'] = ungeocoded
        locations_data['correct_locations'] = geocoded
        locations_cache.set(cache_key, locations_data, expire=3600)
        
        return jsonify({'success': True})
        
    except Exception as e:
        print(f"Błąd podczas aktualizacji współrzędnych: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

def is_route_to_or_from_switzerland(loading_country, unloading_country):
    """
    Sprawdza czy trasa jest do lub ze Szwajcarii na podstawie kodów krajów.
    """
    normalized_loading = normalize_country(loading_country)
    normalized_unloading = normalize_country(unloading_country)
    return normalized_loading == 'Switzerland' or normalized_unloading == 'Switzerland'

def get_route_distance(coord_from, coord_to, loading_country=None, unloading_country=None, avoid_switzerland=True, routing_mode=DEFAULT_ROUTING_MODE):
    # Jeśli trasa jest do/ze Szwajcarii, nie unikamy Szwajcarii
    if loading_country and unloading_country:
        if is_route_to_or_from_switzerland(loading_country, unloading_country):
            avoid_switzerland = False
    
    result = ptv_manager.get_route_distance(coord_from, coord_to, avoid_switzerland, routing_mode)
    return result

def calculate_podlot_from_data(df, description="podlot"):
    """
    Centralna funkcja do obliczania podlotu (dystansu) z DataFrame
    
    Args:
        df: DataFrame z kolumną 'dystans' i opcjonalnie 'Liczba zlecen'
        description: opis dla logowania (np. "regionalny podlot", "podlot historyczny")
    
    Returns:
        float or None: obliczony podlot lub None jeśli brak danych
    """
    if df.empty or 'dystans' not in df.columns:
        return None
    
    # Filtruj tylko rekordy z niepustymi wartościami dystansu
    valid_records = df.dropna(subset=['dystans'])
    
    if valid_records.empty:
        return None
    
    try:
        if len(valid_records) == 1:
            # Pojedynczy rekord - użyj bezpośrednio
            return float(valid_records.iloc[0]['dystans'])
        else:
            # Wiele rekordów - oblicz średnią ważoną
            if 'Liczba zlecen' in valid_records.columns:
                # Wagi na podstawie liczby zleceń
                total_orders = valid_records['Liczba zlecen'].sum()
                if total_orders > 0:
                    weights = valid_records['Liczba zlecen'] / total_orders
                    return (valid_records['dystans'] * weights).sum()
            
            # Fallback - równe wagi dla wszystkich rekordów
            weights = pd.Series(1 / len(valid_records), index=valid_records.index)
            return (valid_records['dystans'] * weights).sum()
    
    except Exception as e:
        print(f"Błąd obliczania {description}: {e}")
        return None

def calculate_weighted_podlot(podlot_hist, z_hist, podlot_gielda, z_gielda):
    """
    Oblicza średni podlot ważony liczbą zleceń z różnych źródeł
    
    Args:
        podlot_hist: podlot z danych historycznych
        z_hist: liczba zleceń historycznych
        podlot_gielda: podlot z danych giełdy
        z_gielda: liczba zleceń z giełdy
    
    Returns:
        float or None: średni ważony podlot lub None jeśli brak danych
    """
    if (z_hist + z_gielda) <= 0:
        return None
    
    hist_value = podlot_hist if podlot_hist is not None else 0
    gielda_value = podlot_gielda if podlot_gielda is not None else 0
    
    # Tylko jeśli mamy jakieś faktyczne wartości podlotu
    if hist_value > 0 or gielda_value > 0:
        return (hist_value * z_hist + gielda_value * z_gielda) / (z_hist + z_gielda)
    
    return None

@app.route("/geocoding_progress")
def geocoding_progress():
    """Endpoint do śledzenia postępu geokodowania"""
    global GEOCODING_CURRENT, GEOCODING_TOTAL
    
    if GEOCODING_TOTAL > 0:
        progress = int((GEOCODING_CURRENT / GEOCODING_TOTAL) * 100)
        return jsonify({
            'progress': progress,
            'current': GEOCODING_CURRENT,
            'total': GEOCODING_TOTAL,
            'status': 'running' if GEOCODING_CURRENT < GEOCODING_TOTAL else 'completed'
        })
    else:
        return jsonify({
            'progress': 0,
            'current': 0,
            'total': 0,
            'status': 'idle'
        })

if __name__ == '__main__':
    load_caches()
    load_margin_matrix()  # Wczytaj domyślną macierz marży (Matrix.xlsx)
    log = logging.getLogger('werkzeug')


    class FilterProgress(logging.Filter):
        def filter(self, record):
            return "/progress" not in record.getMessage()


    log.addFilter(FilterProgress())
    # Wyłączamy debug mode aby uniknąć auto-reload który przerywa przetwarzanie w tle
    app.run(debug=False, host='0.0.0.0')